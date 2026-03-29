
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import matplotlib as mpl
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.ticker import PercentFormatter
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from PyQt6.QtCore import QDate, Qt
from PyQt6.QtGui import QBrush, QColor, QFont, QPalette
from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDateEdit,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)

APP_TITLE = "实验室生产数据统计看板 (PyQt6)"

# 源表固定字段位置（A=1, AI=35）
COL_COUPLING_A = 0
COL_WBP_C = 2
COL_PRODUCT_D = 3
COL_RESEARCHER_E = 4
COL_START_AF = 31
COL_END_AG = 32
COL_SCORE_AI = 34
HEADER_ROW_1 = 2
HEADER_ROW_2 = 3
HEADER_NEW_EMPLOYEE = "是否新人"
HEADER_REQUEST_ACTIVITY = "requestactivity"
UNKNOWN_STAGE = "Unknown"
NON_OTHER_PLACEHOLDER = "请选择FCT偶联类型"
NON_OTHER_REQUIRED_MESSAGE = "请先在筛选器中选择一个FCT偶联类型。"
NO_WIP_MESSAGE = "无在研负载"
MAPPING_DIALOG_TITLE = "手动匹配列"

FIELD_COUPLING_TYPE = "coupling_type"
FIELD_WBP_CODE = "wbp_code"
FIELD_PRODUCT_ID = "product_id"
FIELD_RESEARCHER = "researcher"
FIELD_START_DATE = "start_date"
FIELD_END_DATE = "end_date"
FIELD_SCORE = "score"
FIELD_NEW_EMPLOYEE = "new_employee"
FIELD_REQUEST_ACTIVITY = "request_activity"

REQUIRED_FIELDS = (
    FIELD_COUPLING_TYPE,
    FIELD_WBP_CODE,
    FIELD_PRODUCT_ID,
    FIELD_RESEARCHER,
    FIELD_START_DATE,
    FIELD_END_DATE,
    FIELD_SCORE,
    FIELD_NEW_EMPLOYEE,
    FIELD_REQUEST_ACTIVITY,
)

FIELD_LABELS = {
    FIELD_COUPLING_TYPE: "偶联类型(A)",
    FIELD_WBP_CODE: "WBP Code(C)",
    FIELD_PRODUCT_ID: "Product ID(D)",
    FIELD_RESEARCHER: "实验员(E)",
    FIELD_START_DATE: "起始日期(AF)",
    FIELD_END_DATE: "完成日期(AG)",
    FIELD_SCORE: "积分(AI)",
    FIELD_NEW_EMPLOYEE: "是否新人",
    FIELD_REQUEST_ACTIVITY: "Request activity",
}

DEFAULT_FIXED_COLUMN_MAPPING = {
    FIELD_COUPLING_TYPE: COL_COUPLING_A,
    FIELD_WBP_CODE: COL_WBP_C,
    FIELD_PRODUCT_ID: COL_PRODUCT_D,
    FIELD_RESEARCHER: COL_RESEARCHER_E,
    FIELD_START_DATE: COL_START_AF,
    FIELD_END_DATE: COL_END_AG,
    FIELD_SCORE: COL_SCORE_AI,
}


@dataclass
class FilterState:
    start_date: date
    end_date: date
    researcher: str
    wbp_code: str
    product_id: str
    saturation_threshold: int
    non_other_coupling_type: str


class ColumnMappingCancelledError(Exception):
    pass


class ColumnMappingDialog(QDialog):
    def __init__(
        self,
        parent: QWidget | None,
        candidates: list[dict[str, Any]],
        initial_mapping: dict[str, int | None],
    ) -> None:
        super().__init__(parent)
        self._combos: dict[str, QComboBox] = {}
        self.setWindowTitle(MAPPING_DIALOG_TITLE)
        self.resize(900, 420)

        layout = QVBoxLayout(self)
        intro = QLabel(
            "自动匹配列失败。请为以下必需字段手动选择对应列。\n"
            "确认后的映射会在本次程序运行期间作为默认值复用。"
        )
        intro.setWordWrap(True)
        layout.addWidget(intro)

        grid = QGridLayout()
        grid.addWidget(QLabel("字段"), 0, 0)
        grid.addWidget(QLabel("Excel列"), 0, 1)

        for row_idx, field in enumerate(REQUIRED_FIELDS, start=1):
            label = QLabel(FIELD_LABELS[field])
            combo = QComboBox()
            combo.addItem("请选择列", userData=None)
            for candidate in candidates:
                combo.addItem(candidate["option_text"], userData=candidate["index"])

            initial_idx = initial_mapping.get(field)
            combo_idx = combo.findData(initial_idx)
            combo.setCurrentIndex(combo_idx if combo_idx >= 0 else 0)

            self._combos[field] = combo
            grid.addWidget(label, row_idx, 0)
            grid.addWidget(combo, row_idx, 1)

        layout.addLayout(grid)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def selected_mapping(self) -> dict[str, int | None]:
        return {field: combo.currentData() for field, combo in self._combos.items()}

    def _validate_and_accept(self) -> None:
        mapping = self.selected_mapping()
        missing = [FIELD_LABELS[field] for field, idx in mapping.items() if idx is None]
        if missing:
            QMessageBox.warning(self, "缺少列", f"请先为以下字段选择列：\n{', '.join(missing)}")
            return

        indices = [int(idx) for idx in mapping.values() if idx is not None]
        if len(indices) != len(set(indices)):
            QMessageBox.warning(self, "列重复", "每个字段必须对应不同的列，请不要重复选择同一列。")
            return

        self.accept()


class LabKPIDashboard(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.df_raw = pd.DataFrame()
        self.df_working = pd.DataFrame()
        self.chart_payloads: dict[str, Any] = {}
        self.session_column_mapping_defaults: dict[str, int] = {}

        self.setWindowTitle(APP_TITLE)
        self.resize(1520, 920)
        self._apply_palette()
        self._apply_stylesheet()
        self._configure_matplotlib_font()

        self._build_ui()
        self._auto_pick_default_file()

    @staticmethod
    def _configure_matplotlib_font() -> None:
        mpl.rcParams["font.sans-serif"] = [
            "Microsoft YaHei",
            "SimHei",
            "Segoe UI",
            "Arial Unicode MS",
            "DejaVu Sans",
        ]
        mpl.rcParams["axes.unicode_minus"] = False

    def _apply_palette(self) -> None:
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor("#F3F7FF"))
        palette.setColor(QPalette.ColorRole.Base, QColor("#FFFFFF"))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor("#EFF6FF"))
        palette.setColor(QPalette.ColorRole.Text, QColor("#133047"))
        palette.setColor(QPalette.ColorRole.ButtonText, QColor("#133047"))
        palette.setColor(QPalette.ColorRole.Highlight, QColor("#0EA5E9"))
        palette.setColor(QPalette.ColorRole.HighlightedText, QColor("#FFFFFF"))
        self.setPalette(palette)

    def _apply_stylesheet(self) -> None:
        self.setStyleSheet(
            """
            QWidget {
                font-family: 'Segoe UI', 'Microsoft YaHei UI', sans-serif;
                font-size: 13px;
                color: #16324f;
                background: #f4f8ff;
            }
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f7fbff, stop:0.55 #edf5ff, stop:1 #e7f3ff);
            }
            QLabel#heroTitle {
                font-size: 30px;
                font-weight: 700;
                letter-spacing: 1px;
                color: #0f2f4a;
            }
            QLabel#heroSub {
                color: #476b89;
                font-size: 13px;
            }
            QLabel#chartTitle {
                font-size: 16px;
                font-weight: 700;
                color: #14466c;
                padding: 0 0 4px 0;
            }
            QFrame#card, QGroupBox {
                background: rgba(255, 255, 255, 0.94);
                border: 1px solid #d4e7fb;
                border-radius: 14px;
            }
            QGroupBox {
                margin-top: 10px;
                padding-top: 12px;
                font-weight: 600;
                color: #1a4668;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 6px;
            }
            QLineEdit, QComboBox, QDateEdit, QSpinBox {
                border: 1px solid #bfdcf8;
                border-radius: 8px;
                padding: 6px 10px;
                background: #ffffff;
                min-height: 30px;
            }
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus, QSpinBox:focus {
                border: 2px solid #2ea6ff;
            }
            QPushButton {
                border: 0;
                border-radius: 10px;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #06b6d4, stop:1 #3b82f6);
                color: white;
                font-weight: 600;
                padding: 8px 14px;
                min-height: 34px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #0891b2, stop:1 #2563eb);
            }
            QPushButton:pressed {
                background: #1d4ed8;
            }
            QTabWidget::pane {
                border: 1px solid #d4e7fb;
                border-radius: 12px;
                background: rgba(255, 255, 255, 0.95);
                top: -1px;
            }
            QTabBar::tab {
                background: #e8f3ff;
                border: 1px solid #cde4fb;
                border-bottom: none;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                padding: 8px 14px;
                margin-right: 4px;
                color: #2b5a80;
                font-weight: 600;
            }
            QTabBar::tab:selected {
                background: #ffffff;
                color: #0d4266;
                border-color: #9fd0fa;
            }
            QTableWidget {
                background: white;
                border: 1px solid #d4e7fb;
                border-radius: 10px;
                gridline-color: #e5f0fb;
                selection-background-color: #d8efff;
                selection-color: #113350;
                alternate-background-color: #f6fbff;
            }
            QHeaderView::section {
                background: #e9f4ff;
                border: none;
                border-right: 1px solid #d8ebfd;
                border-bottom: 1px solid #d8ebfd;
                padding: 7px;
                color: #1f4d73;
                font-weight: 700;
            }
            QLabel#periodBadge {
                background: #e9f7ff;
                border: 1px solid #bfe6fb;
                border-radius: 8px;
                padding: 5px 10px;
                color: #155e7a;
                font-weight: 600;
            }
            """
        )

    def _build_ui(self) -> None:
        root = QWidget()
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(20, 16, 20, 16)
        root_layout.setSpacing(14)

        hero = QFrame()
        hero.setObjectName("card")
        hero_layout = QVBoxLayout(hero)
        hero_layout.setContentsMargins(20, 16, 20, 16)
        hero_layout.setSpacing(2)

        title = QLabel("Lab Productivity Navigator")
        title.setObjectName("heroTitle")
        subtitle = QLabel(
            "外链读取 xlsx（不改动原表），按闭区间起止日期分析任务排期、积分、FTE 与人效。"
        )
        subtitle.setObjectName("heroSub")
        hero_layout.addWidget(title)
        hero_layout.addWidget(subtitle)

        root_layout.addWidget(hero)

        control = QFrame()
        control.setObjectName("card")
        control_layout = QVBoxLayout(control)
        control_layout.setContentsMargins(14, 14, 14, 14)
        control_layout.setSpacing(10)

        file_group = QGroupBox("数据源")
        file_grid = QGridLayout(file_group)
        file_grid.setHorizontalSpacing(10)
        file_grid.setVerticalSpacing(8)

        self.file_path_edit = QLineEdit()
        self.file_path_edit.setPlaceholderText("选择生产数据统计 xlsx 文件")
        self.file_path_edit.setReadOnly(True)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(180)

        browse_btn = QPushButton("选择文件")
        browse_btn.clicked.connect(self._choose_file)

        load_btn = QPushButton("加载数据")
        load_btn.clicked.connect(self._load_data)

        file_grid.addWidget(QLabel("Excel 路径"), 0, 0)
        file_grid.addWidget(self.file_path_edit, 0, 1, 1, 5)
        file_grid.addWidget(browse_btn, 0, 6)
        file_grid.addWidget(QLabel("Sheet"), 1, 0)
        file_grid.addWidget(self.sheet_combo, 1, 1)
        file_grid.addWidget(load_btn, 1, 2)

        filter_group = QGroupBox("筛选器")
        filter_grid = QGridLayout(filter_group)
        filter_grid.setHorizontalSpacing(12)
        filter_grid.setVerticalSpacing(8)

        self.period_quick_combo = QComboBox()
        self.period_quick_combo.addItems(["本周", "本月"])
        apply_quick_btn = QPushButton("套用快捷时间")
        apply_quick_btn.clicked.connect(self._apply_quick_period)

        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")

        self.saturation_spin = QSpinBox()
        self.saturation_spin.setRange(1, 200)
        self.saturation_spin.setValue(8)

        self.researcher_combo = QComboBox()
        self.researcher_combo.addItem("全部")

        self.non_other_combo = QComboBox()
        self.non_other_combo.addItem(NON_OTHER_PLACEHOLDER, userData="")
        self.non_other_combo.currentIndexChanged.connect(self._on_non_other_changed)

        self.wbp_edit = QLineEdit()
        self.wbp_edit.setPlaceholderText("精确匹配，例如 WBP-001")

        self.product_edit = QLineEdit()
        self.product_edit.setPlaceholderText("精确匹配，例如 PID-123")

        self.period_badge = QLabel("时间段(闭区间): -")
        self.period_badge.setObjectName("periodBadge")

        apply_btn = QPushButton("应用筛选并刷新")
        apply_btn.clicked.connect(self._refresh_all_tabs)

        self.start_date_edit.dateChanged.connect(self._update_period_badge)
        self.end_date_edit.dateChanged.connect(self._update_period_badge)

        filter_grid.addWidget(QLabel("快捷时间"), 0, 0)
        filter_grid.addWidget(self.period_quick_combo, 0, 1)
        filter_grid.addWidget(apply_quick_btn, 0, 2)
        filter_grid.addWidget(QLabel("起始日期"), 0, 3)
        filter_grid.addWidget(self.start_date_edit, 0, 4)
        filter_grid.addWidget(QLabel("结束日期"), 0, 5)
        filter_grid.addWidget(self.end_date_edit, 0, 6)

        filter_grid.addWidget(QLabel("在研饱和阈值"), 1, 0)
        filter_grid.addWidget(self.saturation_spin, 1, 1)
        filter_grid.addWidget(QLabel("实验员筛选"), 1, 2)
        filter_grid.addWidget(self.researcher_combo, 1, 3)
        filter_grid.addWidget(self.period_badge, 1, 6)

        filter_grid.addWidget(QLabel("WBP Code"), 2, 0)
        filter_grid.addWidget(self.wbp_edit, 2, 1)
        filter_grid.addWidget(QLabel("Product ID"), 2, 2)
        filter_grid.addWidget(self.product_edit, 2, 3)
        filter_grid.addWidget(QLabel("FCT偶联类型"), 2, 4)
        filter_grid.addWidget(self.non_other_combo, 2, 5)
        filter_grid.addWidget(apply_btn, 2, 6)

        control_layout.addWidget(file_group)
        control_layout.addWidget(filter_group)
        root_layout.addWidget(control)

        content_row = QHBoxLayout()
        content_row.setSpacing(12)

        self.tab_widget = QTabWidget()
        self.tab_widget.currentChanged.connect(self._refresh_chart_for_current_tab)
        content_row.addWidget(self.tab_widget, 3)

        chart_card = QFrame()
        chart_card.setObjectName("card")
        chart_layout = QVBoxLayout(chart_card)
        chart_layout.setContentsMargins(12, 12, 12, 12)
        chart_layout.setSpacing(8)

        self.chart_title_label = QLabel("图表视图")
        self.chart_title_label.setObjectName("chartTitle")
        chart_layout.addWidget(self.chart_title_label)

        self.figure = Figure(figsize=(6, 5), dpi=100)
        self.chart_canvas = FigureCanvas(self.figure)
        self.chart_ax = self.figure.add_subplot(111)
        chart_layout.addWidget(self.chart_canvas, 1)

        content_row.addWidget(chart_card, 2)
        root_layout.addLayout(content_row, 1)

        self.detail_table = self._add_table_tab(
            "0. 筛选明细",
            "按当前筛选条件列出相关条目，包含阶段、新人标记与交付判定辅助列。",
        )
        self.wip_table = self._add_table_tab(
            "1. 任务排期",
            "每个实验员当前在研的未完成分子数（Unique C|D）以及是否可接收新任务；该页不受筛选时间段影响。",
        )
        self.score_table = self._add_table_tab(
            "2. 实验员积分",
            "按选定时间段 + 可选筛选条件，统计实验员积分。",
        )
        self.fte_person_table = self._add_fte_overview_tab(
            "3. FTE总览",
            "仅统计老员工，展示按实验员FTE与尚未获得的积分。",
        )
        self.efficiency_table = self._add_table_tab(
            "4. 人效",
            "仅统计老员工，且只将时间段内已完成Bulk阶段的分子视为交付。",
        )
        self.tat_table, self.tat_summary_table, self.tat_hint_label = self._add_tat_tab(
            "5. TAT计算",
            "统计当前所选FCT分子从 Pilot起始 到 Bulk完成 的天数；本模块忽略实验员/WBP/Product筛选。",
        )

        self.setCentralWidget(root)

        self._apply_quick_period(init_only=True)
        self._update_period_badge()
        self._draw_empty_chart("加载并筛选数据后显示图表")

    def _add_table_tab(self, tab_name: str, note: str) -> QTableWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        note_label = QLabel(note)
        note_label.setStyleSheet("color:#47708f; font-size:12px;")
        layout.addWidget(note_label)

        table = QTableWidget()
        self._init_table_widget(table)
        layout.addWidget(table, 1)

        self.tab_widget.addTab(tab, tab_name)
        return table

    def _add_fte_overview_tab(self, tab_name: str, note: str) -> QTableWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        note_label = QLabel(note)
        note_label.setStyleSheet("color:#47708f; font-size:12px;")
        note_label.setWordWrap(True)
        layout.addWidget(note_label)

        person_group = QGroupBox("按实验员FTE")
        person_layout = QVBoxLayout(person_group)
        person_layout.setContentsMargins(10, 10, 10, 10)
        person_table = QTableWidget()
        self._init_table_widget(person_table)
        person_layout.addWidget(person_table)
        layout.addWidget(person_group, 1)

        self.tab_widget.addTab(tab, tab_name)
        return person_table

    def _add_tat_tab(self, tab_name: str, note: str) -> tuple[QTableWidget, QTableWidget, QLabel]:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)

        note_label = QLabel(note)
        note_label.setStyleSheet("color:#47708f; font-size:12px;")
        note_label.setWordWrap(True)
        layout.addWidget(note_label)

        main_group = QGroupBox("TAT明细")
        main_layout = QVBoxLayout(main_group)
        main_layout.setContentsMargins(10, 10, 10, 10)
        tat_table = QTableWidget()
        self._init_table_widget(tat_table)
        main_layout.addWidget(tat_table)
        layout.addWidget(main_group, 1)

        summary_group = QGroupBox("TAT统计")
        summary_layout = QVBoxLayout(summary_group)
        summary_layout.setContentsMargins(10, 10, 10, 10)
        tat_summary_table = QTableWidget()
        self._init_table_widget(tat_summary_table)
        summary_layout.addWidget(tat_summary_table)

        hint_label = QLabel("TAT模块忽略实验员/WBP/Product筛选，仅按时间段和FCT统计。")
        hint_label.setStyleSheet("color:#8A3B12; font-size:12px;")
        hint_label.setWordWrap(True)
        summary_layout.addWidget(hint_label)
        layout.addWidget(summary_group)

        self.tab_widget.addTab(tab, tab_name)
        return tat_table, tat_summary_table, hint_label

    @staticmethod
    def _init_table_widget(table: QTableWidget) -> None:
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(False)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        table.horizontalHeader().setStretchLastSection(True)

    def _auto_pick_default_file(self) -> None:
        cwd = Path.cwd()
        candidates = [p for p in cwd.glob("*.xlsx") if not p.name.startswith("~$")]
        if not candidates:
            return
        target = sorted(candidates, key=lambda x: x.name)[0]
        self.file_path_edit.setText(str(target))
        self._load_sheet_names(target)

    def _choose_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择生产数据统计文件",
            str(Path.cwd()),
            "Excel Files (*.xlsx *.xlsm)",
        )
        if not file_path:
            return
        self.file_path_edit.setText(file_path)
        self._load_sheet_names(Path(file_path))

    def _load_sheet_names(self, path: Path) -> None:
        self.sheet_combo.clear()
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            self.sheet_combo.addItems(list(wb.sheetnames))
            wb.close()
        except Exception as exc:  # pragma: no cover - GUI message path
            QMessageBox.critical(self, "读取失败", f"无法读取 Excel Sheet 列表:\n{exc}")

    def _load_data(self) -> None:
        file_path = self.file_path_edit.text().strip()
        if not file_path:
            QMessageBox.warning(self, "缺少文件", "请先选择 xlsx 文件。")
            return

        sheet_name = self.sheet_combo.currentText().strip()
        if not sheet_name:
            QMessageBox.warning(self, "缺少 Sheet", "请选择数据所在 Sheet。")
            return

        path = Path(file_path)
        if not path.exists():
            QMessageBox.warning(self, "文件不存在", "所选 Excel 文件不存在。")
            return

        try:
            self.df_raw = self._read_source_sheet(path, sheet_name)
        except ColumnMappingCancelledError:
            return
        except Exception as exc:  # pragma: no cover - GUI message path
            QMessageBox.critical(self, "加载失败", f"解析数据失败:\n{exc}")
            return

        if self.df_raw.empty:
            QMessageBox.information(self, "无数据", "第4行起未检测到实验记录。")
            return

        self._rebuild_researcher_filter_options()
        self._rebuild_non_other_filter_options()
        self._refresh_all_tabs()

    def _read_source_sheet(self, file_path: Path, sheet_name: str) -> pd.DataFrame:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet 不存在: {sheet_name}")

            ws = wb[sheet_name]
            header_catalog = self._build_header_catalog(ws)
            epoch = wb.epoch
            resolved_mapping = self._resolve_column_mapping(ws, header_catalog, epoch)
            rows = self._read_rows_with_mapping(ws, resolved_mapping, epoch)
        finally:
            wb.close()

        df = pd.DataFrame(rows)
        if df.empty:
            return df

        df["start_date"] = pd.to_datetime(df["start_date"], errors="coerce")
        df["end_date"] = pd.to_datetime(df["end_date"], errors="coerce")
        df["score"] = pd.to_numeric(df["score"], errors="coerce").fillna(0.0)
        return df

    def _build_header_catalog(self, ws: Any) -> list[dict[str, Any]]:
        rows = list(
            ws.iter_rows(
                min_row=HEADER_ROW_1,
                max_row=HEADER_ROW_2,
                max_col=ws.max_column,
                values_only=True,
            )
        )
        catalog: list[dict[str, Any]] = []
        sample_values: list[Any] = [None] * ws.max_column

        for data_row in ws.iter_rows(min_row=4, max_col=ws.max_column, values_only=True):
            all_found = True
            for idx in range(ws.max_column):
                if sample_values[idx] is not None:
                    continue
                value = self._safe_cell(data_row, idx)
                if isinstance(value, str):
                    if value.strip():
                        sample_values[idx] = value
                elif value not in (None, ""):
                    sample_values[idx] = value
                if sample_values[idx] is None:
                    all_found = False
            if all_found:
                break

        for idx in range(ws.max_column):
            raw_parts: list[str] = []
            normalized_parts: list[str] = []

            for row in rows:
                raw = self._norm_text(self._safe_cell(row, idx))
                if raw and raw not in raw_parts:
                    raw_parts.append(raw)
                normalized = self._normalize_header_text(raw)
                if normalized and normalized not in normalized_parts:
                    normalized_parts.append(normalized)

            combined = " ".join(raw_parts)
            sample_value = sample_values[idx]
            display_text = combined if combined else "(空表头)"
            sample_text = self._norm_text(sample_value) or "-"
            column_letter = get_column_letter(idx + 1)
            catalog.append(
                {
                    "index": idx,
                    "column_letter": column_letter,
                    "display_text": display_text,
                    "sample_text": sample_text,
                    "option_text": f"{column_letter} | {display_text} | 示例: {sample_text}",
                    "normalized_parts": normalized_parts,
                    "normalized_combined": self._normalize_header_text(combined),
                }
            )

        return catalog

    def _resolve_column_mapping(
        self,
        ws: Any,
        header_catalog: list[dict[str, Any]],
        epoch: datetime,
    ) -> dict[str, int]:
        initial_mapping = self._build_initial_column_mapping(header_catalog)
        issues = self._validate_column_mapping(ws, initial_mapping, epoch)
        if not issues:
            return {field: int(idx) for field, idx in initial_mapping.items() if idx is not None}

        mapping = self._prompt_for_column_mapping(header_catalog, initial_mapping)
        if mapping is None:
            raise ColumnMappingCancelledError()

        manual_issues = self._validate_column_mapping(ws, mapping, epoch)
        if manual_issues:
            raise ValueError("人工选择后数据仍不合法：\n- " + "\n- ".join(manual_issues))

        self.session_column_mapping_defaults = {field: int(idx) for field, idx in mapping.items() if idx is not None}
        return self.session_column_mapping_defaults.copy()

    def _build_initial_column_mapping(self, header_catalog: list[dict[str, Any]]) -> dict[str, int | None]:
        mapping: dict[str, int | None] = {field: None for field in REQUIRED_FIELDS}
        max_index = len(header_catalog) - 1

        if self.session_column_mapping_defaults:
            for field, idx in self.session_column_mapping_defaults.items():
                if 0 <= idx <= max_index:
                    mapping[field] = idx

        if mapping[FIELD_NEW_EMPLOYEE] is None:
            mapping[FIELD_NEW_EMPLOYEE] = self._find_column_index(
                header_catalog,
                contains=self._normalize_header_text(HEADER_NEW_EMPLOYEE),
            )
        if mapping[FIELD_REQUEST_ACTIVITY] is None:
            mapping[FIELD_REQUEST_ACTIVITY] = self._find_column_index(
                header_catalog,
                equals=self._normalize_header_text(HEADER_REQUEST_ACTIVITY),
            )

        for field, idx in DEFAULT_FIXED_COLUMN_MAPPING.items():
            if mapping[field] is None and idx <= max_index:
                mapping[field] = idx

        return mapping

    def _validate_column_mapping(
        self,
        ws: Any,
        mapping: dict[str, int | None],
        epoch: datetime,
    ) -> list[str]:
        issues: list[str] = []

        missing_fields = [FIELD_LABELS[field] for field in REQUIRED_FIELDS if mapping.get(field) is None]
        if missing_fields:
            issues.append(f"缺少必需字段列：{', '.join(missing_fields)}")

        indices = [idx for idx in mapping.values() if idx is not None]
        if len(indices) != len(set(indices)):
            issues.append("必需字段中存在重复列映射。")

        if issues:
            return issues

        researcher_has_value = False
        start_has_date = False
        end_has_date = False
        score_has_number = False

        for row in ws.iter_rows(min_row=4, max_col=ws.max_column, values_only=True):
            researcher = self._norm_text(self._safe_cell(row, mapping[FIELD_RESEARCHER]))
            start_date = self._to_date(self._safe_cell(row, mapping[FIELD_START_DATE]), epoch)
            end_date = self._to_date(self._safe_cell(row, mapping[FIELD_END_DATE]), epoch)
            score_raw = self._safe_cell(row, mapping[FIELD_SCORE])

            if researcher:
                researcher_has_value = True
            if start_date is not None:
                start_has_date = True
            if end_date is not None:
                end_has_date = True
            if self._looks_like_number(score_raw):
                score_has_number = True

            if researcher_has_value and score_has_number and start_has_date and end_has_date:
                break

        if not researcher_has_value:
            issues.append("实验员列在数据区没有任何非空值。")
        if not start_has_date and not end_has_date:
            issues.append("起始日期与完成日期列都没有任何可解析日期。")
        if not score_has_number:
            issues.append("积分列没有任何可解析数字。")

        return issues

    def _prompt_for_column_mapping(
        self,
        header_catalog: list[dict[str, Any]],
        initial_mapping: dict[str, int | None],
    ) -> dict[str, int | None] | None:
        dialog = ColumnMappingDialog(self, header_catalog, initial_mapping)
        if dialog.exec() != int(QDialog.DialogCode.Accepted):
            return None
        return dialog.selected_mapping()

    def _read_rows_with_mapping(
        self,
        ws: Any,
        mapping: dict[str, int],
        epoch: datetime,
    ) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []

        for row in ws.iter_rows(min_row=4, max_col=ws.max_column, values_only=True):
            coupling_raw = self._safe_cell(row, mapping[FIELD_COUPLING_TYPE])
            wbp_raw = self._safe_cell(row, mapping[FIELD_WBP_CODE])
            product_raw = self._safe_cell(row, mapping[FIELD_PRODUCT_ID])
            researcher_raw = self._safe_cell(row, mapping[FIELD_RESEARCHER])
            start_raw = self._safe_cell(row, mapping[FIELD_START_DATE])
            end_raw = self._safe_cell(row, mapping[FIELD_END_DATE])
            score_raw = self._safe_cell(row, mapping[FIELD_SCORE])
            new_employee_raw = self._safe_cell(row, mapping[FIELD_NEW_EMPLOYEE])
            request_activity_raw = self._safe_cell(row, mapping[FIELD_REQUEST_ACTIVITY])

            if self._all_empty(
                coupling_raw,
                wbp_raw,
                product_raw,
                researcher_raw,
                start_raw,
                end_raw,
                score_raw,
                new_employee_raw,
                request_activity_raw,
            ):
                continue

            coupling_type = self._norm_text(coupling_raw)
            wbp_code = self._norm_text(wbp_raw)
            product_id = self._norm_text(product_raw)
            researcher = self._norm_text(researcher_raw)
            start_date = self._to_date(start_raw, epoch)
            end_date = self._to_date(end_raw, epoch)
            score = self._to_float(score_raw)
            new_employee_text = self._norm_text(new_employee_raw)
            request_activity = self._norm_text(request_activity_raw)
            stage_group = self._classify_stage(request_activity)

            if not researcher:
                continue

            cd_key = f"{wbp_code}|{product_id}" if (wbp_code or product_id) else ""
            rows.append(
                {
                    "coupling_type": coupling_type,
                    "wbp_code": wbp_code,
                    "product_id": product_id,
                    "researcher": researcher,
                    "start_date": start_date,
                    "end_date": end_date,
                    "score": score,
                    "cd_key": cd_key,
                    "new_employee_raw": new_employee_text,
                    "is_new_employee": new_employee_text == "是",
                    "request_activity_raw": request_activity,
                    "stage_group": stage_group,
                }
            )

        return rows

    def _find_column_index(
        self,
        header_catalog: list[dict[str, Any]],
        *,
        equals: str | None = None,
        contains: str | None = None,
    ) -> int | None:
        for item in header_catalog:
            tokens = set(item["normalized_parts"])
            combined = item["normalized_combined"]
            if combined:
                tokens.add(combined)

            if equals and equals in tokens:
                return int(item["index"])
            if contains and any(contains in token for token in tokens):
                return int(item["index"])

        return None

    @staticmethod
    def _normalize_header_text(value: Any) -> str:
        if value is None:
            return ""
        return "".join(str(value).split()).casefold()

    @staticmethod
    def _classify_stage(value: str) -> str:
        lowered = value.casefold()
        if lowered == "pilot":
            return "Pilot"
        if lowered == "bulk":
            return "Bulk"
        return UNKNOWN_STAGE

    @staticmethod
    def _safe_cell(row: Any, idx: int | None) -> Any:
        if idx is None:
            return None
        if idx < 0 or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _all_empty(*values: Any) -> bool:
        for v in values:
            if isinstance(v, str) and v.strip() != "":
                return False
            if v not in (None, ""):
                return False
        return True

    @staticmethod
    def _norm_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _to_float(value: Any) -> float:
        if value is None or value == "":
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", "")
        try:
            return float(text)
        except ValueError:
            return 0.0

    @staticmethod
    def _looks_like_number(value: Any) -> bool:
        if value is None or value == "":
            return False
        if isinstance(value, (int, float)):
            return True
        text = str(value).strip().replace(",", "")
        if text == "":
            return False
        try:
            float(text)
            return True
        except ValueError:
            return False

    @staticmethod
    def _to_date(value: Any, epoch: datetime) -> date | None:
        if value is None or value == "":
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                converted = from_excel(value, epoch=epoch)
                if isinstance(converted, datetime):
                    return converted.date()
                if isinstance(converted, date):
                    return converted
            except Exception:
                return None

        parsed = pd.to_datetime(str(value).strip(), errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()

    def _rebuild_researcher_filter_options(self) -> None:
        cur = self.researcher_combo.currentText().strip() or "全部"
        names = sorted(x for x in self.df_raw["researcher"].dropna().unique().tolist() if x)

        self.researcher_combo.blockSignals(True)
        self.researcher_combo.clear()
        self.researcher_combo.addItem("全部")
        self.researcher_combo.addItems(names)

        idx = self.researcher_combo.findText(cur)
        self.researcher_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.researcher_combo.blockSignals(False)

    def _rebuild_non_other_filter_options(self) -> None:
        current = self.non_other_combo.currentData() or ""
        types = sorted(x for x in self.df_raw["coupling_type"].dropna().unique().tolist() if x)

        self.non_other_combo.blockSignals(True)
        self.non_other_combo.clear()
        self.non_other_combo.addItem(NON_OTHER_PLACEHOLDER, userData="")
        for item in types:
            self.non_other_combo.addItem(item, userData=item)

        idx = self.non_other_combo.findData(current)
        self.non_other_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.non_other_combo.blockSignals(False)

    def _on_non_other_changed(self, *_: Any) -> None:
        if not self.df_raw.empty:
            self._refresh_all_tabs()

    def _current_filters(self) -> FilterState:
        researcher = self.researcher_combo.currentText().strip()
        if researcher == "全部":
            researcher = ""

        return FilterState(
            start_date=self.start_date_edit.date().toPyDate(),
            end_date=self.end_date_edit.date().toPyDate(),
            researcher=researcher,
            wbp_code=self.wbp_edit.text().strip(),
            product_id=self.product_edit.text().strip(),
            saturation_threshold=int(self.saturation_spin.value()),
            non_other_coupling_type=str(self.non_other_combo.currentData() or "").strip(),
        )

    @staticmethod
    def _period_range(period_type: str, anchor: date) -> tuple[date, date]:
        if period_type == "week":
            start = anchor - timedelta(days=anchor.weekday())
            end = start + timedelta(days=6)
            return start, end

        month_start = anchor.replace(day=1)
        if month_start.month == 12:
            next_month = month_start.replace(year=month_start.year + 1, month=1, day=1)
        else:
            next_month = month_start.replace(month=month_start.month + 1, day=1)
        return month_start, next_month - timedelta(days=1)

    def _apply_quick_period(self, init_only: bool = False) -> None:
        quick_type = self.period_quick_combo.currentText().strip()
        period_type = "week" if quick_type == "本周" else "month"
        start, end = self._period_range(period_type, date.today())

        self.start_date_edit.blockSignals(True)
        self.end_date_edit.blockSignals(True)
        self.start_date_edit.setDate(QDate(start.year, start.month, start.day))
        self.end_date_edit.setDate(QDate(end.year, end.month, end.day))
        self.start_date_edit.blockSignals(False)
        self.end_date_edit.blockSignals(False)

        self._update_period_badge()

        if not init_only and not self.df_raw.empty:
            self._refresh_all_tabs()

    def _update_period_badge(self) -> None:
        start = self.start_date_edit.date().toPyDate()
        end = self.end_date_edit.date().toPyDate()
        self.period_badge.setText(f"时间段(闭区间): {start:%Y-%m-%d} ~ {end:%Y-%m-%d}")

    def _prepare_df_with_stat_date(self, fs: FilterState) -> pd.DataFrame:
        if self.df_raw.empty:
            return pd.DataFrame()

        df = self.df_raw.copy()
        df["stat_date"] = df["end_date"]

        start_ts = pd.Timestamp(fs.start_date)
        end_ts = pd.Timestamp(fs.end_date)
        df["ag_in_period_for_efficiency"] = df["end_date"].between(start_ts, end_ts, inclusive="both")
        df["is_non_other"] = df["coupling_type"] == fs.non_other_coupling_type if fs.non_other_coupling_type else False
        df["efficiency_group"] = df["is_non_other"].map({True: "FCT", False: "非FCT"})
        df["coupling_group"] = df["is_non_other"].map(
            {True: fs.non_other_coupling_type or "FCT", False: "非FCT"}
        )
        df["is_delivered_bulk_record"] = (
            df["ag_in_period_for_efficiency"] & df["is_non_other"] & (df["stage_group"] == "Bulk") & (df["cd_key"] != "")
        )
        return df

    def _filter_common(
        self,
        df: pd.DataFrame,
        fs: FilterState,
        *,
        apply_period: bool,
        apply_researcher: bool,
    ) -> pd.DataFrame:
        if df.empty:
            return df

        out = df.copy()
        if apply_period:
            start_ts = pd.Timestamp(fs.start_date)
            end_ts = pd.Timestamp(fs.end_date)
            out = out[(out["stat_date"] >= start_ts) & (out["stat_date"] <= end_ts)]

        if apply_researcher and fs.researcher:
            out = out[out["researcher"] == fs.researcher]

        if fs.wbp_code:
            out = out[out["wbp_code"] == fs.wbp_code]

        if fs.product_id:
            out = out[out["product_id"] == fs.product_id]

        return out

    def _filter_old_employee_base(
        self,
        df: pd.DataFrame,
        fs: FilterState,
        *,
        apply_period: bool,
        apply_researcher: bool,
        include_other: bool,
    ) -> pd.DataFrame:
        base = self._filter_common(df, fs, apply_period=apply_period, apply_researcher=apply_researcher)
        if base.empty:
            return base

        out = base[~base["is_new_employee"]].copy()
        if not include_other:
            out = out[out["is_non_other"]].copy()
        return out

    def _filter_efficiency_base(self, df: pd.DataFrame, fs: FilterState) -> pd.DataFrame:
        base = self._filter_old_employee_base(
            df,
            fs,
            apply_period=False,
            apply_researcher=True,
            include_other=True,
        )
        if base.empty:
            return base
        return base[base["ag_in_period_for_efficiency"]].copy()

    @staticmethod
    def _compute_delivered_molecule_count(base: pd.DataFrame) -> float:
        if base.empty:
            return 0.0
        delivered = base.loc[base["is_delivered_bulk_record"] & (base["cd_key"] != ""), "cd_key"]
        return float(delivered.nunique())

    @staticmethod
    def _compute_selected_researcher_fte_sum(base: pd.DataFrame) -> tuple[float, list[str]]:
        if base.empty:
            return 0.0, []

        work = base.copy()
        work["fct_score_component"] = work["score"].where(work["is_non_other"], 0.0)
        grouped = (
            work.groupby("researcher", as_index=False)
            .agg(total_score=("score", "sum"), fct_score=("fct_score_component", "sum"))
            .sort_values("researcher")
        )
        grouped["fte_value"] = grouped.apply(
            lambda row: row["fct_score"] / row["total_score"] if row["total_score"] else 0.0,
            axis=1,
        )
        return float(grouped["fte_value"].sum()), grouped["researcher"].astype(str).tolist()

    def _render_message_table(self, table: QTableWidget, message: str) -> pd.DataFrame:
        result = pd.DataFrame([[message]], columns=["说明"])
        self._render_table(table, result)
        return result

    def _refresh_all_tabs(self) -> None:
        if self.df_raw.empty:
            QMessageBox.information(self, "提示", "请先加载数据文件。")
            return

        fs = self._current_filters()
        if fs.start_date > fs.end_date:
            QMessageBox.warning(self, "时间段错误", "起始日期不能晚于结束日期。")
            return

        df = self._prepare_df_with_stat_date(fs)
        self.df_working = df

        detail_df = self._fill_detail_tab(df, fs)
        wip_df = self._fill_wip_tab(df, fs)
        score_df = self._fill_score_tab(df, fs)
        if fs.non_other_coupling_type:
            fte_person_df, _ = self._fill_fte_by_person_tab(df, fs)
            _, efficiency_metrics = self._fill_efficiency_tab(df, fs)
            tat_df, tat_summary_df = self._fill_tat_tab(df, fs)
            tat_message = ""
        else:
            self._render_message_table(self.fte_person_table, NON_OTHER_REQUIRED_MESSAGE)
            self._render_message_table(self.efficiency_table, NON_OTHER_REQUIRED_MESSAGE)
            self._render_message_table(self.tat_table, NON_OTHER_REQUIRED_MESSAGE)
            self._render_message_table(self.tat_summary_table, NON_OTHER_REQUIRED_MESSAGE)
            fte_person_df = pd.DataFrame()
            efficiency_metrics = {}
            tat_df = pd.DataFrame()
            tat_summary_df = pd.DataFrame()
            tat_message = NON_OTHER_REQUIRED_MESSAGE

        self.chart_payloads = {
            "detail": detail_df,
            "wip": wip_df,
            "score": score_df,
            "fte_person": fte_person_df,
            "efficiency_metrics": efficiency_metrics,
            "tat": tat_df,
            "tat_summary": tat_summary_df,
            "tat_message": tat_message,
        }
        self._refresh_chart_for_current_tab()

    def _fill_detail_tab(self, df: pd.DataFrame, fs: FilterState) -> pd.DataFrame:
        base = self._filter_common(df, fs, apply_period=True, apply_researcher=True)

        if base.empty:
            result = pd.DataFrame(
                columns=[
                    "实验员(E)",
                    "偶联类型(A)",
                    "偶联分类(FCT/非FCT)",
                    "Request activity",
                    "阶段归类",
                    "是否新人",
                    "WBP Code(C)",
                    "Product ID(D)",
                    "起始日期(AF)",
                    "完成日期(AG)",
                    "统计日期(stat_date)",
                    "积分(AI)",
                    "CD_Key(C|D)",
                    "AG在时间段内",
                    "计入已完成分子",
                ]
            )
        else:
            result = (
                base[
                    [
                        "researcher",
                        "coupling_type",
                        "coupling_group",
                        "request_activity_raw",
                        "stage_group",
                        "new_employee_raw",
                        "wbp_code",
                        "product_id",
                        "start_date",
                        "end_date",
                        "stat_date",
                        "score",
                        "cd_key",
                        "ag_in_period_for_efficiency",
                        "is_delivered_bulk_record",
                    ]
                ]
                .rename(
                    columns={
                        "researcher": "实验员(E)",
                        "coupling_type": "偶联类型(A)",
                        "coupling_group": "偶联分类(FCT/非FCT)",
                        "request_activity_raw": "Request activity",
                        "stage_group": "阶段归类",
                        "new_employee_raw": "是否新人",
                        "wbp_code": "WBP Code(C)",
                        "product_id": "Product ID(D)",
                        "start_date": "起始日期(AF)",
                        "end_date": "完成日期(AG)",
                        "stat_date": "统计日期(stat_date)",
                        "score": "积分(AI)",
                        "cd_key": "CD_Key(C|D)",
                        "ag_in_period_for_efficiency": "AG在时间段内",
                        "is_delivered_bulk_record": "计入已完成分子",
                    }
                )
                .sort_values(["统计日期(stat_date)", "实验员(E)"], ascending=[True, True])
            )

        self._render_table(self.detail_table, result)
        return result

    def _fill_wip_tab(self, df: pd.DataFrame, fs: FilterState) -> pd.DataFrame:
        # 在研负载只看未完成实验（AG 为空），不受时间段影响
        base = self._filter_common(df, fs, apply_period=False, apply_researcher=True)
        unfinished = base[base["end_date"].isna()].copy()

        if unfinished.empty:
            self._render_message_table(self.wip_table, NO_WIP_MESSAGE)
            return pd.DataFrame(
                columns=["实验员", "未完成记录数", "未完成分子数(Unique C|D)", "饱和判断", "阈值"]
            )
        else:
            grouped = (
                unfinished.groupby("researcher", as_index=False)
                .agg(
                    unfinished_records=("researcher", "count"),
                    unfinished_molecules=("cd_key", lambda s: s[s != ""].nunique()),
                )
                .sort_values(["unfinished_molecules", "unfinished_records"], ascending=[False, False])
            )
            grouped["status"] = grouped["unfinished_molecules"].apply(
                lambda x: "可接新任务" if x <= fs.saturation_threshold else "已饱和"
            )
            grouped["threshold"] = fs.saturation_threshold
            result = grouped.rename(
                columns={
                    "researcher": "实验员",
                    "unfinished_records": "未完成记录数",
                    "unfinished_molecules": "未完成分子数(Unique C|D)",
                    "status": "饱和判断",
                    "threshold": "阈值",
                }
            )

        self._render_table(self.wip_table, result)
        return result

    def _fill_score_tab(self, df: pd.DataFrame, fs: FilterState) -> pd.DataFrame:
        base = self._filter_common(df, fs, apply_period=True, apply_researcher=True)

        if base.empty:
            result = pd.DataFrame(columns=["实验员", "积分合计", "记录数"])
        else:
            grouped = (
                base.groupby("researcher", as_index=False)
                .agg(
                    total_score=("score", "sum"),
                    record_count=("researcher", "count"),
                )
                .sort_values("total_score", ascending=False)
            )
            result = grouped.rename(
                columns={
                    "researcher": "实验员",
                    "total_score": "积分合计",
                    "record_count": "记录数",
                }
            )

        self._render_table(self.score_table, result)
        return result

    def _fill_fte_by_person_tab(self, df: pd.DataFrame, fs: FilterState) -> tuple[pd.DataFrame, dict[str, float]]:
        base_with_other = self._filter_old_employee_base(
            df,
            fs,
            apply_period=True,
            apply_researcher=False,
            include_other=True,
        )
        base_without_other = self._filter_old_employee_base(
            df,
            fs,
            apply_period=True,
            apply_researcher=False,
            include_other=False,
        )

        if base_with_other.empty and base_without_other.empty:
            result = pd.DataFrame(
                columns=[
                    "实验员",
                    "总积分(含非FCT)",
                    "FCT积分",
                    "FTE(含非FCT)",
                    "总积分(不含非FCT)",
                    "FTE(不含非FCT)",
                    "尚未获得的积分",
                ]
            )
            fte_totals = {"with_other": 0.0, "without_other": 0.0}
        else:
            grouped = base_with_other.groupby("researcher", as_index=False).agg(total_with_other=("score", "sum"))
            non_other_grouped = (
                base_with_other[base_with_other["is_non_other"]]
                .groupby("researcher", as_index=False)
                .agg(non_other_score=("score", "sum"))
            )
            merged = grouped.merge(non_other_grouped, on="researcher", how="left").fillna({"non_other_score": 0.0})
            merged["fte_with_other"] = merged.apply(
                lambda row: row["non_other_score"] / row["total_with_other"] if row["total_with_other"] else 0.0,
                axis=1,
            )
            without_grouped = (
                base_without_other.groupby("researcher", as_index=False).agg(total_without_other=("score", "sum"))
                if not base_without_other.empty
                else pd.DataFrame(columns=["researcher", "total_without_other"])
            )
            merged = merged.merge(without_grouped, on="researcher", how="outer").fillna(
                {
                    "total_with_other": 0.0,
                    "non_other_score": 0.0,
                    "fte_with_other": 0.0,
                    "total_without_other": 0.0,
                }
            )
            merged["fte_without_other"] = merged["total_without_other"].apply(lambda value: 1.0 if value else 0.0)
            unearned_base = self._filter_old_employee_base(
                df,
                fs,
                apply_period=False,
                apply_researcher=False,
                include_other=True,
            )
            unearned_grouped = (
                unearned_base[unearned_base["end_date"].isna()]
                .groupby("researcher", as_index=False)
                .agg(unearned_score=("score", "sum"))
                if not unearned_base.empty
                else pd.DataFrame(columns=["researcher", "unearned_score"])
            )
            merged = merged.merge(unearned_grouped, on="researcher", how="outer").fillna(
                {
                    "total_with_other": 0.0,
                    "non_other_score": 0.0,
                    "fte_with_other": 0.0,
                    "total_without_other": 0.0,
                    "fte_without_other": 0.0,
                    "unearned_score": 0.0,
                }
            )
            merged = merged.sort_values(["fte_with_other", "total_with_other"], ascending=[False, False])
            fte_totals = {
                "with_other": float(merged["fte_with_other"].sum()),
                "without_other": float(merged["fte_without_other"].sum()),
            }
            result = merged.rename(
                columns={
                    "researcher": "实验员",
                    "total_with_other": "总积分(含非FCT)",
                    "non_other_score": "FCT积分",
                    "fte_with_other": "FTE(含非FCT)",
                    "total_without_other": "总积分(不含非FCT)",
                    "fte_without_other": "FTE(不含非FCT)",
                    "unearned_score": "尚未获得的积分",
                }
            )

        self._render_table(self.fte_person_table, result, percent_cols={"FTE(含非FCT)", "FTE(不含非FCT)"})
        return result, fte_totals

    def _fill_efficiency_tab(
        self,
        df: pd.DataFrame,
        fs: FilterState,
    ) -> tuple[pd.DataFrame, dict[str, float]]:
        base = self._filter_efficiency_base(df, fs)
        completed_molecules = self._compute_delivered_molecule_count(base)
        selected_researcher_fte_sum, researchers = self._compute_selected_researcher_fte_sum(base)
        efficiency = completed_molecules / selected_researcher_fte_sum if selected_researcher_fte_sum else 0.0
        researcher_text = "、".join(researchers) if researchers else "无"

        rows = [
            ["时间段(AG闭区间)", f"{fs.start_date:%Y-%m-%d} ~ {fs.end_date:%Y-%m-%d}"],
            ["FCT偶联类型", fs.non_other_coupling_type],
            ["实验员集合", researcher_text],
            ["已完成分子数", completed_molecules],
            ["实验员集合FTE值之和", selected_researcher_fte_sum],
            ["人效", efficiency],
        ]
        result = pd.DataFrame(rows, columns=["指标", "值"])
        self._render_table(self.efficiency_table, result)
        return result, {
            "completed_molecules": completed_molecules,
            "selected_researcher_fte_sum": selected_researcher_fte_sum,
            "efficiency": efficiency,
        }

    def _fill_tat_tab(self, df: pd.DataFrame, fs: FilterState) -> tuple[pd.DataFrame, pd.DataFrame]:
        start_ts = pd.Timestamp(fs.start_date)
        end_ts = pd.Timestamp(fs.end_date)

        fct_df = df[df["is_non_other"] & (df["cd_key"] != "")].copy()
        bulk_df = fct_df[
            (fct_df["stage_group"] == "Bulk")
            & fct_df["end_date"].notna()
            & fct_df["end_date"].between(start_ts, end_ts, inclusive="both")
        ].copy()
        pilot_df = fct_df[
            (fct_df["stage_group"] == "Pilot")
            & fct_df["start_date"].notna()
        ].copy()

        if bulk_df.empty or pilot_df.empty:
            result = pd.DataFrame(columns=["分子标识(C|D)", "Pilot起始日", "Bulk完成日", "TAT(天)"])
            summary = pd.DataFrame(
                [["总分子数", 0], ["标红分子数", 0]],
                columns=["指标", "值"],
            )
            self._render_table(self.tat_table, result)
            self._render_table(self.tat_summary_table, summary)
            self.tat_hint_label.setText("TAT模块忽略实验员/WBP/Product筛选，仅按时间段和FCT统计。")
            return result, summary

        bulk_grouped = (
            bulk_df.groupby("cd_key", as_index=False)
            .agg(bulk_end_date=("end_date", "max"))
        )
        pilot_grouped = (
            pilot_df.groupby("cd_key", as_index=False)
            .agg(pilot_start_date=("start_date", "min"))
        )
        merged = bulk_grouped.merge(pilot_grouped, on="cd_key", how="inner")
        merged["tat_days"] = (merged["bulk_end_date"] - merged["pilot_start_date"]).dt.days
        merged = merged.sort_values(["bulk_end_date", "tat_days", "cd_key"], ascending=[False, False, True])

        result = merged.rename(
            columns={
                "cd_key": "分子标识(C|D)",
                "pilot_start_date": "Pilot起始日",
                "bulk_end_date": "Bulk完成日",
                "tat_days": "TAT(天)",
            }
        )
        summary = pd.DataFrame(
            [
                ["总分子数", int(len(result))],
                ["标红分子数", int((result["TAT(天)"] > 10).sum())],
            ],
            columns=["指标", "值"],
        )

        self._render_table(self.tat_table, result)
        self._highlight_tat_rows(result)
        self._render_table(self.tat_summary_table, summary)
        self.tat_hint_label.setText("TAT模块忽略实验员/WBP/Product筛选，仅按时间段和FCT统计。")
        return result, summary

    def _highlight_tat_rows(self, result: pd.DataFrame) -> None:
        if result.empty or "TAT(天)" not in result.columns:
            return

        tat_col_idx = result.columns.get_loc("TAT(天)")
        for row_idx in range(len(result)):
            value = result.iloc[row_idx, tat_col_idx]
            if pd.isna(value) or float(value) <= 10:
                continue
            item = self.tat_table.item(row_idx, tat_col_idx)
            if item is None:
                continue
            item.setForeground(QBrush(QColor("#DC2626")))
            item.setBackground(QBrush(QColor("#FEE2E2")))
            font = item.font()
            font.setBold(True)
            item.setFont(font)

    def _refresh_chart_for_current_tab(self, *_: Any) -> None:
        if not hasattr(self, "tab_widget"):
            return

        idx = self.tab_widget.currentIndex()
        if idx == 3:
            self.figure.clear()
            person_ax = self.figure.add_subplot(111)
            self._style_axis_base(person_ax)
            self.chart_ax = person_ax
            self.chart_title_label.setText("FTE总览图")
            self._plot_fte_person_chart(person_ax)
            self.figure.tight_layout()
            self.chart_canvas.draw_idle()
            return

        self.figure.clear()
        self.chart_ax = self.figure.add_subplot(111)
        self._style_axis_base(self.chart_ax)

        if idx == 0:
            self.chart_title_label.setText("筛选明细趋势图")
            self._plot_detail_chart(self.chart_ax)
        elif idx == 1:
            self.chart_title_label.setText("任务排期图")
            self._plot_wip_chart(self.chart_ax)
        elif idx == 2:
            self.chart_title_label.setText("实验员积分图")
            self._plot_score_chart(self.chart_ax)
        elif idx == 4:
            self.chart_title_label.setText("人效图")
            self._plot_efficiency_chart(self.chart_ax)
        elif idx == 5:
            self.chart_title_label.setText("TAT图")
            self._plot_tat_chart(self.chart_ax)
        else:
            self._draw_empty_chart("图表数据不可用")
            return

        self.figure.tight_layout()
        self.chart_canvas.draw_idle()

    @staticmethod
    def _style_axis_base(ax: Any) -> None:
        ax.set_facecolor("#FCFEFF")
        ax.grid(axis="y", linestyle="--", linewidth=0.8, color="#DCEBFA", alpha=0.9)
        for spine in ["top", "right"]:
            ax.spines[spine].set_visible(False)
        ax.spines["left"].set_color("#B9D5EE")
        ax.spines["bottom"].set_color("#B9D5EE")
        ax.tick_params(colors="#2B5679")

    def _draw_empty_chart(self, text: str) -> None:
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        ax.axis("off")
        ax.text(
            0.5,
            0.5,
            text,
            ha="center",
            va="center",
            fontsize=13,
            color="#5D7994",
            transform=ax.transAxes,
        )
        self.chart_canvas.draw_idle()

    def _plot_detail_chart(self, ax: Any) -> None:
        df = self.chart_payloads.get("detail", pd.DataFrame())
        if df.empty:
            ax.axis("off")
            ax.text(0.5, 0.5, "无筛选明细", ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        dates = pd.to_datetime(df["统计日期(stat_date)"], errors="coerce").dropna()
        if dates.empty:
            ax.axis("off")
            ax.text(0.5, 0.5, "无可用统计日期", ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        trend = dates.dt.date.value_counts().sort_index()
        labels = [d.strftime("%m-%d") for d in trend.index]
        values = trend.values
        bars = ax.bar(labels, values, color="#36A3FF", edgecolor="#1F7BC9", linewidth=1.0)
        ax.set_title("筛选记录数趋势")
        ax.set_xlabel("统计日期")
        ax.set_ylabel("记录数")
        for bar in bars:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h, f"{int(h)}", ha="center", va="bottom", fontsize=9, color="#245981")

    def _plot_wip_chart(self, ax: Any) -> None:
        df = self.chart_payloads.get("wip", pd.DataFrame())
        if df.empty:
            ax.axis("off")
            ax.text(0.5, 0.5, NO_WIP_MESSAGE, ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        x = df["实验员"].astype(str).tolist()
        y = pd.to_numeric(df["未完成分子数(Unique C|D)"], errors="coerce").fillna(0).tolist()
        bars = ax.bar(x, y, color="#06B6D4", edgecolor="#0891B2", linewidth=1.0)
        ax.set_title("实验员在研分子负载")
        ax.set_xlabel("实验员")
        ax.set_ylabel("Unique C|D")
        for bar in bars:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h, f"{int(h)}", ha="center", va="bottom", fontsize=9, color="#245981")

    def _plot_score_chart(self, ax: Any) -> None:
        df = self.chart_payloads.get("score", pd.DataFrame())
        if df.empty:
            ax.axis("off")
            ax.text(0.5, 0.5, "无实验员积分数据", ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        x = df["实验员"].astype(str).tolist()
        y = pd.to_numeric(df["积分合计"], errors="coerce").fillna(0).tolist()
        bars = ax.bar(x, y, color="#2DD4BF", edgecolor="#0F766E", linewidth=1.0)
        ax.set_title("实验员积分排行")
        ax.set_xlabel("实验员")
        ax.set_ylabel("积分")
        for bar in bars:
            h = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, h, f"{h:.1f}", ha="center", va="bottom", fontsize=9, color="#245981")

    def _plot_fte_person_chart(self, ax: Any) -> None:
        df = self.chart_payloads.get("fte_person", pd.DataFrame())
        if df.empty or "实验员" not in df.columns:
            ax.axis("off")
            ax.text(0.5, 0.5, NON_OTHER_REQUIRED_MESSAGE, ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        x = df["实验员"].astype(str).tolist()
        y_with = pd.to_numeric(df["FTE(含非FCT)"], errors="coerce").fillna(0).tolist()
        y_without = pd.to_numeric(df["FTE(不含非FCT)"], errors="coerce").fillna(0).tolist()
        positions = range(len(x))
        width = 0.38

        bars_with = ax.bar(
            [p - width / 2 for p in positions],
            y_with,
            width=width,
            color="#22C55E",
            edgecolor="#15803D",
            linewidth=1.0,
            label="含非FCT",
        )
        bars_without = ax.bar(
            [p + width / 2 for p in positions],
            y_without,
            width=width,
            color="#F59E0B",
            edgecolor="#B45309",
            linewidth=1.0,
            label="不含非FCT",
        )
        ax.set_title("按实验员FTE")
        ax.set_xlabel("实验员")
        ax.set_ylabel("FTE")
        ax.set_xticks(list(positions))
        ax.set_xticklabels(x)
        ax.yaxis.set_major_formatter(PercentFormatter(1.0))
        ax.legend()
        for bars in [bars_with, bars_without]:
            for bar in bars:
                h = bar.get_height()
                ax.text(bar.get_x() + bar.get_width() / 2, h, f"{h * 100:.1f}%", ha="center", va="bottom", fontsize=9, color="#245981")

    def _plot_efficiency_chart(self, ax: Any) -> None:
        metrics = self.chart_payloads.get("efficiency_metrics", {})
        if not metrics:
            ax.axis("off")
            ax.text(0.5, 0.5, NON_OTHER_REQUIRED_MESSAGE, ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        labels = [
            "已完成分子数",
            "实验员集合FTE值之和",
            "人效",
        ]
        values = [
            float(metrics.get("completed_molecules", 0.0)),
            float(metrics.get("selected_researcher_fte_sum", 0.0)),
            float(metrics.get("efficiency", 0.0)),
        ]
        if all(v == 0 for v in values):
            ax.axis("off")
            ax.text(0.5, 0.5, "无人效统计数据", ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        bars = ax.bar(
            labels,
            values,
            color=["#38BDF8", "#F59E0B", "#8B5CF6"],
            edgecolor="#475569",
            linewidth=0.8,
        )
        ax.set_title("人效")
        ax.set_ylabel("数值")
        ax.tick_params(axis="x", rotation=10)
        for i, bar in enumerate(bars):
            h = bar.get_height()
            fmt = f"{h:.3f}" if i == 2 else f"{h:.2f}".rstrip("0").rstrip(".")
            ax.text(bar.get_x() + bar.get_width() / 2, h, fmt, ha="center", va="bottom", fontsize=9, color="#245981")

    def _plot_tat_chart(self, ax: Any) -> None:
        df = self.chart_payloads.get("tat", pd.DataFrame())
        message = str(self.chart_payloads.get("tat_message", "") or "")
        if message:
            ax.axis("off")
            ax.text(0.5, 0.5, message, ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return
        if df.empty or "分子标识(C|D)" not in df.columns:
            ax.axis("off")
            ax.text(0.5, 0.5, "无TAT数据", ha="center", va="center", color="#5D7994", transform=ax.transAxes)
            return

        x = df["分子标识(C|D)"].astype(str).tolist()
        y = pd.to_numeric(df["TAT(天)"], errors="coerce").fillna(0).tolist()
        colors = ["#EF4444" if value > 10 else "#0EA5E9" for value in y]
        edge_colors = ["#B91C1C" if value > 10 else "#0369A1" for value in y]
        bars = ax.bar(x, y, color=colors, edgecolor=edge_colors, linewidth=1.0)
        ax.set_title("FCT分子TAT")
        ax.set_xlabel("分子标识(C|D)")
        ax.set_ylabel("TAT(天)")
        ax.tick_params(axis="x", rotation=20)
        for bar, value in zip(bars, y):
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                f"{int(value) if float(value).is_integer() else value}",
                ha="center",
                va="bottom",
                fontsize=9,
                color="#245981",
            )

    def _render_table(
        self,
        table: QTableWidget,
        df: pd.DataFrame,
        *,
        percent_cols: set[str] | None = None,
        percent_rows: set[str] | None = None,
    ) -> None:
        percent_cols = percent_cols or set()
        percent_rows = percent_rows or set()

        table.clear()
        table.setRowCount(0)

        if df.empty:
            table.setColumnCount(1)
            table.setHorizontalHeaderLabels(["结果"])
            table.setRowCount(1)
            table.setItem(0, 0, QTableWidgetItem("无匹配数据"))
            return

        headers = df.columns.tolist()
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.setRowCount(len(df))

        first_col_name = headers[0] if headers else ""

        for r, (_, row) in enumerate(df.iterrows()):
            row_label = str(row[first_col_name]) if first_col_name else ""
            for c, col in enumerate(headers):
                val = row[col]
                txt = self._format_cell(val, col, row_label, percent_cols, percent_rows)
                item = QTableWidgetItem(txt)
                if isinstance(val, bool):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                elif isinstance(val, (int, float)):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight)
                else:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                table.setItem(r, c, item)

        table.resizeColumnsToContents()

    @staticmethod
    def _format_cell(
        val: Any,
        col_name: str,
        row_label: str,
        percent_cols: set[str],
        percent_rows: set[str],
    ) -> str:
        if pd.isna(val):
            return ""

        if isinstance(val, pd.Timestamp):
            return val.strftime("%Y-%m-%d")

        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")

        if isinstance(val, date):
            return val.strftime("%Y-%m-%d")

        if isinstance(val, bool):
            return "是" if val else "否"

        if isinstance(val, (int, float)):
            if col_name in percent_cols or row_label in percent_rows:
                return f"{float(val) * 100:.2f}%"

            num = float(val)
            if abs(num - round(num)) < 1e-9:
                return f"{int(round(num)):,}"
            return f"{num:,.3f}".rstrip("0").rstrip(".")

        return str(val)


def main() -> None:
    app = QApplication([])
    app.setFont(QFont("Segoe UI", 10))
    win = LabKPIDashboard()
    win.show()
    app.exec()


if __name__ == "__main__":
    main()
