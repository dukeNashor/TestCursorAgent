"""
视图层 - PyQt版本
使用PyQt构建用户界面，保持缓存机制确保流畅性能
"""
import sys
import os
import io
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QLineEdit, QTextEdit, QComboBox, QScrollArea,
    QListWidget, QListWidgetItem, QFrame, QSplitter, QMessageBox, QFileDialog,
    QDialog, QDialogButtonBox, QSpinBox, QDoubleSpinBox, QGroupBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QTabWidget, QProgressBar, QDateEdit, QInputDialog, QSizePolicy, QCheckBox
)
from PyQt5.QtCore import Qt, QSize, pyqtSignal, QThread, QTimer, QDate
from PyQt5.QtGui import QPixmap, QFont, QColor, QImage

# 从模块导入
from material.models import Material, Order, OrderStatus, Priority
from material.controller import MaterialController, OrderController, ReportController
from adc.models import ADC, ADCSpec, ADCOutbound, ADCInbound, ADCMovementItem
from adc.controller import ADCController, PRESET_SPECS
from adc_workflow.models import ADCWorkflow, ADCWorkflowStep, ADCExperimentResult, AppUser
from adc_workflow.controller import ADCWorkflowController
from adc_workflow.request_schema import ordered_request_items_for_display
from adc_workflow import sp_dar8
from database import (
    load_config, save_config, get_database_list, add_database, 
    remove_database, set_current_database, DatabaseManager
)


class EmojiPicker(QDialog):
    """Emoji选择器"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择Emoji")
        self.setFixedSize(400, 300)
        self.result = None
        
        # 常用emoji列表
        emojis = [
            "🧪", "🔬", "⚗️", "🧬", "🦠", "💊", "💉", "🧫", "🔍", "📊",
            "📈", "📉", "⚠️", "✅", "❌", "🔴", "🟡", "🟢", "🔵", "⚪",
            "📝", "📋", "📌", "🔗", "💡", "🔧", "⚙️", "🔩", "📦", "📋",
            "🏷️", "📅", "⏰", "📍", "🎯", "💯", "⭐", "🔥", "💎", "🌟"
        ]
        
        layout = QGridLayout()
        self.setLayout(layout)
        
        row, col = 0, 0
        for emoji in emojis:
            btn = QPushButton(emoji)
            btn.setFixedSize(40, 40)
            btn.setFont(QFont("Arial", 16))
            btn.clicked.connect(lambda checked, e=emoji: self._select_emoji(e))
            layout.addWidget(btn, row, col)
            col += 1
            if col >= 10:
                col = 0
                row += 1
    
    def _select_emoji(self, emoji):
        self.result = emoji
        self.accept()


class MaterialDialog(QDialog):
    """物料编辑对话框"""
    
    def __init__(self, parent=None, material: Optional[Material] = None, material_controller=None):
        super().__init__(parent)
        self.material = material
        self.material_controller = material_controller
        self.result = None
        self.image_paths = []
        
        self.setWindowTitle("编辑物料" if material else "添加物料")
        self.setFixedSize(600, 700)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        scroll_layout = QVBoxLayout()
        content.setLayout(scroll_layout)
        
        # 物料名称
        scroll_layout.addWidget(QLabel("物料名称 *:"))
        self.name_edit = QLineEdit(self.material.name if self.material else "")
        scroll_layout.addWidget(self.name_edit)
        
        # 类别
        scroll_layout.addWidget(QLabel("类别 *:"))
        self.category_combo = QComboBox()
        self.category_combo.addItems(["试剂", "耗材", "设备", "工具", "其他"])
        if self.material:
            self.category_combo.setCurrentText(self.material.category)
        scroll_layout.addWidget(self.category_combo)
        
        # 数量
        scroll_layout.addWidget(QLabel("数量 *:"))
        self.quantity_edit = QSpinBox()
        self.quantity_edit.setMaximum(999999)
        self.quantity_edit.setValue(self.material.quantity if self.material else 0)
        scroll_layout.addWidget(self.quantity_edit)
        
        # 单位
        scroll_layout.addWidget(QLabel("单位 *:"))
        self.unit_combo = QComboBox()
        self.unit_combo.addItems(["个", "瓶", "盒", "包", "升", "毫升", "克", "千克", "米", "厘米"])
        if self.material:
            self.unit_combo.setCurrentText(self.material.unit)
        scroll_layout.addWidget(self.unit_combo)
        
        # 最低库存
        scroll_layout.addWidget(QLabel("最低库存:"))
        self.min_stock_edit = QSpinBox()
        self.min_stock_edit.setMaximum(999999)
        self.min_stock_edit.setValue(self.material.min_stock if self.material else 0)
        scroll_layout.addWidget(self.min_stock_edit)
        
        # 存放位置
        scroll_layout.addWidget(QLabel("存放位置:"))
        self.location_edit = QLineEdit(self.material.location if self.material else "")
        scroll_layout.addWidget(self.location_edit)
        
        # 供应商
        scroll_layout.addWidget(QLabel("供应商:"))
        self.supplier_edit = QLineEdit(self.material.supplier if self.material else "")
        scroll_layout.addWidget(self.supplier_edit)
        
        # 描述
        scroll_layout.addWidget(QLabel("描述:"))
        desc_layout = QHBoxLayout()
        self.desc_text = QTextEdit()
        self.desc_text.setMaximumHeight(100)
        if self.material and self.material.description:
            self.desc_text.setPlainText(self.material.description)
        desc_layout.addWidget(self.desc_text)
        
        emoji_btn = QPushButton("😀")
        emoji_btn.clicked.connect(self._insert_emoji)
        desc_layout.addWidget(emoji_btn)
        scroll_layout.addLayout(desc_layout)
        
        # 图片管理
        scroll_layout.addWidget(QLabel("图片:"))
        img_layout = QHBoxLayout()
        self.image_list = QListWidget()
        self.image_list.setMaximumHeight(100)
        img_layout.addWidget(self.image_list)
        
        img_btn_layout = QVBoxLayout()
        add_btn = QPushButton("添加")
        add_btn.clicked.connect(self._add_image)
        remove_btn = QPushButton("删除")
        remove_btn.clicked.connect(self._remove_image)
        view_btn = QPushButton("查看")
        view_btn.clicked.connect(self._view_image)
        
        img_btn_layout.addWidget(add_btn)
        img_btn_layout.addWidget(remove_btn)
        img_btn_layout.addWidget(view_btn)
        img_layout.addLayout(img_btn_layout)
        scroll_layout.addLayout(img_layout)
        
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _insert_emoji(self):
        emoji_picker = EmojiPicker(self)
        if emoji_picker.exec_() == QDialog.Accepted:
            emoji = emoji_picker.result
            if emoji:
                self.desc_text.insertPlainText(emoji)
    
    def _add_image(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "选择图片", "", "图片文件 (*.jpg *.jpeg *.png *.gif *.bmp)"
        )
        if filename:
            self.image_paths.append(filename)
            self.image_list.addItem(os.path.basename(filename))
    
    def _remove_image(self):
        current_item = self.image_list.currentItem()
        if current_item:
            index = self.image_list.row(current_item)
            self.image_list.takeItem(index)
            del self.image_paths[index]
    
    def _view_image(self):
        current_item = self.image_list.currentItem()
        if not current_item:
            QMessageBox.warning(self, "警告", "请选择要查看的图片")
            return
        
        index = self.image_list.row(current_item)
        image_path = self.image_paths[index]
        
        dialog = QDialog(self)
        dialog.setWindowTitle("查看图片")
        dialog.setFixedSize(800, 600)
        layout = QVBoxLayout()
        
        pixmap = QPixmap(image_path)
        label = QLabel()
        label.setPixmap(pixmap.scaled(700, 500, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        label.setAlignment(Qt.AlignCenter)
        layout.addWidget(label)
        
        path_label = QLabel(image_path)
        path_label.setWordWrap(True)
        layout.addWidget(path_label)
        
        dialog.setLayout(layout)
        dialog.exec_()
    
    def _save(self):
        if not self.name_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入物料名称")
            return
        
        if not self.category_combo.currentText():
            QMessageBox.critical(self, "错误", "请选择类别")
            return
        
        if not self.unit_combo.currentText():
            QMessageBox.critical(self, "错误", "请输入单位")
            return
        
        # 读取图片文件为二进制数据
        image_data_list = []
        for image_path in self.image_paths:
            if os.path.exists(image_path):
                with open(image_path, 'rb') as f:
                    image_bytes = f.read()
                image_data_list.append(image_bytes)
        
        material = Material(
            id=self.material.id if self.material else None,
            name=self.name_edit.text().strip(),
            category=self.category_combo.currentText(),
            description=self.desc_text.toPlainText().strip(),
            quantity=self.quantity_edit.value(),
            unit=self.unit_combo.currentText(),
            min_stock=self.min_stock_edit.value(),
            location=self.location_edit.text().strip(),
            supplier=self.supplier_edit.text().strip(),
            images=image_data_list
        )
        
        self.result = material
        self.accept()


class OrderDialog(QDialog):
    """订单编辑对话框"""
    
    def __init__(self, parent=None, order: Optional[Order] = None, material_controller: MaterialController = None):
        super().__init__(parent)
        self.order = order
        self.material_controller = material_controller
        self.result = None
        self.materials = []
        
        self.setWindowTitle("编辑订单" if order else "创建订单")
        self.setFixedSize(800, 700)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 订单信息
        info_group = QGroupBox("订单信息")
        info_layout = QGridLayout()
        
        info_layout.addWidget(QLabel("申请人 *:"), 0, 0)
        self.requester_edit = QLineEdit(self.order.requester if self.order else "")
        info_layout.addWidget(self.requester_edit, 0, 1)
        
        info_layout.addWidget(QLabel("部门:"), 0, 2)
        self.department_edit = QLineEdit(self.order.department if self.order else "")
        info_layout.addWidget(self.department_edit, 0, 3)
        
        info_layout.addWidget(QLabel("优先级:"), 1, 0)
        self.priority_combo = QComboBox()
        self.priority_combo.addItems([p.value for p in Priority])
        if self.order:
            self.priority_combo.setCurrentText(self.order.priority)
        info_layout.addWidget(self.priority_combo, 1, 1)
        
        info_layout.addWidget(QLabel("状态:"), 1, 2)
        self.status_combo = QComboBox()
        self.status_combo.addItems([s.value for s in OrderStatus])
        if self.order:
            self.status_combo.setCurrentText(self.order.status)
        info_layout.addWidget(self.status_combo, 1, 3)
        
        info_layout.addWidget(QLabel("备注:"), 2, 0)
        self.notes_text = QTextEdit()
        self.notes_text.setMaximumHeight(80)
        if self.order and self.order.notes:
            self.notes_text.setPlainText(self.order.notes)
        info_layout.addWidget(self.notes_text, 2, 1, 1, 3)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # 物料列表
        materials_group = QGroupBox("物料列表")
        materials_layout = QVBoxLayout()
        
        self.materials_table = QTableWidget()
        self.materials_table.setColumnCount(5)
        self.materials_table.setHorizontalHeaderLabels(["物料名称", "类别", "数量", "单位", "备注"])
        self.materials_table.horizontalHeader().setStretchLastSection(True)
        materials_layout.addWidget(self.materials_table)
        
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("添加物料")
        add_btn.clicked.connect(self._add_material)
        edit_btn = QPushButton("编辑物料")
        edit_btn.clicked.connect(self._edit_material)
        remove_btn = QPushButton("删除物料")
        remove_btn.clicked.connect(self._remove_material)
        
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(remove_btn)
        materials_layout.addLayout(btn_layout)
        
        materials_group.setLayout(materials_layout)
        layout.addWidget(materials_group)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _add_material(self):
        pass
    
    def _edit_material(self):
        pass
    
    def _remove_material(self):
        pass
    
    def _save(self):
        if not self.requester_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入申请人")
            return
        
        order = Order(
            id=self.order.id if self.order else None,
            order_number=self.order.order_number if self.order else "",
            requester=self.requester_edit.text().strip(),
            department=self.department_edit.text().strip(),
            status=self.status_combo.currentText(),
            priority=self.priority_combo.currentText(),
            notes=self.notes_text.toPlainText().strip(),
            materials=self.materials
        )
        
        self.result = order
        self.accept()


class MaterialCard(QFrame):
    """物料卡片"""
    
    clicked = pyqtSignal(int)  # material_id
    
    def __init__(self, material: Material, parent=None):
        super().__init__(parent)
        self.material = material
        self.setFrameStyle(QFrame.Box)
        self.setLineWidth(2)
        self.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 5px;
            }
        """)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QHBoxLayout()
        self.setLayout(layout)
        
        # 左侧图片
        img_label = QLabel()
        if self.material.images and len(self.material.images) > 0:
            try:
                img_bytes = self.material.images[0]
                if isinstance(img_bytes, bytes):
                    img = QImage.fromData(img_bytes)
                    pixmap = QPixmap.fromImage(img)
                    img_label.setPixmap(pixmap.scaled(120, 120, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                img_label.setAlignment(Qt.AlignCenter)
            except:
                img_label.setText("📷\n无图片")
                img_label.setAlignment(Qt.AlignCenter)
        else:
            img_label.setText("📷\n无图片")
            img_label.setAlignment(Qt.AlignCenter)
        
        img_label.setFixedSize(150, 150)
        layout.addWidget(img_label)
        
        # 右侧信息
        info_layout = QVBoxLayout()
        
        # 标题
        title_layout = QHBoxLayout()
        name_label = QLabel(self.material.name)
        name_label.setFont(QFont("Microsoft YaHei", 16, QFont.Bold))
        title_layout.addWidget(name_label)
        
        id_label = QLabel(f"ID: {self.material.id}")
        id_label.setStyleSheet("background-color: #e9ecef; padding: 5px; border-radius: 3px;")
        title_layout.addWidget(id_label)
        info_layout.addLayout(title_layout)
        
        # 类别
        category_colors = {
            "试剂": "#28a745",
            "耗材": "#17a2b8",
            "设备": "#ffc107",
            "工具": "#fd7e14",
            "其他": "#6c757d"
        }
        category_color = category_colors.get(self.material.category, "#6c757d")
        category_label = QLabel(self.material.category)
        category_label.setStyleSheet(f"background-color: {category_color}; color: white; padding: 5px; border-radius: 3px;")
        category_label.setFixedWidth(80)
        info_layout.addWidget(category_label)
        
        # 信息
        info_text = f"数量: {self.material.quantity} {self.material.unit}"
        if self.material.quantity <= self.material.min_stock:
            info_text += f" ⚠️ 库存不足"
        info_label = QLabel(info_text)
        info_layout.addWidget(info_label)
        
        if self.material.location:
            location_label = QLabel(f"📍 {self.material.location}")
            info_layout.addWidget(location_label)
        
        if self.material.supplier:
            supplier_label = QLabel(f"🏢 {self.material.supplier}")
            info_layout.addWidget(supplier_label)
        
        layout.addLayout(info_layout)
        
        # 鼠标点击事件
        self.mousePressEvent = self._on_click
    
    def _on_click(self, event):
        self.clicked.emit(self.material.id)
    
    def set_selected(self, selected: bool):
        """设置选中状态"""
        if selected:
            self.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border: 3px solid #28a745;
                    border-radius: 5px;
                }
            """)
        else:
            self.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border-radius: 5px;
                }
            """)


class MaterialDetailPanel(QWidget):
    """物料详情面板"""
    
    edit_requested = pyqtSignal(int)  # material_id
    delete_requested = pyqtSignal(int)  # material_id
    
    def __init__(self, material: Material, parent=None):
        super().__init__(parent)
        self.material = material
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        scroll_layout = QVBoxLayout()
        content.setLayout(scroll_layout)
        
        # 物料名称
        name_label = QLabel(self.material.name)
        name_label.setFont(QFont("Microsoft YaHei", 18, QFont.Bold))
        name_label.setAlignment(Qt.AlignCenter)
        scroll_layout.addWidget(name_label)
        
        # 类别
        category_colors = {
            "试剂": "#28a745",
            "耗材": "#17a2b8",
            "设备": "#ffc107",
            "工具": "#fd7e14",
            "其他": "#6c757d"
        }
        category_color = category_colors.get(self.material.category, "#6c757d")
        category_label = QLabel(self.material.category)
        category_label.setStyleSheet(f"background-color: {category_color}; color: white; padding: 10px; border-radius: 5px;")
        category_label.setAlignment(Qt.AlignCenter)
        scroll_layout.addWidget(category_label)
        
        # 基本信息
        info_group = QGroupBox("基本信息")
        info_layout = QVBoxLayout()
        info_layout.addWidget(QLabel(f"ID: {self.material.id}"))
        info_layout.addWidget(QLabel(f"数量: {self.material.quantity} {self.material.unit}"))
        
        min_stock_text = f"最低库存: {self.material.min_stock}"
        if self.material.quantity <= self.material.min_stock:
            min_stock_label = QLabel(min_stock_text)
            min_stock_label.setStyleSheet("color: #dc3545;")
            info_layout.addWidget(min_stock_label)
        else:
            info_layout.addWidget(QLabel(min_stock_text))
        
        if self.material.location:
            info_layout.addWidget(QLabel(f"📍 位置: {self.material.location}"))
        if self.material.supplier:
            info_layout.addWidget(QLabel(f"🏢 供应商: {self.material.supplier}"))
        
        info_group.setLayout(info_layout)
        scroll_layout.addWidget(info_group)
        
        # 描述
        if self.material.description:
            desc_group = QGroupBox("描述")
            desc_layout = QVBoxLayout()
            desc_text = QTextEdit()
            desc_text.setPlainText(self.material.description)
            desc_text.setReadOnly(True)
            desc_text.setMaximumHeight(100)
            desc_layout.addWidget(desc_text)
            desc_group.setLayout(desc_layout)
            scroll_layout.addWidget(desc_group)
        
        # 图片
        if self.material.images and len(self.material.images) > 0:
            img_group = QGroupBox("图片")
            img_layout = QVBoxLayout()
            
            max_images = 3
            for idx, img_bytes in enumerate(self.material.images[:max_images]):
                if isinstance(img_bytes, bytes):
                    try:
                        img = QImage.fromData(img_bytes)
                        pixmap = QPixmap.fromImage(img)
                        label = QLabel()
                        label.setPixmap(pixmap.scaled(200, 200, Qt.KeepAspectRatio, Qt.SmoothTransformation))
                        label.setAlignment(Qt.AlignCenter)
                        img_layout.addWidget(label)
                    except:
                        pass
            
            if len(self.material.images) > max_images:
                img_layout.addWidget(QLabel(f"...还有 {len(self.material.images) - max_images} 张图片"))
            
            img_group.setLayout(img_layout)
            scroll_layout.addWidget(img_group)
        
        scroll_layout.addStretch()
        
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # 按钮
        btn_layout = QVBoxLayout()
        edit_btn = QPushButton("编辑物料")
        edit_btn.clicked.connect(lambda: self.edit_requested.emit(self.material.id))
        delete_btn = QPushButton("删除物料")
        delete_btn.clicked.connect(lambda: self.delete_requested.emit(self.material.id))
        
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        layout.addLayout(btn_layout)


# ==================== ADC 相关UI组件 ====================

class ADCSpecDialog(QDialog):
    """ADC规格编辑对话框"""
    
    def __init__(self, parent=None, spec: Optional[ADCSpec] = None, preset_specs: List[float] = None):
        super().__init__(parent)
        self.spec = spec
        self.preset_specs = preset_specs or PRESET_SPECS
        self.result = None
        
        self.setWindowTitle("编辑规格" if spec else "添加规格")
        self.setFixedSize(400, 200)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 规格选择
        spec_layout = QHBoxLayout()
        spec_layout.addWidget(QLabel("规格 (mg):"))
        
        self.spec_combo = QComboBox()
        self.spec_combo.setEditable(True)
        for preset in self.preset_specs:
            self.spec_combo.addItem(f"{preset}")
        if self.spec:
            self.spec_combo.setCurrentText(f"{self.spec.spec_mg}")
        spec_layout.addWidget(self.spec_combo)
        layout.addLayout(spec_layout)
        
        # 数量
        qty_layout = QHBoxLayout()
        qty_layout.addWidget(QLabel("数量 (小管数):"))
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setMaximum(999999)
        self.quantity_spin.setValue(self.spec.quantity if self.spec else 0)
        qty_layout.addWidget(self.quantity_spin)
        layout.addLayout(qty_layout)
        
        layout.addStretch()
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _save(self):
        try:
            spec_mg = float(self.spec_combo.currentText())
        except ValueError:
            QMessageBox.critical(self, "错误", "请输入有效的规格数值")
            return
        
        if spec_mg <= 0:
            QMessageBox.critical(self, "错误", "规格必须大于0")
            return
        
        if self.quantity_spin.value() < 0:
            QMessageBox.critical(self, "错误", "数量不能为负数")
            return
        
        self.result = {
            'spec_mg': spec_mg,
            'quantity': self.quantity_spin.value()
        }
        self.accept()


class ADCDialog(QDialog):
    """ADC编辑对话框"""
    
    def __init__(self, parent=None, adc: Optional[ADC] = None):
        super().__init__(parent)
        self.adc = adc
        self.result = None
        self.specs = []  # 规格列表
        
        if adc and adc.specs:
            self.specs = [{'spec_mg': s.spec_mg, 'quantity': s.quantity} 
                         if isinstance(s, ADCSpec) else s for s in adc.specs]
        
        self.setWindowTitle("编辑ADC" if adc else "添加ADC")
        self.setFixedSize(700, 700)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        scroll_layout = QVBoxLayout()
        content.setLayout(scroll_layout)
        
        # 基本信息
        info_group = QGroupBox("基本信息")
        info_layout = QGridLayout()
        
        # Lot Number
        info_layout.addWidget(QLabel("Lot Number *:"), 0, 0)
        self.lot_number_edit = QLineEdit(self.adc.lot_number if self.adc else "")
        info_layout.addWidget(self.lot_number_edit, 0, 1)
        
        # Sample ID
        info_layout.addWidget(QLabel("Sample ID *:"), 1, 0)
        self.sample_id_edit = QLineEdit(self.adc.sample_id if self.adc else "")
        info_layout.addWidget(self.sample_id_edit, 1, 1)
        
        # Owner
        info_layout.addWidget(QLabel("Owner:"), 2, 0)
        self.owner_edit = QLineEdit(self.adc.owner if self.adc else "")
        info_layout.addWidget(self.owner_edit, 2, 1)
        
        # Concentration
        info_layout.addWidget(QLabel("Concentration (mg/mL):"), 3, 0)
        self.concentration_spin = QDoubleSpinBox()
        self.concentration_spin.setMaximum(9999.99)
        self.concentration_spin.setDecimals(2)
        self.concentration_spin.setValue(self.adc.concentration if self.adc else 0.0)
        info_layout.addWidget(self.concentration_spin, 3, 1)
        
        # Storage Temp
        info_layout.addWidget(QLabel("Storage Temp:"), 4, 0)
        self.storage_temp_combo = QComboBox()
        self.storage_temp_combo.setEditable(True)
        self.storage_temp_combo.addItems(["-80°C", "-20°C", "4°C", "RT"])
        if self.adc and self.adc.storage_temp:
            self.storage_temp_combo.setCurrentText(self.adc.storage_temp)
        info_layout.addWidget(self.storage_temp_combo, 4, 1)
        
        # Storage Position
        info_layout.addWidget(QLabel("Storage Position:"), 5, 0)
        self.storage_position_edit = QLineEdit(self.adc.storage_position if self.adc else "")
        info_layout.addWidget(self.storage_position_edit, 5, 1)
        
        # Antibody
        info_layout.addWidget(QLabel("Antibody:"), 6, 0)
        self.antibody_edit = QLineEdit(self.adc.antibody if self.adc else "")
        info_layout.addWidget(self.antibody_edit, 6, 1)
        
        # Linker-payload
        info_layout.addWidget(QLabel("Linker-payload:"), 7, 0)
        self.linker_payload_edit = QLineEdit(self.adc.linker_payload if self.adc else "")
        info_layout.addWidget(self.linker_payload_edit, 7, 1)
        
        # Description
        info_layout.addWidget(QLabel("Description:"), 8, 0)
        self.desc_text = QTextEdit()
        self.desc_text.setMaximumHeight(80)
        if self.adc and self.adc.description:
            self.desc_text.setPlainText(self.adc.description)
        info_layout.addWidget(self.desc_text, 8, 1)
        
        info_group.setLayout(info_layout)
        scroll_layout.addWidget(info_group)
        
        # 规格管理
        specs_group = QGroupBox("规格库存")
        specs_layout = QVBoxLayout()
        
        self.specs_table = QTableWidget()
        self.specs_table.setColumnCount(3)
        self.specs_table.setHorizontalHeaderLabels(["规格 (mg)", "数量 (小管)", "小计 (mg)"])
        self.specs_table.horizontalHeader().setStretchLastSection(True)
        self.specs_table.setEditTriggers(QTableWidget.NoEditTriggers)  # 设为只读
        self.specs_table.setSelectionBehavior(QTableWidget.SelectRows)  # 选中整行
        specs_layout.addWidget(self.specs_table)
        
        spec_btn_layout = QHBoxLayout()
        add_spec_btn = QPushButton("添加规格")
        add_spec_btn.clicked.connect(self._add_spec)
        edit_spec_btn = QPushButton("编辑规格")
        edit_spec_btn.clicked.connect(self._edit_spec)
        remove_spec_btn = QPushButton("删除规格")
        remove_spec_btn.clicked.connect(self._remove_spec)
        
        spec_btn_layout.addWidget(add_spec_btn)
        spec_btn_layout.addWidget(edit_spec_btn)
        spec_btn_layout.addWidget(remove_spec_btn)
        specs_layout.addLayout(spec_btn_layout)
        
        # 汇总显示（必须在调用_refresh_specs_table之前创建）
        self.total_label = QLabel()
        self.total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #007bff;")
        specs_layout.addWidget(self.total_label)
        
        # 刷新表格和汇总
        self._refresh_specs_table()
        
        specs_group.setLayout(specs_layout)
        scroll_layout.addWidget(specs_group)
        
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _refresh_specs_table(self):
        """刷新规格表格"""
        self.specs_table.setRowCount(len(self.specs))
        for row, spec in enumerate(self.specs):
            spec_mg = spec['spec_mg']
            quantity = spec['quantity']
            subtotal = spec_mg * quantity
            
            self.specs_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
            self.specs_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
            self.specs_table.setItem(row, 2, QTableWidgetItem(f"{subtotal:.2f}"))
        
        self._update_total_label()
    
    def _update_total_label(self):
        """更新汇总标签"""
        total_mg = sum(s['spec_mg'] * s['quantity'] for s in self.specs)
        total_vials = sum(s['quantity'] for s in self.specs)
        self.total_label.setText(f"汇总: {total_vials} 个小管, 共计 {total_mg:.2f} mg")
    
    def _add_spec(self):
        """添加规格"""
        dialog = ADCSpecDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.specs.append(dialog.result)
            self._refresh_specs_table()
    
    def _edit_spec(self):
        """编辑规格"""
        current_row = self.specs_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要编辑的规格")
            return
        
        spec_data = self.specs[current_row]
        spec = ADCSpec(spec_mg=spec_data['spec_mg'], quantity=spec_data['quantity'])
        
        dialog = ADCSpecDialog(self, spec)
        if dialog.exec_() == QDialog.Accepted:
            self.specs[current_row] = dialog.result
            self._refresh_specs_table()
    
    def _remove_spec(self):
        """删除规格"""
        current_row = self.specs_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要删除的规格")
            return
        
        if QMessageBox.question(self, "确认", "确定要删除这个规格吗？") == QMessageBox.Yes:
            del self.specs[current_row]
            self._refresh_specs_table()
    
    def _save(self):
        """保存"""
        if not self.lot_number_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入Lot Number")
            return
        
        if not self.sample_id_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入Sample ID")
            return
        
        adc = ADC(
            id=self.adc.id if self.adc else None,
            lot_number=self.lot_number_edit.text().strip(),
            sample_id=self.sample_id_edit.text().strip(),
            description=self.desc_text.toPlainText().strip(),
            concentration=self.concentration_spin.value(),
            owner=self.owner_edit.text().strip(),
            storage_temp=self.storage_temp_combo.currentText(),
            storage_position=self.storage_position_edit.text().strip(),
            antibody=self.antibody_edit.text().strip(),
            linker_payload=self.linker_payload_edit.text().strip(),
            specs=[ADCSpec(spec_mg=s['spec_mg'], quantity=s['quantity']) for s in self.specs]
        )
        
        self.result = adc
        self.accept()


class ADCCard(QFrame):
    """ADC卡片"""
    
    clicked = pyqtSignal(int)  # adc_id
    
    def __init__(self, adc: ADC, parent=None):
        super().__init__(parent)
        self.adc = adc
        self.setFrameStyle(QFrame.Box)
        self.setLineWidth(2)
        self.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 5px;
            }
        """)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 标题行
        title_layout = QHBoxLayout()
        
        lot_label = QLabel(f"Lot#: {self.adc.lot_number}")
        lot_label.setFont(QFont("Microsoft YaHei", 12, QFont.Bold))
        title_layout.addWidget(lot_label)
        
        title_layout.addStretch()
        layout.addLayout(title_layout)
        
        # Sample ID
        sample_label = QLabel(f"Sample ID: {self.adc.sample_id}")
        sample_label.setStyleSheet("color: #6c757d;")
        layout.addWidget(sample_label)
        
        # Owner
        if self.adc.owner:
            owner_label = QLabel(f"👤 {self.adc.owner}")
            layout.addWidget(owner_label)
        
        # 存储信息
        storage_info = []
        if self.adc.storage_temp:
            storage_info.append(self.adc.storage_temp)
        if self.adc.storage_position:
            storage_info.append(self.adc.storage_position)
        if storage_info:
            storage_label = QLabel(f"📍 {' / '.join(storage_info)}")
            layout.addWidget(storage_label)
        
        # 汇总信息
        total_mg = self.adc.get_total_mg()
        total_vials = self.adc.get_total_vials()
        summary_label = QLabel(f"📦 {total_vials} 管 | 总量: {total_mg:.2f} mg")
        summary_label.setStyleSheet("font-weight: bold; color: #007bff;")
        layout.addWidget(summary_label)
        
        # 鼠标点击事件
        self.mousePressEvent = self._on_click
    
    def _on_click(self, event):
        self.clicked.emit(self.adc.id)
    
    def set_selected(self, selected: bool):
        """设置选中状态"""
        if selected:
            self.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border: 3px solid #007bff;
                    border-radius: 5px;
                }
            """)
        else:
            self.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border-radius: 5px;
                }
            """)


class ADCDetailPanel(QWidget):
    """ADC详情面板"""
    
    edit_requested = pyqtSignal(int)  # adc_id
    delete_requested = pyqtSignal(int)  # adc_id
    
    def __init__(self, adc: ADC, parent=None):
        super().__init__(parent)
        self.adc = adc
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        scroll_layout = QVBoxLayout()
        content.setLayout(scroll_layout)
        
        # Lot Number 标题
        lot_label = QLabel(f"Lot#: {self.adc.lot_number}")
        lot_label.setFont(QFont("Microsoft YaHei", 18, QFont.Bold))
        lot_label.setAlignment(Qt.AlignCenter)
        scroll_layout.addWidget(lot_label)
        
        # Sample ID
        sample_label = QLabel(f"Sample ID: {self.adc.sample_id}")
        sample_label.setFont(QFont("Microsoft YaHei", 14))
        sample_label.setAlignment(Qt.AlignCenter)
        sample_label.setStyleSheet("color: #6c757d;")
        scroll_layout.addWidget(sample_label)
        
        # 基本信息
        info_group = QGroupBox("基本信息")
        info_layout = QVBoxLayout()
        
        # 始终显示所有属性，缺失时显示"N/A"
        missing_style = "color: #000000; font-style: italic;"
        
        owner_text = self.adc.owner if self.adc.owner else '<span style="' + missing_style + '">N/A</span>'
        owner_label = QLabel(f"👤 Owner: {owner_text}")
        owner_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(owner_label)
        
        concentration_text = f"{self.adc.concentration} mg/mL" if self.adc.concentration > 0 else '<span style="' + missing_style + '">N/A</span>'
        concentration_label = QLabel(f"💉 Concentration: {concentration_text}")
        concentration_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(concentration_label)
        
        storage_temp_text = self.adc.storage_temp if self.adc.storage_temp else '<span style="' + missing_style + '">N/A</span>'
        storage_temp_label = QLabel(f"🌡️ Storage Temp: {storage_temp_text}")
        storage_temp_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(storage_temp_label)
        
        storage_position_text = self.adc.storage_position if self.adc.storage_position else '<span style="' + missing_style + '">N/A</span>'
        storage_position_label = QLabel(f"📍 Storage Position: {storage_position_text}")
        storage_position_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(storage_position_label)
        
        antibody_text = self.adc.antibody if self.adc.antibody else '<span style="' + missing_style + '">N/A</span>'
        antibody_label = QLabel(f"🧬 Antibody: {antibody_text}")
        antibody_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(antibody_label)
        
        linker_payload_text = self.adc.linker_payload if self.adc.linker_payload else '<span style="' + missing_style + '">N/A</span>'
        linker_payload_label = QLabel(f"🔗 Linker-payload: {linker_payload_text}")
        linker_payload_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(linker_payload_label)
        
        if self.adc.created_at:
            created_str = self.adc.created_at.strftime('%Y-%m-%d %H:%M') if isinstance(self.adc.created_at, datetime) else str(self.adc.created_at)
        else:
            created_str = '<span style="' + missing_style + '">N/A</span>'
        created_label = QLabel(f"📅 入库时间: {created_str}")
        created_label.setTextFormat(Qt.RichText)
        info_layout.addWidget(created_label)
        
        info_group.setLayout(info_layout)
        scroll_layout.addWidget(info_group)
        
        # 描述
        if self.adc.description:
            desc_group = QGroupBox("描述")
            desc_layout = QVBoxLayout()
            desc_text = QTextEdit()
            desc_text.setPlainText(self.adc.description)
            desc_text.setReadOnly(True)
            desc_text.setMaximumHeight(80)
            desc_layout.addWidget(desc_text)
            desc_group.setLayout(desc_layout)
            scroll_layout.addWidget(desc_group)
        
        # 规格列表
        specs_group = QGroupBox("规格库存")
        specs_layout = QVBoxLayout()
        
        specs_table = QTableWidget()
        specs_table.setColumnCount(3)
        specs_table.setHorizontalHeaderLabels(["规格 (mg)", "数量 (小管)", "小计 (mg)"])
        specs_table.horizontalHeader().setStretchLastSection(True)
        specs_table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        specs = self.adc.specs if self.adc.specs else []
        specs_table.setRowCount(len(specs))
        
        for row, spec in enumerate(specs):
            if isinstance(spec, ADCSpec):
                spec_mg = spec.spec_mg
                quantity = spec.quantity
            else:
                spec_mg = spec.get('spec_mg', 0)
                quantity = spec.get('quantity', 0)
            
            subtotal = spec_mg * quantity
            specs_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
            specs_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
            specs_table.setItem(row, 2, QTableWidgetItem(f"{subtotal:.2f}"))
        
        specs_layout.addWidget(specs_table)
        
        # 汇总
        total_mg = self.adc.get_total_mg()
        total_vials = self.adc.get_total_vials()
        total_label = QLabel(f"汇总: {total_vials} 个小管, 共计 {total_mg:.2f} mg")
        total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #007bff;")
        specs_layout.addWidget(total_label)
        
        specs_group.setLayout(specs_layout)
        scroll_layout.addWidget(specs_group)
        
        scroll_layout.addStretch()
        
        scroll.setWidget(content)
        layout.addWidget(scroll)
        
        # 按钮
        btn_layout = QVBoxLayout()
        edit_btn = QPushButton("编辑ADC")
        edit_btn.clicked.connect(lambda: self.edit_requested.emit(self.adc.id))
        delete_btn = QPushButton("删除ADC")
        delete_btn.clicked.connect(lambda: self.delete_requested.emit(self.adc.id))
        
        btn_layout.addWidget(edit_btn)
        btn_layout.addWidget(delete_btn)
        layout.addLayout(btn_layout)


# ==================== ADC 出入库对话框 ====================

class ADCMovementItemDialog(QDialog):
    """出入库明细编辑对话框"""
    
    def __init__(self, parent=None, item: Optional[Dict] = None, preset_specs: List[float] = None):
        super().__init__(parent)
        self.item = item
        self.preset_specs = preset_specs or PRESET_SPECS
        self.result = None
        
        self.setWindowTitle("编辑明细" if item else "添加明细")
        self.setFixedSize(400, 180)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 规格选择
        spec_layout = QHBoxLayout()
        spec_layout.addWidget(QLabel("规格 (mg):"))
        self.spec_combo = QComboBox()
        self.spec_combo.setEditable(True)
        for preset in self.preset_specs:
            self.spec_combo.addItem(f"{preset}")
        if self.item:
            self.spec_combo.setCurrentText(f"{self.item.get('spec_mg', '')}")
        spec_layout.addWidget(self.spec_combo)
        layout.addLayout(spec_layout)
        
        # 数量
        qty_layout = QHBoxLayout()
        qty_layout.addWidget(QLabel("数量 (小管数):"))
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setMinimum(1)
        self.quantity_spin.setMaximum(999999)
        self.quantity_spin.setValue(self.item.get('quantity', 1) if self.item else 1)
        qty_layout.addWidget(self.quantity_spin)
        layout.addLayout(qty_layout)
        
        layout.addStretch()
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _save(self):
        try:
            spec_mg = float(self.spec_combo.currentText())
        except ValueError:
            QMessageBox.critical(self, "错误", "请输入有效的规格数值")
            return
        
        if spec_mg <= 0:
            QMessageBox.critical(self, "错误", "规格必须大于0")
            return
        
        self.result = {
            'spec_mg': spec_mg,
            'quantity': self.quantity_spin.value()
        }
        self.accept()


class ADCOutboundDialog(QDialog):
    """ADC出库对话框"""
    
    def __init__(self, parent=None, adc_controller: ADCController = None):
        super().__init__(parent)
        self.adc_controller = adc_controller
        self.result = None
        self.items = []  # 出库明细列表
        
        self.setWindowTitle("ADC出库")
        self.setFixedSize(800, 700)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 基本信息
        info_group = QGroupBox("出库信息")
        info_layout = QGridLayout()
        
        # Lot Number选择
        info_layout.addWidget(QLabel("Lot Number *:"), 0, 0)
        self.lot_combo = QComboBox()
        self.lot_combo.setEditable(True)
        self.lot_combo.currentTextChanged.connect(self._on_lot_changed)
        if self.adc_controller:
            for adc in self.adc_controller.get_all_adcs():
                self.lot_combo.addItem(adc.lot_number)
        info_layout.addWidget(self.lot_combo, 0, 1)
        
        # 需求人
        info_layout.addWidget(QLabel("需求人 *:"), 1, 0)
        self.requester_edit = QLineEdit()
        info_layout.addWidget(self.requester_edit, 1, 1)
        
        # 出库人
        info_layout.addWidget(QLabel("出库人 *:"), 2, 0)
        self.operator_edit = QLineEdit()
        info_layout.addWidget(self.operator_edit, 2, 1)
        
        # 寄送地址
        info_layout.addWidget(QLabel("寄送地址:"), 3, 0)
        self.address_edit = QLineEdit()
        info_layout.addWidget(self.address_edit, 3, 1)
        
        # 备注
        info_layout.addWidget(QLabel("备注:"), 4, 0)
        self.notes_edit = QLineEdit()
        info_layout.addWidget(self.notes_edit, 4, 1)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # 当前库存（只读显示）
        stock_group = QGroupBox("当前库存")
        stock_layout = QVBoxLayout()
        
        self.stock_table = QTableWidget()
        self.stock_table.setColumnCount(3)
        self.stock_table.setHorizontalHeaderLabels(["规格 (mg)", "当前库存 (小管)", "库存 (mg)"])
        self.stock_table.horizontalHeader().setStretchLastSection(True)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectRows)  # 整行选择
        stock_layout.addWidget(self.stock_table)
        
        stock_group.setLayout(stock_layout)
        layout.addWidget(stock_group)
        
        # 出库明细
        items_group = QGroupBox("出库明细")
        items_layout = QVBoxLayout()
        
        self.items_table = QTableWidget()
        self.items_table.setColumnCount(3)
        self.items_table.setHorizontalHeaderLabels(["规格 (mg)", "数量 (小管)", "小计 (mg)"])
        self.items_table.horizontalHeader().setStretchLastSection(True)
        self.items_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.items_table.setSelectionBehavior(QTableWidget.SelectRows)  # 整行选择
        items_layout.addWidget(self.items_table)
        
        item_btn_layout = QHBoxLayout()
        add_btn = QPushButton("添加")
        add_btn.clicked.connect(self._add_item)
        edit_btn = QPushButton("编辑")
        edit_btn.clicked.connect(self._edit_item)
        remove_btn = QPushButton("删除")
        remove_btn.clicked.connect(self._remove_item)
        item_btn_layout.addWidget(add_btn)
        item_btn_layout.addWidget(edit_btn)
        item_btn_layout.addWidget(remove_btn)
        items_layout.addLayout(item_btn_layout)
        
        self.total_label = QLabel("合计: 0 个小管, 0.00 mg")
        self.total_label.setStyleSheet("font-weight: bold; color: #dc3545;")
        items_layout.addWidget(self.total_label)
        
        items_group.setLayout(items_layout)
        layout.addWidget(items_group)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        # 初始化库存显示
        self._on_lot_changed(self.lot_combo.currentText())
    
    def _on_lot_changed(self, lot_number: str):
        """Lot Number变更时，更新库存显示"""
        if not hasattr(self, 'stock_table'):
            return
        
        self.stock_table.setRowCount(0)
        if self.adc_controller and lot_number:
            adc = self.adc_controller.get_adc_by_lot_number(lot_number)
            if adc and adc.specs:
                self.stock_table.setRowCount(len(adc.specs))
                for row, spec in enumerate(adc.specs):
                    spec_mg = spec.spec_mg if isinstance(spec, ADCSpec) else spec.get('spec_mg', 0)
                    quantity = spec.quantity if isinstance(spec, ADCSpec) else spec.get('quantity', 0)
                    total = spec_mg * quantity
                    
                    self.stock_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
                    self.stock_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
                    self.stock_table.setItem(row, 2, QTableWidgetItem(f"{total:.2f}"))
        
        # 动态调整表格高度：显示所有行但不超过最大高度
        row_count = self.stock_table.rowCount()
        row_height = 30  # 每行高度
        header_height = 30  # 表头高度
        max_height = 180  # 最大高度
        min_height = 60  # 最小高度
        
        if row_count > 0:
            calculated_height = header_height + row_count * row_height + 5
            table_height = min(max(calculated_height, min_height), max_height)
        else:
            table_height = min_height
        
        self.stock_table.setFixedHeight(table_height)
    
    def _refresh_items_table(self):
        self.items_table.setRowCount(len(self.items))
        total_mg = 0.0
        total_vials = 0
        for row, item in enumerate(self.items):
            spec_mg = item['spec_mg']
            quantity = item['quantity']
            subtotal = spec_mg * quantity
            total_mg += subtotal
            total_vials += quantity
            
            self.items_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
            self.items_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
            self.items_table.setItem(row, 2, QTableWidgetItem(f"{subtotal:.2f}"))
        
        self.total_label.setText(f"合计: {total_vials} 个小管, {total_mg:.2f} mg")
    
    def _get_available_specs(self) -> List[float]:
        """获取当前选中LotNumber的可用规格列表"""
        lot_number = self.lot_combo.currentText().strip()
        if self.adc_controller and lot_number:
            adc = self.adc_controller.get_adc_by_lot_number(lot_number)
            if adc and adc.specs:
                specs = []
                for spec in adc.specs:
                    spec_mg = spec.spec_mg if isinstance(spec, ADCSpec) else spec.get('spec_mg', 0)
                    quantity = spec.quantity if isinstance(spec, ADCSpec) else spec.get('quantity', 0)
                    if quantity > 0:  # 只显示有库存的规格
                        specs.append(spec_mg)
                return specs
        return PRESET_SPECS  # 如果没有可用规格，返回默认列表
    
    def _add_item(self):
        available_specs = self._get_available_specs()
        dialog = ADCMovementItemDialog(self, preset_specs=available_specs)
        if dialog.exec_() == QDialog.Accepted:
            self.items.append(dialog.result)
            self._refresh_items_table()
    
    def _edit_item(self):
        current_row = self.items_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要编辑的明细")
            return
        
        available_specs = self._get_available_specs()
        dialog = ADCMovementItemDialog(self, self.items[current_row], preset_specs=available_specs)
        if dialog.exec_() == QDialog.Accepted:
            self.items[current_row] = dialog.result
            self._refresh_items_table()
    
    def _remove_item(self):
        current_row = self.items_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要删除的明细")
            return
        
        del self.items[current_row]
        self._refresh_items_table()
    
    def _save(self):
        if not self.lot_combo.currentText().strip():
            QMessageBox.critical(self, "错误", "请选择Lot Number")
            return
        if not self.requester_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入需求人")
            return
        if not self.operator_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入出库人")
            return
        if not self.items:
            QMessageBox.critical(self, "错误", "请添加出库明细")
            return
        
        outbound = ADCOutbound(
            lot_number=self.lot_combo.currentText().strip(),
            requester=self.requester_edit.text().strip(),
            operator=self.operator_edit.text().strip(),
            shipping_address=self.address_edit.text().strip(),
            shipping_date=datetime.now(),  # 使用当前时间
            notes=self.notes_edit.text().strip(),
            items=[ADCMovementItem(spec_mg=i['spec_mg'], quantity=i['quantity']) for i in self.items]
        )
        
        self.result = outbound
        self.accept()


class ADCInboundDialog(QDialog):
    """ADC入库对话框"""
    
    def __init__(self, parent=None, adc_controller: ADCController = None):
        super().__init__(parent)
        self.adc_controller = adc_controller
        self.result = None
        self.items = []  # 入库明细列表
        
        self.setWindowTitle("ADC入库")
        self.setFixedSize(800, 700)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        # 基本信息
        info_group = QGroupBox("入库信息")
        info_layout = QGridLayout()
        
        # Lot Number选择
        info_layout.addWidget(QLabel("Lot Number *:"), 0, 0)
        self.lot_combo = QComboBox()
        self.lot_combo.setEditable(True)
        self.lot_combo.currentTextChanged.connect(self._on_lot_changed)
        if self.adc_controller:
            for adc in self.adc_controller.get_all_adcs():
                self.lot_combo.addItem(adc.lot_number)
        info_layout.addWidget(self.lot_combo, 0, 1)
        
        # 入库人
        info_layout.addWidget(QLabel("入库人 *:"), 1, 0)
        self.operator_edit = QLineEdit()
        self.operator_edit.textChanged.connect(self._on_operator_changed)
        info_layout.addWidget(self.operator_edit, 1, 1)
        
        # Owner
        info_layout.addWidget(QLabel("Owner:"), 2, 0)
        self.owner_edit = QLineEdit()
        info_layout.addWidget(self.owner_edit, 2, 1)
        
        # 存放地址
        info_layout.addWidget(QLabel("存放地址:"), 3, 0)
        self.position_edit = QLineEdit()
        info_layout.addWidget(self.position_edit, 3, 1)
        
        # 备注
        info_layout.addWidget(QLabel("备注:"), 4, 0)
        self.notes_edit = QLineEdit()
        info_layout.addWidget(self.notes_edit, 4, 1)
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # 当前库存（只读显示）
        stock_group = QGroupBox("当前库存")
        stock_layout = QVBoxLayout()
        
        self.stock_table = QTableWidget()
        self.stock_table.setColumnCount(3)
        self.stock_table.setHorizontalHeaderLabels(["规格 (mg)", "当前库存 (小管)", "库存 (mg)"])
        self.stock_table.horizontalHeader().setStretchLastSection(True)
        self.stock_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectRows)  # 整行选择
        stock_layout.addWidget(self.stock_table)
        
        stock_group.setLayout(stock_layout)
        layout.addWidget(stock_group)
        
        # 入库明细
        items_group = QGroupBox("入库明细")
        items_layout = QVBoxLayout()
        
        self.items_table = QTableWidget()
        self.items_table.setColumnCount(3)
        self.items_table.setHorizontalHeaderLabels(["规格 (mg)", "数量 (小管)", "小计 (mg)"])
        self.items_table.horizontalHeader().setStretchLastSection(True)
        self.items_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.items_table.setSelectionBehavior(QTableWidget.SelectRows)  # 整行选择
        items_layout.addWidget(self.items_table)
        
        item_btn_layout = QHBoxLayout()
        add_btn = QPushButton("添加")
        add_btn.clicked.connect(self._add_item)
        edit_btn = QPushButton("编辑")
        edit_btn.clicked.connect(self._edit_item)
        remove_btn = QPushButton("删除")
        remove_btn.clicked.connect(self._remove_item)
        item_btn_layout.addWidget(add_btn)
        item_btn_layout.addWidget(edit_btn)
        item_btn_layout.addWidget(remove_btn)
        items_layout.addLayout(item_btn_layout)
        
        self.total_label = QLabel("合计: 0 个小管, 0.00 mg")
        self.total_label.setStyleSheet("font-weight: bold; color: #28a745;")
        items_layout.addWidget(self.total_label)
        
        items_group.setLayout(items_layout)
        layout.addWidget(items_group)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        # 初始化存放地址
        self._on_lot_changed(self.lot_combo.currentText())
    
    def _on_lot_changed(self, lot_number: str):
        """Lot Number变更时，自动填充存放地址和更新库存显示"""
        # 检查 UI 元素是否已创建（避免初始化时的信号触发）
        if not hasattr(self, 'position_edit') or not hasattr(self, 'stock_table'):
            return
        
        self.stock_table.setRowCount(0)
        if self.adc_controller and lot_number:
            adc = self.adc_controller.get_adc_by_lot_number(lot_number)
            if adc:
                self.position_edit.setText(adc.storage_position)
                # 更新库存显示
                if adc.specs:
                    self.stock_table.setRowCount(len(adc.specs))
                    for row, spec in enumerate(adc.specs):
                        spec_mg = spec.spec_mg if isinstance(spec, ADCSpec) else spec.get('spec_mg', 0)
                        quantity = spec.quantity if isinstance(spec, ADCSpec) else spec.get('quantity', 0)
                        total = spec_mg * quantity
                        
                        self.stock_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
                        self.stock_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
                        self.stock_table.setItem(row, 2, QTableWidgetItem(f"{total:.2f}"))
        
        # 动态调整表格高度：显示所有行但不超过最大高度
        row_count = self.stock_table.rowCount()
        row_height = 30  # 每行高度
        header_height = 30  # 表头高度
        max_height = 180  # 最大高度
        min_height = 60  # 最小高度
        
        if row_count > 0:
            calculated_height = header_height + row_count * row_height + 5
            table_height = min(max(calculated_height, min_height), max_height)
        else:
            table_height = min_height
        
        self.stock_table.setFixedHeight(table_height)
    
    def _on_operator_changed(self, text: str):
        """入库人变更时，自动填充Owner"""
        if not hasattr(self, 'owner_edit'):
            return
        if not self.owner_edit.text():
            self.owner_edit.setText(text)
    
    def _refresh_items_table(self):
        self.items_table.setRowCount(len(self.items))
        total_mg = 0.0
        total_vials = 0
        for row, item in enumerate(self.items):
            spec_mg = item['spec_mg']
            quantity = item['quantity']
            subtotal = spec_mg * quantity
            total_mg += subtotal
            total_vials += quantity
            
            self.items_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
            self.items_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
            self.items_table.setItem(row, 2, QTableWidgetItem(f"{subtotal:.2f}"))
        
        self.total_label.setText(f"合计: {total_vials} 个小管, {total_mg:.2f} mg")
    
    def _add_item(self):
        dialog = ADCMovementItemDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.items.append(dialog.result)
            self._refresh_items_table()
    
    def _edit_item(self):
        current_row = self.items_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要编辑的明细")
            return
        
        dialog = ADCMovementItemDialog(self, self.items[current_row])
        if dialog.exec_() == QDialog.Accepted:
            self.items[current_row] = dialog.result
            self._refresh_items_table()
    
    def _remove_item(self):
        current_row = self.items_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要删除的明细")
            return
        
        del self.items[current_row]
        self._refresh_items_table()
    
    def _save(self):
        if not self.lot_combo.currentText().strip():
            QMessageBox.critical(self, "错误", "请选择Lot Number")
            return
        if not self.operator_edit.text().strip():
            QMessageBox.critical(self, "错误", "请输入入库人")
            return
        if not self.items:
            QMessageBox.critical(self, "错误", "请添加入库明细")
            return
        
        inbound = ADCInbound(
            lot_number=self.lot_combo.currentText().strip(),
            operator=self.operator_edit.text().strip(),
            owner=self.owner_edit.text().strip() or self.operator_edit.text().strip(),
            storage_position=self.position_edit.text().strip(),
            storage_date=datetime.now(),  # 使用当前时间
            notes=self.notes_edit.text().strip(),
            items=[ADCMovementItem(spec_mg=i['spec_mg'], quantity=i['quantity']) for i in self.items]
        )
        
        self.result = inbound
        self.accept()


class MainWindow(QMainWindow):
    """主窗口"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("生物实验室库存管理系统")
        self.setGeometry(100, 100, 1400, 900)
        
        # 初始化数据库和控制器
        self.db_manager = DatabaseManager()
        self._init_controllers()
        
        # 物料相关缓存
        self.material_cards = {}
        self.detail_panels = {}
        self.selected_material_id = None
        
        # ADC相关缓存
        self.adc_cards = {}
        self.adc_detail_panels = {}
        self.selected_adc_id = None
        
        self.setup_ui()
        self.refresh_data()
    
    def _init_controllers(self):
        """初始化所有控制器"""
        self.material_controller = MaterialController(self.db_manager)
        self.order_controller = OrderController(self.db_manager, self.material_controller)
        self.report_controller = ReportController(self.db_manager)
        self.adc_controller = ADCController(self.db_manager)
        self.workflow_controller = ADCWorkflowController(self.db_manager)
    
    def setup_ui(self):
        """设置用户界面"""
        # 创建中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # 数据库选择工具栏
        db_toolbar = QHBoxLayout()
        db_toolbar.addWidget(QLabel("当前数据库:"))
        
        self.db_combo = QComboBox()
        self.db_combo.setMinimumWidth(200)
        self._refresh_db_combo()
        self.db_combo.currentIndexChanged.connect(self._on_db_changed)
        db_toolbar.addWidget(self.db_combo)
        
        add_db_btn = QPushButton("添加")
        add_db_btn.clicked.connect(self._add_database)
        db_toolbar.addWidget(add_db_btn)
        
        remove_db_btn = QPushButton("移除")
        remove_db_btn.clicked.connect(self._remove_database)
        db_toolbar.addWidget(remove_db_btn)
        
        db_toolbar.addStretch()
        
        # 显示当前数据库路径
        self.db_path_label = QLabel()
        self.db_path_label.setStyleSheet("color: #666; font-size: 11px;")
        self._update_db_path_label()
        db_toolbar.addWidget(self.db_path_label)
        
        main_layout.addLayout(db_toolbar)
        
        # 创建标签页
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # ADC管理标签页
        adc_tab = QWidget()
        self.setup_adc_tab(adc_tab)
        self.tabs.addTab(adc_tab, "ADC管理")
        
        # ADC出入库管理标签页
        adc_movement_tab = QWidget()
        self.setup_adc_movement_tab(adc_movement_tab)
        self.tabs.addTab(adc_movement_tab, "ADC出入库")
        
        # ADC实验流程标签页
        adc_workflow_tab = QWidget()
        self.setup_adc_workflow_tab(adc_workflow_tab)
        self.tabs.addTab(adc_workflow_tab, "ADC实验流程")
        
        # 状态栏
        self.statusBar().showMessage("就绪 - 支持多用户并发访问")
        
        # 配置信息
        config = load_config()
        config_path = os.path.join(os.path.dirname(__file__), "config.json")
        config_display = f"📄 配置: {os.path.basename(config_path)} | "
        if config.get("database_path"):
            config_display += f"数据库: {config['database_path']}"
        else:
            config_display += "数据库: inventory.db"
        self.statusBar().addPermanentWidget(QLabel(config_display))
    
    def setup_material_tab(self, parent):
        """设置物料管理标签页"""
        layout = QVBoxLayout()
        parent.setLayout(layout)
        
        # 工具栏
        toolbar = QHBoxLayout()
        
        add_btn = QPushButton("添加物料")
        add_btn.clicked.connect(self.add_material)
        toolbar.addWidget(add_btn)
        
        edit_btn = QPushButton("编辑物料")
        edit_btn.clicked.connect(self.edit_material)
        toolbar.addWidget(edit_btn)
        
        delete_btn = QPushButton("删除物料")
        delete_btn.clicked.connect(self.delete_material)
        toolbar.addWidget(delete_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.refresh_materials)
        toolbar.addWidget(refresh_btn)
        
        toolbar.addWidget(QLabel("搜索:"))
        self.material_search_edit = QLineEdit()
        self.material_search_edit.textChanged.connect(self.search_materials)
        toolbar.addWidget(self.material_search_edit)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # 创建分割器
        splitter = QSplitter(Qt.Horizontal)
        
        # 左侧物料列表
        list_widget = QWidget()
        list_layout = QVBoxLayout()
        list_widget.setLayout(list_layout)
        
        self.material_scroll = QScrollArea()
        self.material_scroll.setWidgetResizable(True)
        self.material_scroll.setWidget(QWidget())
        list_layout.addWidget(self.material_scroll)
        
        splitter.addWidget(list_widget)
        splitter.setStretchFactor(0, 2)
        
        # 右侧详情面板
        self.detail_widget = QWidget()
        self.detail_layout = QVBoxLayout()
        self.detail_widget.setLayout(self.detail_layout)
        
        self.detail_placeholder = QLabel("请点击左侧物料卡片查看详情")
        self.detail_placeholder.setAlignment(Qt.AlignCenter)
        self.detail_layout.addWidget(self.detail_placeholder)
        
        splitter.addWidget(self.detail_widget)
        splitter.setStretchFactor(1, 1)
        
        layout.addWidget(splitter)
    
    def setup_order_tab(self, parent):
        """设置订单管理标签页"""
        layout = QVBoxLayout()
        parent.setLayout(layout)
        
        # 工具栏
        toolbar = QHBoxLayout()
        
        create_btn = QPushButton("创建订单")
        create_btn.clicked.connect(self.create_order)
        toolbar.addWidget(create_btn)
        
        edit_btn = QPushButton("编辑订单")
        edit_btn.clicked.connect(self.edit_order)
        toolbar.addWidget(edit_btn)
        
        complete_btn = QPushButton("完成订单")
        complete_btn.clicked.connect(self.complete_order)
        toolbar.addWidget(complete_btn)
        
        cancel_btn = QPushButton("取消订单")
        cancel_btn.clicked.connect(self.cancel_order)
        toolbar.addWidget(cancel_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.refresh_orders)
        toolbar.addWidget(refresh_btn)
        
        toolbar.addWidget(QLabel("状态:"))
        self.order_status_combo = QComboBox()
        self.order_status_combo.addItems(["all", "pending", "in_progress", "completed", "cancelled"])
        self.order_status_combo.currentTextChanged.connect(self.filter_orders)
        toolbar.addWidget(self.order_status_combo)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # 订单表格
        self.order_table = QTableWidget()
        self.order_table.setColumnCount(7)
        self.order_table.setHorizontalHeaderLabels(["ID", "订单号", "申请人", "部门", "状态", "优先级", "创建时间"])
        self.order_table.horizontalHeader().setStretchLastSection(True)
        self.order_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.order_table)
    
    def setup_adc_tab(self, parent):
        """设置ADC管理标签页"""
        layout = QVBoxLayout()
        parent.setLayout(layout)
        
        # 第一行工具栏：操作按钮
        toolbar_row1 = QHBoxLayout()
        
        add_btn = QPushButton("添加ADC")
        add_btn.clicked.connect(self.add_adc)
        toolbar_row1.addWidget(add_btn)
        
        edit_btn = QPushButton("编辑ADC")
        edit_btn.clicked.connect(self.edit_adc)
        toolbar_row1.addWidget(edit_btn)
        
        delete_btn = QPushButton("删除ADC")
        delete_btn.clicked.connect(self.delete_adc)
        toolbar_row1.addWidget(delete_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.refresh_adcs)
        toolbar_row1.addWidget(refresh_btn)
        
        export_csv_btn = QPushButton("导出CSV")
        export_csv_btn.clicked.connect(self.export_adc_to_csv)
        toolbar_row1.addWidget(export_csv_btn)
        
        export_excel_btn = QPushButton("导出Excel")
        export_excel_btn.clicked.connect(self.export_adc_to_excel)
        toolbar_row1.addWidget(export_excel_btn)
        
        toolbar_row1.addStretch()
        layout.addLayout(toolbar_row1)
        
        # 第二行工具栏：搜索框
        toolbar_row2 = QHBoxLayout()
        
        toolbar_row2.addWidget(QLabel("搜索:"))
        
        toolbar_row2.addWidget(QLabel("LotNumber:"))
        self.adc_lot_search_edit = QLineEdit()
        self.adc_lot_search_edit.setPlaceholderText("搜索LotNumber")
        self.adc_lot_search_edit.textChanged.connect(self.search_adcs)
        toolbar_row2.addWidget(self.adc_lot_search_edit)
        
        toolbar_row2.addWidget(QLabel("SampleID:"))
        self.adc_search_edit = QLineEdit()
        self.adc_search_edit.setPlaceholderText("搜索SampleID")
        self.adc_search_edit.textChanged.connect(self.search_adcs)
        toolbar_row2.addWidget(self.adc_search_edit)
        
        toolbar_row2.addWidget(QLabel("Antibody:"))
        self.adc_antibody_search_edit = QLineEdit()
        self.adc_antibody_search_edit.setPlaceholderText("搜索Antibody")
        self.adc_antibody_search_edit.textChanged.connect(self.search_adcs)
        toolbar_row2.addWidget(self.adc_antibody_search_edit)
        
        toolbar_row2.addWidget(QLabel("Linker-payload:"))
        self.adc_linker_search_edit = QLineEdit()
        self.adc_linker_search_edit.setPlaceholderText("搜索Linker-payload")
        self.adc_linker_search_edit.textChanged.connect(self.search_adcs)
        toolbar_row2.addWidget(self.adc_linker_search_edit)
        
        toolbar_row2.addStretch()
        layout.addLayout(toolbar_row2)
        
        # 创建分割器
        splitter = QSplitter(Qt.Horizontal)
        
        # 左侧ADC列表
        list_widget = QWidget()
        list_layout = QVBoxLayout()
        list_widget.setLayout(list_layout)
        
        self.adc_scroll = QScrollArea()
        self.adc_scroll.setWidgetResizable(True)
        self.adc_scroll.setWidget(QWidget())
        list_layout.addWidget(self.adc_scroll)
        
        splitter.addWidget(list_widget)
        splitter.setStretchFactor(0, 2)
        
        # 右侧详情面板
        self.adc_detail_widget = QWidget()
        self.adc_detail_layout = QVBoxLayout()
        self.adc_detail_widget.setLayout(self.adc_detail_layout)
        
        self.adc_detail_placeholder = QLabel("请点击左侧ADC卡片查看详情")
        self.adc_detail_placeholder.setAlignment(Qt.AlignCenter)
        self.adc_detail_layout.addWidget(self.adc_detail_placeholder)
        
        splitter.addWidget(self.adc_detail_widget)
        splitter.setStretchFactor(1, 1)
        
        layout.addWidget(splitter)
    
    def setup_adc_movement_tab(self, parent):
        """设置ADC出入库管理标签页"""
        layout = QVBoxLayout()
        parent.setLayout(layout)
        
        # 第一行：操作按钮
        toolbar_row1 = QHBoxLayout()
        
        inbound_btn = QPushButton("入库")
        inbound_btn.setStyleSheet("background-color: #28a745; color: white;")
        inbound_btn.clicked.connect(self.adc_inbound)
        toolbar_row1.addWidget(inbound_btn)
        
        outbound_btn = QPushButton("出库")
        outbound_btn.setStyleSheet("background-color: #dc3545; color: white;")
        outbound_btn.clicked.connect(self.adc_outbound)
        toolbar_row1.addWidget(outbound_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.refresh_adc_movements)
        toolbar_row1.addWidget(refresh_btn)
        
        toolbar_row1.addStretch()
        layout.addLayout(toolbar_row1)
        
        # 第二行：搜索条件
        toolbar_row2 = QHBoxLayout()
        
        # 类型筛选
        toolbar_row2.addWidget(QLabel("类型:"))
        self.movement_type_combo = QComboBox()
        self.movement_type_combo.addItems(["全部", "入库", "出库"])
        self.movement_type_combo.setMinimumWidth(80)
        self.movement_type_combo.currentIndexChanged.connect(self.search_adc_movements)
        toolbar_row2.addWidget(self.movement_type_combo)
        
        # LotNumber搜索
        toolbar_row2.addWidget(QLabel("LotNumber:"))
        self.movement_search_edit = QLineEdit()
        self.movement_search_edit.setPlaceholderText("输入LotNumber")
        self.movement_search_edit.textChanged.connect(self.search_adc_movements)
        toolbar_row2.addWidget(self.movement_search_edit)
        
        # 操作人搜索
        toolbar_row2.addWidget(QLabel("操作人:"))
        self.movement_operator_edit = QLineEdit()
        self.movement_operator_edit.setPlaceholderText("输入操作人")
        self.movement_operator_edit.textChanged.connect(self.search_adc_movements)
        toolbar_row2.addWidget(self.movement_operator_edit)
        
        # 日期范围筛选
        toolbar_row2.addWidget(QLabel("日期从:"))
        self.movement_date_from = QDateEdit()
        self.movement_date_from.setCalendarPopup(True)
        self.movement_date_from.setDisplayFormat("yyyy-MM-dd")
        self.movement_date_from.setSpecialValueText("不限")
        self.movement_date_from.setMinimumDate(QDate(2000, 1, 1))
        self.movement_date_from.setDate(QDate(2000, 1, 1))  # 设为最小值表示不限
        self.movement_date_from.dateChanged.connect(self.search_adc_movements)
        toolbar_row2.addWidget(self.movement_date_from)
        
        toolbar_row2.addWidget(QLabel("到:"))
        self.movement_date_to = QDateEdit()
        self.movement_date_to.setCalendarPopup(True)
        self.movement_date_to.setDisplayFormat("yyyy-MM-dd")
        self.movement_date_to.setSpecialValueText("不限")
        self.movement_date_to.setMinimumDate(QDate(2000, 1, 1))
        self.movement_date_to.setDate(QDate(2099, 12, 31))  # 设为未来日期表示不限
        self.movement_date_to.dateChanged.connect(self.search_adc_movements)
        toolbar_row2.addWidget(self.movement_date_to)
        
        # 清除搜索条件按钮
        clear_search_btn = QPushButton("清除筛选")
        clear_search_btn.clicked.connect(self.clear_movement_search)
        toolbar_row2.addWidget(clear_search_btn)
        
        toolbar_row2.addStretch()
        layout.addLayout(toolbar_row2)
        
        # 使用分割器分割左右区域
        splitter = QSplitter(Qt.Horizontal)
        
        # 左侧：出入库记录表格
        left_widget = QWidget()
        left_layout = QVBoxLayout()
        left_widget.setLayout(left_layout)
        
        self.movement_table = QTableWidget()
        self.movement_table.setColumnCount(7)
        self.movement_table.setHorizontalHeaderLabels([
            "类型", "Lot Number", "操作人", "日期", "明细", "合计(mg)", "备注"
        ])
        self.movement_table.horizontalHeader().setStretchLastSection(True)
        self.movement_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.movement_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.movement_table.itemSelectionChanged.connect(self._on_movement_selected)
        left_layout.addWidget(self.movement_table)
        
        splitter.addWidget(left_widget)
        
        # 右侧：详情面板
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_widget.setLayout(right_layout)
        
        # 右侧最上方：选中记录详情
        detail_group = QGroupBox("选中记录详情")
        detail_layout = QVBoxLayout()
        
        self.movement_lot_label = QLabel("请选择一条记录")
        self.movement_lot_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #007bff;")
        detail_layout.addWidget(self.movement_lot_label)
        
        self.movement_detail_label = QLabel("")
        self.movement_detail_label.setWordWrap(True)
        self.movement_detail_label.setStyleSheet("background-color: #f8f9fa; padding: 10px; border-radius: 5px;")
        detail_layout.addWidget(self.movement_detail_label)
        
        detail_group.setLayout(detail_layout)
        right_layout.addWidget(detail_group)
        
        # 右侧中间：该LotNumber的出入库历史
        history_group = QGroupBox("出入库历史")
        history_layout = QVBoxLayout()
        
        self.movement_history_table = QTableWidget()
        self.movement_history_table.setColumnCount(5)
        self.movement_history_table.setHorizontalHeaderLabels([
            "类型", "操作人", "日期", "明细", "合计(mg)"
        ])
        self.movement_history_table.horizontalHeader().setStretchLastSection(True)
        self.movement_history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.movement_history_table.setEditTriggers(QTableWidget.NoEditTriggers)
        history_layout.addWidget(self.movement_history_table)
        
        history_group.setLayout(history_layout)
        right_layout.addWidget(history_group)
        
        # 右侧下方：该LotNumber的当前库存
        stock_group = QGroupBox("当前库存")
        stock_layout = QVBoxLayout()
        
        self.movement_stock_table = QTableWidget()
        self.movement_stock_table.setColumnCount(3)
        self.movement_stock_table.setHorizontalHeaderLabels([
            "规格 (mg)", "数量 (小管)", "小计 (mg)"
        ])
        self.movement_stock_table.horizontalHeader().setStretchLastSection(True)
        self.movement_stock_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.movement_stock_table.setEditTriggers(QTableWidget.NoEditTriggers)
        stock_layout.addWidget(self.movement_stock_table)
        
        self.movement_stock_total_label = QLabel("汇总: 0 小管 / 0.00 mg")
        self.movement_stock_total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #28a745;")
        stock_layout.addWidget(self.movement_stock_total_label)
        
        stock_group.setLayout(stock_layout)
        right_layout.addWidget(stock_group)
        
        splitter.addWidget(right_widget)
        
        # 设置分割比例
        splitter.setStretchFactor(0, 2)
        splitter.setStretchFactor(1, 1)
        
        layout.addWidget(splitter)
    
    def _on_movement_selected(self):
        """出入库记录选中事件处理"""
        current_row = self.movement_table.currentRow()
        if current_row < 0:
            return
        
        # 获取选中行的LotNumber
        lot_number_item = self.movement_table.item(current_row, 1)
        if not lot_number_item:
            return
        
        lot_number = lot_number_item.text()
        self.movement_lot_label.setText(f"Lot Number: {lot_number}")
        
        # 更新选中记录详情
        self._update_movement_detail(current_row)
        
        # 更新出入库历史
        self._update_movement_history(lot_number)
        
        # 更新当前库存
        self._update_movement_stock(lot_number)
    
    def _update_movement_detail(self, row: int):
        """更新选中记录的详细信息"""
        if not hasattr(self, '_current_movements') or row >= len(self._current_movements):
            self.movement_detail_label.setText("")
            return
        
        movement = self._current_movements[row]
        record = movement['record']
        
        # 构建详情文本
        details = []
        
        if movement['type'] == 'outbound':
            # 出库记录详情
            details.append(f"<b>类型:</b> 出库")
            details.append(f"<b>需求人:</b> {record.requester}")
            details.append(f"<b>出库人:</b> {record.operator}")
            details.append(f"<b>寄送地址:</b> {record.shipping_address or '-'}")
            if record.shipping_date:
                if isinstance(record.shipping_date, datetime):
                    date_str = record.shipping_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    date_str = str(record.shipping_date)
                details.append(f"<b>寄送日期:</b> {date_str}")
            if record.notes:
                details.append(f"<b>备注:</b> {record.notes}")
        else:
            # 入库记录详情
            details.append(f"<b>类型:</b> 入库")
            details.append(f"<b>入库人:</b> {record.operator}")
            details.append(f"<b>Owner:</b> {record.owner or '-'}")
            details.append(f"<b>存放地址:</b> {record.storage_position or '-'}")
            if record.storage_date:
                if isinstance(record.storage_date, datetime):
                    date_str = record.storage_date.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    date_str = str(record.storage_date)
                details.append(f"<b>存放日期:</b> {date_str}")
            if record.notes:
                details.append(f"<b>备注:</b> {record.notes}")
        
        # 明细信息
        items = movement['items']
        items_str = ", ".join([
            f"{item.spec_mg}mg×{item.quantity}" if isinstance(item, ADCMovementItem) 
            else f"{item.get('spec_mg', 0)}mg×{item.get('quantity', 0)}"
            for item in items
        ])
        details.append(f"<b>明细:</b> {items_str}")
        
        # 合计
        total_mg = sum([
            item.spec_mg * item.quantity if isinstance(item, ADCMovementItem)
            else item.get('spec_mg', 0) * item.get('quantity', 0)
            for item in items
        ])
        total_vials = sum([
            item.quantity if isinstance(item, ADCMovementItem)
            else item.get('quantity', 0)
            for item in items
        ])
        details.append(f"<b>合计:</b> {total_vials} 小管 / {total_mg:.2f} mg")
        
        self.movement_detail_label.setText("<br>".join(details))
    
    def _update_movement_history(self, lot_number: str):
        """更新出入库历史表格"""
        movements = self.adc_controller.search_movements_by_lot_number(lot_number)
        
        self.movement_history_table.setRowCount(len(movements))
        
        for row, movement in enumerate(movements):
            # 类型
            type_text = "入库" if movement['type'] == 'inbound' else "出库"
            type_item = QTableWidgetItem(type_text)
            if movement['type'] == 'inbound':
                type_item.setBackground(QColor("#d4edda"))
            else:
                type_item.setBackground(QColor("#f8d7da"))
            self.movement_history_table.setItem(row, 0, type_item)
            
            # 操作人
            self.movement_history_table.setItem(row, 1, QTableWidgetItem(movement['operator']))
            
            # 日期
            date_str = ""
            if movement['date']:
                if isinstance(movement['date'], datetime):
                    date_str = movement['date'].strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(movement['date'], str):
                    try:
                        dt = datetime.strptime(movement['date'], '%Y-%m-%d %H:%M:%S.%f')
                        date_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        date_str = str(movement['date'])
                else:
                    date_str = str(movement['date'])
            self.movement_history_table.setItem(row, 2, QTableWidgetItem(date_str))
            
            # 明细
            items = movement['items']
            items_str = ", ".join([
                f"{item.spec_mg}mg×{item.quantity}" if isinstance(item, ADCMovementItem) 
                else f"{item.get('spec_mg', 0)}mg×{item.get('quantity', 0)}"
                for item in items
            ])
            self.movement_history_table.setItem(row, 3, QTableWidgetItem(items_str))
            
            # 合计
            total_mg = sum([
                item.spec_mg * item.quantity if isinstance(item, ADCMovementItem)
                else item.get('spec_mg', 0) * item.get('quantity', 0)
                for item in items
            ])
            self.movement_history_table.setItem(row, 4, QTableWidgetItem(f"{total_mg:.2f}"))
    
    def _update_movement_stock(self, lot_number: str):
        """更新当前库存表格"""
        adc = self.adc_controller.get_adc_by_lot_number(lot_number)
        
        if not adc:
            self.movement_stock_table.setRowCount(0)
            self.movement_stock_total_label.setText("汇总: 0 小管 / 0.00 mg")
            return
        
        specs = adc.specs if adc.specs else []
        self.movement_stock_table.setRowCount(len(specs))
        
        total_vials = 0
        total_mg = 0.0
        
        for row, spec in enumerate(specs):
            spec_mg = spec.spec_mg if isinstance(spec, ADCSpec) else spec.get('spec_mg', 0)
            quantity = spec.quantity if isinstance(spec, ADCSpec) else spec.get('quantity', 0)
            subtotal = spec_mg * quantity
            
            total_vials += quantity
            total_mg += subtotal
            
            self.movement_stock_table.setItem(row, 0, QTableWidgetItem(f"{spec_mg}"))
            self.movement_stock_table.setItem(row, 1, QTableWidgetItem(f"{quantity}"))
            self.movement_stock_table.setItem(row, 2, QTableWidgetItem(f"{subtotal:.2f}"))
        
        self.movement_stock_total_label.setText(f"汇总: {total_vials} 小管 / {total_mg:.2f} mg")
    
    # ==================== ADC实验流程 Tab ====================
    
    def setup_adc_workflow_tab(self, parent):
        """设置ADC实验流程标签页"""
        layout = QVBoxLayout()
        parent.setLayout(layout)
        
        # 当前用户与操作栏
        toolbar = QHBoxLayout()
        toolbar.addWidget(QLabel("当前用户:"))
        self.workflow_user_combo = QComboBox()
        self.workflow_user_combo.setMinimumWidth(120)
        self.workflow_user_combo.currentIndexChanged.connect(self._on_workflow_user_changed)
        toolbar.addWidget(self.workflow_user_combo)
        self.workflow_role_label = QLabel("")
        self.workflow_role_label.setStyleSheet("color: #666;")
        toolbar.addWidget(self.workflow_role_label)
        refresh_user_btn = QPushButton("刷新用户")
        refresh_user_btn.clicked.connect(self._refresh_workflow_user_combo)
        toolbar.addWidget(refresh_user_btn)
        toolbar.addStretch()
        self.workflow_import_btn = QPushButton("导入偶联任务文件")
        self.workflow_import_btn.clicked.connect(self._workflow_import_xlsx)
        toolbar.addWidget(self.workflow_import_btn)
        refresh_btn = QPushButton("刷新列表")
        refresh_btn.clicked.connect(self._refresh_workflow_list)
        toolbar.addWidget(refresh_btn)
        layout.addLayout(toolbar)
        
        splitter = QSplitter(Qt.Horizontal)
        
        # 左侧：流程列表
        left_w = QWidget()
        left_layout = QVBoxLayout()
        left_w.setLayout(left_layout)
        left_layout.addWidget(QLabel("实验流程列表"))
        self.workflow_table = QTableWidget()
        self.workflow_table.setColumnCount(5)
        self.workflow_table.setHorizontalHeaderLabels(["ID", "Request SN", "纯化步骤流程", "创建人", "创建时间"])
        self.workflow_table.horizontalHeader().setStretchLastSection(True)
        self.workflow_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.workflow_table.itemSelectionChanged.connect(self._on_workflow_selected)
        left_layout.addWidget(self.workflow_table)
        splitter.addWidget(left_w)
        
        # 右侧：详情
        right_w = QWidget()
        right_layout = QVBoxLayout()
        right_w.setLayout(right_layout)
        self.workflow_detail_placeholder = QLabel("请选择左侧一条实验流程")
        self.workflow_detail_placeholder.setAlignment(Qt.AlignCenter)
        right_layout.addWidget(self.workflow_detail_placeholder)
        
        self.workflow_detail_stack = QWidget()
        detail_stack_layout = QVBoxLayout()
        self.workflow_detail_stack.setLayout(detail_stack_layout)
        
        detail_stack_layout.addWidget(QLabel("Request 信息"))
        self.workflow_request_table = QTableWidget()
        self.workflow_request_table.setColumnCount(4)
        self.workflow_request_table.setHorizontalHeaderLabels(["字段名", "类型", "必填/可选", "值"])
        header = self.workflow_request_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setStretchLastSection(True)
        self.workflow_request_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.workflow_request_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.workflow_request_table.setAlternatingRowColors(True)
        self.workflow_request_table.setSizePolicy(
            self.workflow_request_table.sizePolicy().horizontalPolicy(),
            QSizePolicy.Expanding
        )
        self.workflow_request_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: #fff;
            }
            QTableWidget::item {
                padding: 4px 8px;
                color: #212529;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: #ecf0f1;
                padding: 8px 10px;
                font-weight: bold;
                border: none;
                border-right: 1px solid #34495e;
            }
            QHeaderView::section:last-child {
                border-right: none;
            }
        """)
        detail_stack_layout.addWidget(self.workflow_request_table, 1)
        
        detail_stack_layout.addWidget(QLabel("纯化步骤（可增删改）"))
        steps_btn_layout = QHBoxLayout()
        self.workflow_btn_step_up = QPushButton("上移")
        self.workflow_btn_step_up.clicked.connect(self._workflow_step_move_up)
        steps_btn_layout.addWidget(self.workflow_btn_step_up)
        self.workflow_btn_step_down = QPushButton("下移")
        self.workflow_btn_step_down.clicked.connect(self._workflow_step_move_down)
        steps_btn_layout.addWidget(self.workflow_btn_step_down)
        self.workflow_btn_step_add = QPushButton("添加步骤")
        self.workflow_btn_step_add.clicked.connect(self._workflow_step_add)
        steps_btn_layout.addWidget(self.workflow_btn_step_add)
        self.workflow_btn_step_remove = QPushButton("删除步骤")
        self.workflow_btn_step_remove.clicked.connect(self._workflow_step_remove)
        steps_btn_layout.addWidget(self.workflow_btn_step_remove)
        detail_stack_layout.addLayout(steps_btn_layout)
        self.workflow_steps_table = QTableWidget()
        self.workflow_steps_table.setColumnCount(3)
        self.workflow_steps_table.setHorizontalHeaderLabels(["顺序", "步骤类型", "参数/Estimated recovery"])
        self.workflow_steps_table.horizontalHeader().setStretchLastSection(True)
        self.workflow_steps_table.setMaximumHeight(140)
        detail_stack_layout.addWidget(self.workflow_steps_table, 0)
        
        action_btn_layout = QHBoxLayout()
        btn_feed = QPushButton("生成投料表")
        btn_feed.clicked.connect(self._workflow_show_feed_table)
        action_btn_layout.addWidget(btn_feed)
        self.workflow_btn_add_result = QPushButton("添加实验结果")
        self.workflow_btn_add_result.clicked.connect(self._workflow_add_result)
        action_btn_layout.addWidget(self.workflow_btn_add_result)
        self.workflow_btn_del_wf = QPushButton("删除本流程")
        self.workflow_btn_del_wf.clicked.connect(self._workflow_delete)
        action_btn_layout.addWidget(self.workflow_btn_del_wf)
        detail_stack_layout.addLayout(action_btn_layout)
        
        detail_stack_layout.addWidget(QLabel("实验结果列表"))
        self.workflow_results_table = QTableWidget()
        self.workflow_results_table.setColumnCount(6)
        self.workflow_results_table.setHorizontalHeaderLabels(
            ["Sample ID", "Lot No.", "Conc.(mg/mL)", "Yield(%)", "Purification Method", "操作"]
        )
        self.workflow_results_table.horizontalHeader().setStretchLastSection(True)
        self.workflow_results_table.setMaximumHeight(100)
        detail_stack_layout.addWidget(self.workflow_results_table, 0)
        
        right_layout.addWidget(self.workflow_detail_stack)
        self.workflow_detail_stack.hide()
        splitter.addWidget(right_w)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        layout.addWidget(splitter)
        
        self._refresh_workflow_user_combo()
        self._refresh_workflow_list()
    
    def _refresh_workflow_user_combo(self):
        self.workflow_user_combo.blockSignals(True)
        self.workflow_user_combo.clear()
        users = self.workflow_controller.get_all_users()
        for u in users:
            self.workflow_user_combo.addItem(u.username, u.id)
        if users:
            self.workflow_user_combo.setCurrentIndex(0)
        self.workflow_user_combo.blockSignals(False)
        self._on_workflow_user_changed()
    
    def _on_workflow_user_changed(self):
        idx = self.workflow_user_combo.currentIndex()
        if idx < 0:
            self.workflow_role_label.setText("")
            self.workflow_import_btn.setEnabled(False)
            return
        user_id = self.workflow_user_combo.currentData()
        if user_id is None:
            users = self.workflow_controller.get_all_users()
            if 0 <= idx < len(users):
                user_id = users[idx].id
        user = self.workflow_controller.get_user_by_id(user_id) if user_id else None
        self.workflow_role_label.setText(f"角色: {user.role}" if user else "")
        can_create = self.workflow_controller.can_create_workflow(user_id, user.role if user else "")
        self.workflow_import_btn.setEnabled(can_create)
        self._refresh_workflow_list()
    
    def _get_current_workflow_user_id_and_role(self):
        idx = self.workflow_user_combo.currentIndex()
        if idx < 0:
            return None, ""
        user_id = self.workflow_user_combo.currentData()
        if user_id is None:
            users = self.workflow_controller.get_all_users()
            if 0 <= idx < len(users):
                user_id = users[idx].id
        user = self.workflow_controller.get_user_by_id(user_id) if user_id else None
        return (user_id, user.role) if user else (None, "")
    
    def _workflow_require_can_edit(self) -> Optional[Tuple[int, str, ADCWorkflow]]:
        """若当前用户对当前 workflow 有编辑权限则返回 (user_id, role, workflow)，否则弹窗并返回 None。"""
        user_id, role = self._get_current_workflow_user_id_and_role()
        w = getattr(self, "_current_workflow", None)
        if not w or not user_id or not self.workflow_controller.can_edit_workflow(w, user_id, role):
            QMessageBox.warning(self, "权限", "您没有权限编辑此流程。")
            return None
        return (user_id, role, w)
    
    def _workflow_require_can_delete(self) -> Optional[Tuple[int, str, ADCWorkflow]]:
        """若当前用户对当前 workflow 有删除权限则返回 (user_id, role, workflow)，否则弹窗并返回 None。"""
        user_id, role = self._get_current_workflow_user_id_and_role()
        w = getattr(self, "_current_workflow", None)
        if not w or not user_id or not self.workflow_controller.can_delete_workflow(w, user_id, role):
            QMessageBox.warning(self, "权限", "您没有权限删除此流程。")
            return None
        return (user_id, role, w)
    
    def _refresh_workflow_list(self):
        user_id, role = self._get_current_workflow_user_id_and_role()
        if user_id is None:
            self.workflow_table.setRowCount(0)
            return
        workflows = self.workflow_controller.get_workflows_for_user(user_id, role)
        self.workflow_table.setRowCount(len(workflows))
        user_id_to_name = {u.id: u.username for u in self.workflow_controller.get_all_users()}
        for row, w in enumerate(workflows):
            self.workflow_table.setItem(row, 0, QTableWidgetItem(str(w.id)))
            self.workflow_table.setItem(row, 1, QTableWidgetItem(w.request_sn or ""))
            self.workflow_table.setItem(row, 2, QTableWidgetItem(w.purification_flow_string or ""))
            self.workflow_table.setItem(row, 3, QTableWidgetItem(user_id_to_name.get(w.created_by_user_id, "")))
            created = w.created_at.strftime("%Y-%m-%d %H:%M") if w.created_at else ""
            self.workflow_table.setItem(row, 4, QTableWidgetItem(created))
        self.workflow_table.setRowHidden(0, False)
    
    def _on_workflow_selected(self):
        current_row = self.workflow_table.currentRow()
        if current_row < 0:
            self.workflow_detail_placeholder.show()
            self.workflow_detail_stack.hide()
            return
        wf_id_item = self.workflow_table.item(current_row, 0)
        if not wf_id_item:
            return
        try:
            wf_id = int(wf_id_item.text())
        except ValueError:
            return
        workflow = self.workflow_controller.get_workflow_by_id(wf_id)
        if not workflow:
            return
        self.workflow_detail_placeholder.hide()
        self.workflow_detail_stack.show()
        self._current_workflow_id = wf_id
        self._current_workflow = workflow
        user_id, role = self._get_current_workflow_user_id_and_role()
        can_edit = self.workflow_controller.can_edit_workflow(workflow, user_id, role) if user_id is not None else False
        self.workflow_btn_step_up.setEnabled(can_edit)
        self.workflow_btn_step_down.setEnabled(can_edit)
        self.workflow_btn_step_add.setEnabled(can_edit)
        self.workflow_btn_step_remove.setEnabled(can_edit)
        self.workflow_btn_add_result.setEnabled(can_edit)
        self.workflow_btn_del_wf.setEnabled(can_edit)
        import json
        try:
            raw = json.loads(workflow.raw_request_json) if workflow.raw_request_json else {}
        except Exception:
            raw = {}
        ordered = ordered_request_items_for_display(raw)
        self.workflow_request_table.setRowCount(len(ordered))
        for row, (key, type_str, optional_label, value_str) in enumerate(ordered):
            def _center_item(text):
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                return item
            self.workflow_request_table.setItem(row, 0, _center_item(key))
            self.workflow_request_table.setItem(row, 1, _center_item(type_str))
            self.workflow_request_table.setItem(row, 2, _center_item(optional_label))
            value_item = QTableWidgetItem(value_str)
            value_item.setTextAlignment(Qt.AlignCenter)
            if value_str == "null":
                value_item.setForeground(QColor("#6c757d"))
                f = value_item.font()
                f.setItalic(True)
                value_item.setFont(f)
            self.workflow_request_table.setItem(row, 3, value_item)
        step_types = {t.id: t.name for t in self.workflow_controller.get_all_step_types(active_only=False)}
        self.workflow_steps_table.setRowCount(len(workflow.steps))
        for row, s in enumerate(workflow.steps):
            self.workflow_steps_table.setItem(row, 0, QTableWidgetItem(str(s.step_order + 1)))
            self.workflow_steps_table.setItem(row, 1, QTableWidgetItem(step_types.get(s.step_type_id, "")))
            self.workflow_steps_table.setItem(row, 2, QTableWidgetItem(s.params_json or "{}"))
        results = self.workflow_controller.get_experiment_results(wf_id)
        self.workflow_results_table.setRowCount(len(results))
        for row, r in enumerate(results):
            self.workflow_results_table.setItem(row, 0, QTableWidgetItem(r.sample_id))
            self.workflow_results_table.setItem(row, 1, QTableWidgetItem(r.lot_no))
            self.workflow_results_table.setItem(row, 2, QTableWidgetItem(str(r.conc_mg_ml)))
            self.workflow_results_table.setItem(row, 3, QTableWidgetItem(str(r.yield_pct)))
            self.workflow_results_table.setItem(row, 4, QTableWidgetItem(r.purification_method or ""))
            btn = QPushButton("删除")
            btn.setProperty("result_id", r.id)
            btn.setEnabled(can_edit)
            btn.clicked.connect(lambda checked, rid=r.id: self._workflow_delete_result(rid))
            self.workflow_results_table.setCellWidget(row, 5, btn)
    
    def _workflow_import_xlsx(self):
        user_id, role = self._get_current_workflow_user_id_and_role()
        if not self.workflow_controller.can_create_workflow(user_id, role):
            QMessageBox.warning(self, "提示", "请先选择当前用户。")
            return
        path, _ = QFileDialog.getOpenFileName(self, "选择偶联任务文件", "", "Excel (*.xlsx);;All (*)")
        if not path:
            return
        ok, msg, ids = self.workflow_controller.import_task_xlsx(path, user_id)
        if ok:
            QMessageBox.information(self, "导入结果", msg)
            self._refresh_workflow_list()
            # 选中第一行（新导入的流程）并刷新详情，避免仍显示旧选中项的空 Request 信息
            if self.workflow_table.rowCount() > 0:
                self.workflow_table.setCurrentCell(0, 0)
                self._on_workflow_selected()
        else:
            QMessageBox.warning(self, "导入失败", msg)
    
    def _workflow_step_move_up(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        if self._workflow_require_can_edit() is None:
            return
        row = self.workflow_steps_table.currentRow()
        if row <= 0:
            return
        w = self._current_workflow
        steps = list(w.steps)
        steps[row], steps[row - 1] = steps[row - 1], steps[row]
        names = []
        type_id_to_name = {t.id: t.name for t in self.workflow_controller.get_all_step_types(active_only=False)}
        for s in steps:
            names.append(type_id_to_name.get(s.step_type_id, ""))
        self.workflow_controller.update_workflow_steps(self._current_workflow_id, names)
        self._on_workflow_selected()
    
    def _workflow_step_move_down(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        if self._workflow_require_can_edit() is None:
            return
        row = self.workflow_steps_table.currentRow()
        if row < 0 or row >= self.workflow_steps_table.rowCount() - 1:
            return
        w = self._current_workflow
        steps = list(w.steps)
        steps[row], steps[row + 1] = steps[row + 1], steps[row]
        type_id_to_name = {t.id: t.name for t in self.workflow_controller.get_all_step_types(active_only=False)}
        names = [type_id_to_name.get(s.step_type_id, "") for s in steps]
        self.workflow_controller.update_workflow_steps(self._current_workflow_id, names)
        self._on_workflow_selected()
    
    def _workflow_step_add(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        if self._workflow_require_can_edit() is None:
            return
        types = self.workflow_controller.get_all_step_types(active_only=True)
        if not types:
            QMessageBox.information(self, "提示", "暂无纯化步骤类型。")
            return
        name, ok = QInputDialog.getItem(self, "添加步骤", "选择步骤类型:", [t.name for t in types], 0, False)
        if not ok or not name:
            return
        w = self._current_workflow
        type_id_to_name = {t.id: t.name for t in self.workflow_controller.get_all_step_types(active_only=False)}
        names = [type_id_to_name.get(s.step_type_id, "") for s in w.steps]
        names.append(name)
        self.workflow_controller.update_workflow_steps(self._current_workflow_id, names)
        self._on_workflow_selected()
    
    def _workflow_step_remove(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        if self._workflow_require_can_edit() is None:
            return
        row = self.workflow_steps_table.currentRow()
        if row < 0:
            return
        w = self._current_workflow
        type_id_to_name = {t.id: t.name for t in self.workflow_controller.get_all_step_types(active_only=False)}
        names = [type_id_to_name.get(s.step_type_id, "") for i, s in enumerate(w.steps) if i != row]
        self.workflow_controller.update_workflow_steps(self._current_workflow_id, names)
        flow_str = "+".join(names)
        self.workflow_controller.update_workflow_purification_string(self._current_workflow_id, flow_str)
        self._on_workflow_selected()
    
    def _workflow_show_feed_table(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        data = self.workflow_controller.get_feed_table_data(self._current_workflow_id)
        if not data:
            return
        raw_request = data.get("raw_request") or {}
        ordered_request = data.get("ordered_request") or []

        dlg = QDialog(self)
        dlg.setWindowTitle("ADC实验流程信息投料表 / Setup Param")
        dlg.setMinimumSize(900, 600)

        # 统一浅色现代风格样式
        dlg.setStyleSheet("""
            QDialog {
                background-color: #f5f7fb;
            }
            QGroupBox {
                background-color: #ffffff;
                border: 1px solid #d0d7e2;
                border-radius: 6px;
                margin-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 4px;
                color: #34495e;
                font-weight: bold;
            }
            QTableWidget {
                background-color: #ffffff;
                gridline-color: #e1e4eb;
                selection-background-color: #dbeafe;
                selection-color: #1f2933;
                alternate-background-color: #f8fafc;
            }
            QHeaderView::section {
                background-color: #edf2ff;
                color: #34495e;
                padding: 4px;
                border: 1px solid #d0d7e2;
                font-weight: bold;
            }
            QPushButton {
                background-color: #2d7ff9;
                color: #ffffff;
                border-radius: 4px;
                padding: 4px 10px;
            }
            QPushButton:disabled {
                background-color: #cfd8e3;
                color: #7b8794;
            }
            QPushButton:hover:!disabled {
                background-color: #1f6fe0;
            }
            QDialogButtonBox QPushButton {
                min-width: 80px;
                background-color: #ffffff;
                color: #1f2933;
                border: 1px solid #d0d7e2;
            }
            QDialogButtonBox QPushButton:hover {
                background-color: #f0f4ff;
            }
        """)

        main_layout = QVBoxLayout()
        dlg.setLayout(main_layout)

        # 顶部：基本信息卡片
        request_sn = data.get("request_sn") or ""
        wbp_code = raw_request.get("WBP Code", "")
        product_id = raw_request.get("Product ID", "")
        summary_frame = QFrame()
        summary_frame.setFrameShape(QFrame.StyledPanel)
        summary_frame.setStyleSheet("""
            QFrame {
                background-color: #ffffff;
                border: 1px solid #d0d7e2;
                border-radius: 8px;
            }
        """)
        summary_layout = QGridLayout()
        summary_frame.setLayout(summary_layout)
        title_label = QLabel("ADC 实验流程投料表 - DAR8 Setup Param")
        t_font = title_label.font()
        t_font.setPointSize(t_font.pointSize() + 1)
        t_font.setBold(True)
        title_label.setFont(t_font)
        summary_layout.addWidget(title_label, 0, 0, 1, 2)

        def _add_summary_row(r: int, name: str, value: str) -> None:
            name_label = QLabel(name)
            name_label.setStyleSheet("color: #6b7280;")
            value_label = QLabel(value)
            v_font = value_label.font()
            v_font.setBold(True)
            value_label.setFont(v_font)
            summary_layout.addWidget(name_label, r, 0)
            summary_layout.addWidget(value_label, r, 1)

        _add_summary_row(1, "Request SN:", str(request_sn))
        _add_summary_row(2, "WBP Code:", str(wbp_code))
        _add_summary_row(3, "Product ID:", str(product_id))

        main_layout.addWidget(summary_frame)

        # 中部：三列布局（左：Request 详细；中：SP 输入；右：结果+说明）
        center_layout = QHBoxLayout()
        main_layout.addLayout(center_layout, 1)

        # 左列：Request 详细信息（四列表，复用 ADC 实验流程 tab 风格）
        left_request_widget = QWidget()
        left_request_layout = QVBoxLayout()
        left_request_widget.setLayout(left_request_layout)
        center_layout.addWidget(left_request_widget, 1)

        left_request_layout.addWidget(QLabel("Request 输入（来自偶联任务）"))
        request_detail_table = QTableWidget()
        request_detail_table.setColumnCount(4)
        request_detail_table.setHorizontalHeaderLabels(["字段名", "类型", "必填/可选", "值"])
        from adc_workflow.request_schema import ordered_request_items_for_display
        request_items = ordered_request_items_for_display(raw_request)
        request_detail_table.setRowCount(len(request_items))
        for row, (key, type_str, optional_label, value_str) in enumerate(request_items):
            def _center_item(text: str) -> QTableWidgetItem:
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignCenter)
                return item
            request_detail_table.setItem(row, 0, _center_item(key))
            request_detail_table.setItem(row, 1, _center_item(type_str))
            request_detail_table.setItem(row, 2, _center_item(optional_label))
            value_item = QTableWidgetItem(value_str)
            value_item.setTextAlignment(Qt.AlignCenter)
            if value_str == "null":
                value_item.setForeground(QColor("#6c757d"))
                f = value_item.font()
                f.setItalic(True)
                value_item.setFont(f)
            request_detail_table.setItem(row, 3, value_item)
        request_detail_table.horizontalHeader().setStretchLastSection(True)
        request_detail_table.setAlternatingRowColors(True)
        left_request_layout.addWidget(request_detail_table)

        # 中列：SP 类型选择 + DAR8 输入
        middle_widget = QWidget()
        middle_layout = QVBoxLayout()
        middle_widget.setLayout(middle_layout)
        center_layout.addWidget(middle_widget, 1)

        # SP 类型选择
        sp_type_layout = QHBoxLayout()
        sp_type_label = QLabel("Setup Param 类型:")
        sp_type_combo = QComboBox()
        sp_type_combo.addItems(["DAR8", "DAR4", "Deblocking", "Thiomab"])
        sp_type_layout.addWidget(sp_type_label)
        sp_type_layout.addWidget(sp_type_combo)
        middle_layout.addLayout(sp_type_layout)

        # DAR8 输入区
        dar8_group = QGroupBox("DAR8 Setup Param 输入（用户输入）")
        dar8_form = QGridLayout()
        dar8_group.setLayout(dar8_form)

        row = 0

        def _add_spin(label_text, default, minimum, maximum, step=0.1):
            nonlocal row
            label = QLabel(label_text)
            spin = QDoubleSpinBox()
            spin.setRange(minimum, maximum)
            spin.setSingleStep(step)
            spin.setValue(default)
            dar8_form.addWidget(label, row, 0)
            dar8_form.addWidget(spin, row, 1)
            row += 1
            return spin

        tcep_eq_spin = _add_spin("TCEP 当量:", 8.0, 0.0, 1000.0, 0.1)
        tcep_stock_spin = _add_spin("TCEP stock (mM):", 8.0, 0.0, 1000.0, 0.1)
        ratio_spin = _add_spin("Conjugation organic solvent ratio (%):", 0.0, 0.0, 100.0, 1.0)
        xlp_spin = _add_spin("x LP/Ab:", 12.0, 0.0, 1000.0, 0.5)

        # 可选输入：checkbox + spin
        def _add_optional_spin(label_text):
            nonlocal row
            cb = QCheckBox(label_text)
            spin = QDoubleSpinBox()
            spin.setRange(0.0, 1000.0)
            spin.setSingleStep(0.5)
            spin.setEnabled(False)
            def _on_cb_changed(state):
                spin.setEnabled(state == Qt.Checked)
            cb.stateChanged.connect(_on_cb_changed)
            dar8_form.addWidget(cb, row, 0)
            dar8_form.addWidget(spin, row, 1)
            row += 1
            return cb, spin

        add_tcep_cb, add_tcep_spin = _add_optional_spin("Add additional TCEP (eq):")
        add_lp_cb, add_lp_spin = _add_optional_spin("Add additional LP:")
        add_time_cb, add_time_spin = _add_optional_spin("Additional reaction time (h):")

        # Reaction status
        status_label = QLabel("Reaction status:")
        status_combo = QComboBox()
        status_combo.addItems(["", "clear", "cloudy", "precipitate"])
        dar8_form.addWidget(status_label, row, 0)
        dar8_form.addWidget(status_combo, row, 1)
        row += 1

        # 重新计算按钮
        recalc_btn = QPushButton("重新计算 DAR8 参数")
        recalc_btn.setToolTip("根据当前所有用户输入重新计算右侧的 DAR8 Setup Param 结果。")
        dar8_form.addWidget(recalc_btn, row, 0, 1, 2)
        row += 1

        # 底部提示
        hint_label = QLabel("提示：修改任意输入参数后，点击上方按钮以刷新右侧结果与计算说明。")
        hint_label.setStyleSheet("color: #6b7280;")
        dar8_form.addWidget(hint_label, row, 0, 1, 2)
        row += 1

        middle_layout.addWidget(dar8_group)

        # 右列：SP 结果表 + 说明
        right_widget = QWidget()
        right_layout = QVBoxLayout()
        right_widget.setLayout(right_layout)
        center_layout.addWidget(right_widget, 1)

        result_table = QTableWidget()
        result_table.setColumnCount(2)
        result_table.setHorizontalHeaderLabels(["名称", "数值"])
        result_table.horizontalHeader().setStretchLastSection(True)
        right_layout.addWidget(QLabel("Setup Param 结果"))
        right_layout.addWidget(result_table, 3)

        explain_label = QLabel("计算过程说明")
        explain_text = QTextEdit()
        explain_text.setReadOnly(True)
        explain_text.setMinimumHeight(160)
        right_layout.addWidget(explain_label)
        right_layout.addWidget(explain_text, 2)

        # 底部图例：解释不同颜色/分组含义
        legend_layout = QHBoxLayout()
        legend_layout.addStretch(1)
        def _legend_item(color: str, text: str) -> QWidget:
            w = QWidget()
            lay = QHBoxLayout()
            lay.setContentsMargins(0, 0, 0, 0)
            w.setLayout(lay)
            box = QFrame()
            box.setFixedSize(14, 14)
            box.setStyleSheet(f"background-color: {color}; border: 1px solid #cbd5e1; border-radius: 3px;")
            lbl = QLabel(text)
            lbl.setStyleSheet("color: #6b7280;")
            lay.addWidget(box)
            lay.addWidget(lbl)
            return w
        legend_layout.addWidget(_legend_item("#e5e7eb", "Request 输入"))
        legend_layout.addWidget(_legend_item("#e6f4ea", "用户输入（回显）"))
        legend_layout.addWidget(_legend_item("#fff8e1", "Reduction 输出"))
        legend_layout.addWidget(_legend_item("#e3f2fd", "Conjugation 输出"))
        right_layout.addLayout(legend_layout)

        # 底部按钮
        bbox = QDialogButtonBox(QDialogButtonBox.Close)
        bbox.rejected.connect(dlg.reject)
        bbox.accepted.connect(dlg.accept)
        main_layout.addWidget(bbox)

        # ------ 内部状态与工具 ------
        dar8_meta = sp_dar8.get_dar8_field_meta_dict()
        dar8_fields = sp_dar8.DAR8_FIELDS
        current_result = None
        row_by_key: Dict[str, int] = {}
        row_base_color: Dict[int, QColor] = {}

        def _collect_user_inputs() -> Dict[str, Any]:
            inputs = {
                "tcep_eq": tcep_eq_spin.value(),
                "tcep_stock_mM": tcep_stock_spin.value(),
                "conj_org_ratio_percent": ratio_spin.value(),
                "x_lp_per_ab": xlp_spin.value(),
            }
            if add_tcep_cb.isChecked():
                inputs["add_additional_tcep_eq"] = add_tcep_spin.value()
            if add_lp_cb.isChecked():
                inputs["add_additional_lp"] = add_lp_spin.value()
            if add_time_cb.isChecked():
                inputs["additional_reaction_time_h"] = add_time_spin.value()
            status = status_combo.currentText().strip()
            if status:
                inputs["reaction_status"] = status
            return inputs

        def _format_value(v: Any) -> str:
            from adc_workflow.sp_core import format_number
            return format_number(v, digits=3)

        def _refresh_result_table():
            nonlocal current_result, row_by_key, row_base_color
            user_inputs = _collect_user_inputs()
            current_result = sp_dar8.calculate_dar8_sp(raw_request, user_inputs)
            # 按分组构建分节行
            sections = [
                ("input_request", "Request 输入", QColor("#e5e7eb")),
                ("input_user", "用户输入（回显）", QColor("#e6f4ea")),
                ("output_reduction", "Antibody Reduction 输出", QColor("#fff8e1")),
                ("output_conjugation", "Antibody Conjugation 输出", QColor("#e3f2fd")),
                ("meta", "元数据", QColor("#f3f4f6")),
            ]
            fields_by_group: Dict[str, list] = {}
            for f in dar8_fields:
                fields_by_group.setdefault(f.group, []).append(f)
            for g in fields_by_group.values():
                g.sort(key=lambda f: f.display_name.lower())

            rows_items = []
            for group_key, group_title, color in sections:
                group_fields = fields_by_group.get(group_key, [])
                if not group_fields:
                    continue
                rows_items.append(("header", group_title, group_key, color, None))
                for fmeta in group_fields:
                    rows_items.append(("field", fmeta, group_key, color, fmeta.key))

            result_table.setRowCount(len(rows_items))
            row_by_key = {}
            row_base_color = {}
            for row_idx, item in enumerate(rows_items):
                kind, payload, group_key, base_color, key = item
                if kind == "header":
                    header_item = QTableWidgetItem(str(payload))
                    header_item.setFlags(header_item.flags() & ~Qt.ItemIsSelectable & ~Qt.ItemIsEditable)
                    h_font = header_item.font()
                    h_font.setBold(True)
                    header_item.setFont(h_font)
                    header_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    header_item.setBackground(base_color)
                    result_table.setItem(row_idx, 0, header_item)
                    empty_item = QTableWidgetItem("")
                    empty_item.setFlags(empty_item.flags() & ~Qt.ItemIsSelectable & ~Qt.ItemIsEditable)
                    empty_item.setBackground(base_color)
                    result_table.setItem(row_idx, 1, empty_item)
                    row_base_color[row_idx] = base_color
                else:
                    fmeta: Any = payload
                    val = current_result.get_value(fmeta.key)
                    display_name = fmeta.display_name + (f" ({fmeta.unit})" if fmeta.unit else "")
                    name_item = QTableWidgetItem(display_name)
                    name_item.setData(Qt.UserRole, fmeta.key)
                    # tooltip 展示公式或说明
                    tip_parts = [display_name]
                    if fmeta.description:
                        tip_parts.append(fmeta.description)
                    if fmeta.formula_text:
                        tip_parts.append(f"公式: {fmeta.formula_text}")
                    name_item.setToolTip("\n".join(tip_parts))
                    value_str = _format_value(val) if fmeta.data_type in ("float", "optional_float") else ("" if val is None else str(val))
                    value_item = QTableWidgetItem(value_str)
                    name_item.setTextAlignment(Qt.AlignCenter)
                    value_item.setTextAlignment(Qt.AlignCenter)
                    # 分组底色
                    cell_color = QColor(base_color)
                    if fmeta.is_important and group_key == "output_reduction":
                        cell_color = QColor("#ffe8a1")
                    name_item.setBackground(cell_color)
                    value_item.setBackground(cell_color)
                    if fmeta.is_important:
                        f_font = name_item.font()
                        f_font.setBold(True)
                        name_item.setFont(f_font)
                        value_item.setFont(f_font)
                    result_table.setItem(row_idx, 0, name_item)
                    result_table.setItem(row_idx, 1, value_item)
                    row_by_key[fmeta.key] = row_idx
                    row_base_color[row_idx] = cell_color

        def _reset_highlight():
            rows = result_table.rowCount()
            cols = result_table.columnCount()
            for r in range(rows):
                for c in range(cols):
                    item = result_table.item(r, c)
                    if not item:
                        continue
                    # 恢复为基础底色
                    base = row_base_color.get(r)
                    if base is not None:
                        item.setBackground(base)

        def _on_result_cell_clicked(row_idx: int, column: int):
            if current_result is None:
                return
            item = result_table.item(row_idx, 0)
            if not item:
                return
            key = item.data(Qt.UserRole)
            if not key:
                return
            fmeta = dar8_meta.get(key)
            if not fmeta:
                return
            value = current_result.get_value(key)
            # 文字说明
            lines = []
            # 分组前缀
            group_prefix = ""
            if fmeta.group == "input_request":
                group_prefix = "【Request 输入】"
            elif fmeta.group == "input_user":
                group_prefix = "【用户输入】"
            elif fmeta.group == "output_reduction":
                group_prefix = "【Reduction 输出】"
            elif fmeta.group == "output_conjugation":
                group_prefix = "【Conjugation 输出】"
            elif fmeta.group == "meta":
                group_prefix = "【元数据】"

            title = f"{fmeta.display_name}"
            if fmeta.unit:
                title += f" ({fmeta.unit})"
            if group_prefix:
                lines.append(f"{group_prefix} 字段：{title}")
            else:
                lines.append(f"字段：{title}")
            lines.append(f"当前值：{_format_value(value) if fmeta.data_type in ('float', 'optional_float') else ('' if value is None else str(value))}")
            lines.append(f"数据来源：{fmeta.source}")
            if fmeta.description:
                lines.append(f"说明：{fmeta.description}")
            if fmeta.formula_text:
                lines.append(f"公式：{fmeta.formula_text}")
            if fmeta.depends_on:
                lines.append("依赖字段：")
                for dep_key in fmeta.depends_on:
                    dep_meta = dar8_meta.get(dep_key)
                    dep_val = current_result.get_value(dep_key)
                    dep_name = dep_meta.display_name if dep_meta else dep_key
                    dep_unit = f" ({dep_meta.unit})" if dep_meta and dep_meta.unit else ""
                    dep_str = _format_value(dep_val) if isinstance(dep_val, (int, float)) or dep_val is None else str(dep_val)
                    lines.append(f"  - {dep_name}{dep_unit} = {dep_str}")
            else:
                lines.append("该字段为原始输入或固定值，无上游依赖。")
            explain_text.setPlainText("\n".join(lines))

            # 高亮当前字段与依赖字段
            _reset_highlight()
            # 当前字段行：淡绿色
            for c in range(result_table.columnCount()):
                it = result_table.item(row_idx, c)
                if it:
                    it.setBackground(QColor("#d4edda"))
            # 依赖字段行：淡蓝色
            for dep_key in fmeta.depends_on:
                dep_row = row_by_key.get(dep_key)
                if dep_row is None:
                    continue
                for c in range(result_table.columnCount()):
                    it = result_table.item(dep_row, c)
                    if it:
                        it.setBackground(QColor("#d1ecf1"))

        result_table.cellClicked.connect(_on_result_cell_clicked)
        recalc_btn.clicked.connect(_refresh_result_table)

        # SP 类型切换逻辑：目前仅实现 DAR8，其它类型仅提示
        def _on_sp_type_changed(index: int):
            sp_type = sp_type_combo.currentText()
            enabled = sp_type == "DAR8"
            dar8_group.setEnabled(enabled)
            result_table.setEnabled(enabled)
            explain_text.setEnabled(enabled)
            if not enabled:
                explain_text.setPlainText("当前仅支持 DAR8 类型的 Setup Param 计算。")

        sp_type_combo.currentIndexChanged.connect(_on_sp_type_changed)
        _on_sp_type_changed(sp_type_combo.currentIndex())

        # 首次打开时自动计算一次
        _refresh_result_table()
        dlg.exec_()
    
    def _workflow_add_result(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        t = self._workflow_require_can_edit()
        if t is None:
            return
        user_id, role, w = t
        default_purification = w.purification_flow_string or ""
        try:
            import json
            raw = json.loads(w.raw_request_json) if w.raw_request_json else {}
        except Exception:
            raw = {}
        default_sample = (raw.get("Product ID") or raw.get("Request ID") or w.request_sn or "")
        if isinstance(default_sample, (int, float)):
            default_sample = str(int(default_sample))
        dlg = QDialog(self)
        dlg.setWindowTitle("添加实验结果")
        dlg.setMinimumSize(400, 420)
        form = QGridLayout()
        dlg.setLayout(form)
        r = 0
        sample_edit = QLineEdit(default_sample)
        form.addWidget(QLabel("Sample ID:"), r, 0)
        form.addWidget(sample_edit, r, 1)
        r += 1
        lot_edit = QLineEdit()
        lot_edit.setPlaceholderText("如 WBPX1111-260208001")
        form.addWidget(QLabel("Lot No.:"), r, 0)
        form.addWidget(lot_edit, r, 1)
        r += 1
        conc_spin = QDoubleSpinBox()
        form.addWidget(QLabel("Conc.(mg/mL):"), r, 0)
        form.addWidget(conc_spin, r, 1)
        r += 1
        amount_spin = QDoubleSpinBox()
        form.addWidget(QLabel("Amount(mg):"), r, 0)
        form.addWidget(amount_spin, r, 1)
        r += 1
        yield_spin = QDoubleSpinBox()
        form.addWidget(QLabel("Yield(%):"), r, 0)
        form.addWidget(yield_spin, r, 1)
        r += 1
        msdar_spin = QDoubleSpinBox()
        form.addWidget(QLabel("MS-DAR:"), r, 0)
        form.addWidget(msdar_spin, r, 1)
        r += 1
        mono_spin = QDoubleSpinBox()
        form.addWidget(QLabel("Monomer(%):"), r, 0)
        form.addWidget(mono_spin, r, 1)
        r += 1
        free_spin = QDoubleSpinBox()
        form.addWidget(QLabel("Free drug(%):"), r, 0)
        form.addWidget(free_spin, r, 1)
        r += 1
        endo_edit = QLineEdit()
        form.addWidget(QLabel("Endotoxin:"), r, 0)
        form.addWidget(endo_edit, r, 1)
        r += 1
        aliq_edit = QLineEdit()
        form.addWidget(QLabel("Aliquot:"), r, 0)
        form.addWidget(aliq_edit, r, 1)
        r += 1
        puri_edit = QLineEdit(default_purification)
        form.addWidget(QLabel("Purification Method:"), r, 0)
        form.addWidget(puri_edit, r, 1)
        r += 1
        bbox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        bbox.accepted.connect(dlg.accept)
        bbox.rejected.connect(dlg.reject)
        form.addWidget(bbox, r, 0, 1, 2)
        
        def get_float(w):
            return w.value() if w else 0.0
        def get_text(w):
            return w.text().strip() if w else ""
        
        if dlg.exec_() == QDialog.Accepted:
            lot_no = get_text(lot_edit)
            if lot_no and not self._is_lot_no_format_ok(lot_no):
                if QMessageBox.Yes != QMessageBox.question(
                    self, "Lot No. 格式",
                    "Lot No. 建议格式为 WBPX项目编号-日期-任务ID（如 WBPX1111-260208001）。是否仍要保存？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                ):
                    return
            self.workflow_controller.add_experiment_result(
                self._current_workflow_id,
                user_id,
                sample_id=get_text(sample_edit),
                lot_no=lot_no,
                conc_mg_ml=get_float(conc_spin),
                amount_mg=get_float(amount_spin),
                yield_pct=get_float(yield_spin),
                ms_dar=get_float(msdar_spin),
                monomer_pct=get_float(mono_spin),
                free_drug_pct=get_float(free_spin),
                endotoxin=get_text(endo_edit),
                aliquot=get_text(aliq_edit),
                purification_method=get_text(puri_edit),
            )
            self._on_workflow_selected()
    
    def _is_lot_no_format_ok(self, lot_no: str) -> bool:
        """Lot No. 建议格式 WBPX1111-260208001（项目编号-日期-任务ID）"""
        import re
        return bool(re.match(r"^WBPX\d+-\d{6}\d*$", lot_no.strip()))
    
    def _workflow_delete_result(self, result_id):
        if self._workflow_require_can_edit() is None:
            return
        if QMessageBox.Yes != QMessageBox.question(self, "确认", "确定删除该实验结果？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No):
            return
        self.workflow_controller.delete_experiment_result(result_id)
        self._on_workflow_selected()
    
    def _workflow_delete(self):
        if not getattr(self, "_current_workflow_id", None):
            return
        if self._workflow_require_can_delete() is None:
            return
        if QMessageBox.Yes != QMessageBox.question(self, "确认", "确定删除本实验流程及其步骤、实验结果？", QMessageBox.Yes | QMessageBox.No, QMessageBox.No):
            return
        self.workflow_controller.delete_workflow(self._current_workflow_id)
        self._current_workflow_id = None
        self._current_workflow = None
        self.workflow_detail_placeholder.show()
        self.workflow_detail_stack.hide()
        self._refresh_workflow_list()
    
    def setup_report_tab(self, parent):
        """设置报告生成标签页"""
        layout = QVBoxLayout()
        parent.setLayout(layout)
        
        info_label = QLabel("选择要生成报告的订单，系统将自动生成包含订单详细信息的HTML报告。")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        self.report_table = QTableWidget()
        self.report_table.setColumnCount(7)
        self.report_table.setHorizontalHeaderLabels(["ID", "订单号", "申请人", "部门", "状态", "优先级", "创建时间"])
        self.report_table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.report_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.report_table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(self.report_table)
        
        btn_layout = QHBoxLayout()
        generate_btn = QPushButton("生成报告")
        generate_btn.clicked.connect(self.generate_report)
        btn_layout.addWidget(generate_btn)
        
        refresh_btn = QPushButton("刷新订单列表")
        refresh_btn.clicked.connect(self.refresh_report_orders)
        btn_layout.addWidget(refresh_btn)
        
        layout.addLayout(btn_layout)
    
    # ==================== 数据库切换相关方法 ====================
    
    def _refresh_db_combo(self):
        """刷新数据库下拉菜单"""
        self.db_combo.blockSignals(True)
        self.db_combo.clear()
        
        databases = get_database_list()
        config = load_config()
        current_idx = config.get("current_database", 0)
        
        for db in databases:
            self.db_combo.addItem(db.get("name", "未命名"))
        
        if 0 <= current_idx < len(databases):
            self.db_combo.setCurrentIndex(current_idx)
        
        self.db_combo.blockSignals(False)
    
    def _update_db_path_label(self):
        """更新数据库路径标签"""
        self.db_path_label.setText(f"路径: {self.db_manager.db_path}")
    
    def _on_db_changed(self, index: int):
        """数据库选择变更事件"""
        if index < 0:
            return
        
        databases = get_database_list()
        if index >= len(databases):
            return
        
        db_path = databases[index].get("path", "inventory.db")
        
        # 保存选择
        set_current_database(index)
        
        # 切换数据库
        self.db_manager.switch_database(db_path)
        
        # 重新初始化控制器
        self._init_controllers()
        
        # 清空缓存
        self.material_cards.clear()
        self.detail_panels.clear()
        self.selected_material_id = None
        self.adc_cards.clear()
        self.adc_detail_panels.clear()
        self.selected_adc_id = None
        
        # 更新路径标签
        self._update_db_path_label()
        
        # 刷新数据
        self.refresh_data()
        
        QMessageBox.information(self, "提示", f"已切换到数据库: {databases[index].get('name', '未命名')}")
    
    def _add_database(self):
        """添加数据库"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择数据库文件",
            "",
            "SQLite Database (*.db);;All Files (*)"
        )
        
        if not file_path:
            return
        
        # 获取数据库名称
        from PyQt5.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(
            self,
            "数据库名称",
            "请输入数据库显示名称:",
            text=os.path.basename(file_path).replace('.db', '')
        )
        
        if not ok or not name:
            return
        
        if add_database(name, file_path):
            self._refresh_db_combo()
            # 自动切换到新添加的数据库
            new_index = len(get_database_list()) - 1
            self.db_combo.setCurrentIndex(new_index)
            QMessageBox.information(self, "成功", f"已添加并切换到数据库: {name}")
        else:
            QMessageBox.warning(self, "警告", "该数据库已存在于列表中")
    
    def _remove_database(self):
        """移除数据库"""
        current_idx = self.db_combo.currentIndex()
        databases = get_database_list()
        
        if len(databases) <= 1:
            QMessageBox.warning(self, "警告", "至少需要保留一个数据库")
            return
        
        db_name = databases[current_idx].get("name", "未命名")
        
        reply = QMessageBox.question(
            self,
            "确认移除",
            f"确定要从列表中移除数据库 '{db_name}' 吗？\n（不会删除数据库文件）",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            if remove_database(current_idx):
                self._refresh_db_combo()
                # 如果移除的是当前数据库，需要切换到其他数据库
                self._on_db_changed(self.db_combo.currentIndex())
                QMessageBox.information(self, "成功", f"已移除数据库: {db_name}")
    
    def refresh_data(self):
        """刷新所有数据"""
        self.refresh_adcs()
        self.refresh_adc_movements()
        if hasattr(self, "_refresh_workflow_user_combo"):
            self._refresh_workflow_user_combo()
        if hasattr(self, "_refresh_workflow_list"):
            self._refresh_workflow_list()
    
    # ==================== 物料相关方法 ====================
    
    def refresh_materials(self):
        """刷新物料列表"""
        materials = self.material_controller.get_all_materials()
        self.update_material_cards(materials)
    
    def update_material_cards(self, materials: List[Material]):
        """更新物料卡片"""
        # 清空现有卡片
        for card in self.material_cards.values():
            card.deleteLater()
        self.material_cards.clear()
        
        # 清空详情面板缓存
        for panel in self.detail_panels.values():
            panel.deleteLater()
        self.detail_panels.clear()
        
        self.selected_material_id = None
        
        # 创建新卡片
        container = QWidget()
        layout = QVBoxLayout()
        container.setLayout(layout)
        
        for material in materials:
            card = MaterialCard(material)
            card.clicked.connect(self._on_material_card_clicked)
            layout.addWidget(card)
            self.material_cards[material.id] = card
        
        layout.addStretch()
        
        self.material_scroll.setWidget(container)
        
        # 显示placeholder
        self.detail_placeholder.show()
    
    def _on_material_card_clicked(self, material_id: int):
        """物料卡片点击事件"""
        # 取消之前选中的卡片
        if self.selected_material_id:
            if self.selected_material_id in self.material_cards:
                self.material_cards[self.selected_material_id].set_selected(False)
        
        # 选中当前卡片
        if material_id in self.material_cards:
            self.material_cards[material_id].set_selected(True)
        self.selected_material_id = material_id
        
        # 显示详情
        self._show_material_detail(material_id)
    
    def _show_material_detail(self, material_id: int):
        """显示物料详情"""
        # 隐藏placeholder
        self.detail_placeholder.hide()
        
        # 如果已经有缓存的面板，直接显示
        if material_id in self.detail_panels:
            for mid, panel in self.detail_panels.items():
                panel.hide()
            self.detail_panels[material_id].show()
            return
        
        # 从缓存获取物料信息
        material = self.material_controller.get_material(material_id)
        if not material:
            return
        
        # 创建新的详情面板并缓存
        panel = MaterialDetailPanel(material, self.detail_widget)
        panel.edit_requested.connect(self.edit_material_by_id)
        panel.delete_requested.connect(self.delete_material_by_id)
        self.detail_panels[material_id] = panel
        self.detail_layout.addWidget(panel)
    
    def add_material(self):
        """添加物料"""
        dialog = MaterialDialog(self, material_controller=self.material_controller)
        if dialog.exec_() == QDialog.Accepted:
            material = dialog.result
            if material:
                try:
                    self.material_controller.create_material(material)
                    QMessageBox.information(self, "成功", "物料添加成功")
                    self.refresh_materials()
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"添加失败: {str(e)}")
    
    def edit_material(self):
        """编辑物料"""
        if not self.selected_material_id:
            QMessageBox.warning(self, "警告", "请先选择一个物料")
            return
        
        self.edit_material_by_id(self.selected_material_id)
    
    def edit_material_by_id(self, material_id: int):
        """根据ID编辑物料"""
        material = self.material_controller.get_material(material_id)
        if not material:
            QMessageBox.critical(self, "错误", "物料不存在")
            return
        
        dialog = MaterialDialog(self, material, self.material_controller)
        if dialog.exec_() == QDialog.Accepted:
            updated_material = dialog.result
            if updated_material:
                try:
                    success, message = self.material_controller.update_material(updated_material)
                    if success:
                        QMessageBox.information(self, "成功", message)
                        self.refresh_materials()
                    else:
                        QMessageBox.critical(self, "错误", message)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"更新失败: {str(e)}")
    
    def delete_material(self):
        """删除物料"""
        if not self.selected_material_id:
            QMessageBox.warning(self, "警告", "请先选择一个物料")
            return
        
        self.delete_material_by_id(self.selected_material_id)
    
    def delete_material_by_id(self, material_id: int):
        """根据ID删除物料"""
        if QMessageBox.question(self, "确认", "确定要删除这个物料吗？") == QMessageBox.Yes:
            try:
                self.material_controller.delete_material(material_id)
                QMessageBox.information(self, "成功", "物料删除成功")
                self.refresh_materials()
            except Exception as e:
                QMessageBox.critical(self, "错误", f"删除失败: {str(e)}")
    
    def search_materials(self):
        """搜索物料"""
        keyword = self.material_search_edit.text()
        if keyword:
            materials = self.material_controller.search_materials(keyword)
        else:
            materials = self.material_controller.get_all_materials()
        
        self.update_material_cards(materials)
    
    # ==================== 订单相关方法 ====================
    
    def refresh_orders(self):
        """刷新订单列表"""
        orders = self.order_controller.get_all_orders()
        self.order_table.setRowCount(len(orders))
        
        for row, order in enumerate(orders):
            self.order_table.setItem(row, 0, QTableWidgetItem(str(order.id)))
            self.order_table.setItem(row, 1, QTableWidgetItem(order.order_number))
            self.order_table.setItem(row, 2, QTableWidgetItem(order.requester))
            self.order_table.setItem(row, 3, QTableWidgetItem(order.department or ""))
            self.order_table.setItem(row, 4, QTableWidgetItem(order.status))
            self.order_table.setItem(row, 5, QTableWidgetItem(order.priority))
            created_at = order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else 'N/A'
            self.order_table.setItem(row, 6, QTableWidgetItem(created_at))
    
    def refresh_report_orders(self):
        """刷新报告页面的订单列表"""
        orders = self.order_controller.get_all_orders()
        self.report_table.setRowCount(len(orders))
        
        for row, order in enumerate(orders):
            self.report_table.setItem(row, 0, QTableWidgetItem(str(order.id)))
            self.report_table.setItem(row, 1, QTableWidgetItem(order.order_number))
            self.report_table.setItem(row, 2, QTableWidgetItem(order.requester))
            self.report_table.setItem(row, 3, QTableWidgetItem(order.department or ""))
            self.report_table.setItem(row, 4, QTableWidgetItem(order.status))
            self.report_table.setItem(row, 5, QTableWidgetItem(order.priority))
            created_at = order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else 'N/A'
            self.report_table.setItem(row, 6, QTableWidgetItem(created_at))
    
    def create_order(self):
        """创建订单"""
        dialog = OrderDialog(self, material_controller=self.material_controller)
        if dialog.exec_() == QDialog.Accepted:
            order = dialog.result
            if order:
                try:
                    self.order_controller.create_order(order)
                    QMessageBox.information(self, "成功", "订单创建成功")
                    self.refresh_orders()
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"创建失败: {str(e)}")
    
    def edit_order(self):
        """编辑订单"""
        current_row = self.order_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要编辑的订单")
            return
        
        order_id = int(self.order_table.item(current_row, 0).text())
        order = self.order_controller.get_order(order_id)
        
        if order:
            dialog = OrderDialog(self, order, self.material_controller)
            if dialog.exec_() == QDialog.Accepted:
                updated_order = dialog.result
                if updated_order:
                    try:
                        self.order_controller.update_order(updated_order)
                        QMessageBox.information(self, "成功", "订单更新成功")
                        self.refresh_orders()
                    except Exception as e:
                        QMessageBox.critical(self, "错误", f"更新失败: {str(e)}")
    
    def complete_order(self):
        """完成订单"""
        current_row = self.order_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要完成的订单")
            return
        
        order_id = int(self.order_table.item(current_row, 0).text())
        order_number = self.order_table.item(current_row, 1).text()
        
        if QMessageBox.question(self, "确认完成订单", 
                              f"确定要完成订单 {order_number} 吗？\n\n"
                              f"此操作将：\n"
                              f"• 更新订单状态为已完成\n"
                              f"• 减少相关物料的库存\n"
                              f"• 记录库存变动历史\n\n"
                              f"此操作不可撤销！") == QMessageBox.Yes:
            try:
                success, message = self.order_controller.complete_order(order_id)
                if success:
                    QMessageBox.information(self, "成功", message)
                    self.refresh_orders()
                    self.refresh_materials()
                else:
                    QMessageBox.critical(self, "错误", message)
            except Exception as e:
                QMessageBox.critical(self, "错误", f"完成失败: {str(e)}")
    
    def cancel_order(self):
        """取消订单"""
        current_row = self.order_table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "警告", "请选择要取消的订单")
            return
        
        if QMessageBox.question(self, "确认", "确定要取消选中的订单吗？") == QMessageBox.Yes:
            order_id = int(self.order_table.item(current_row, 0).text())
            try:
                self.order_controller.cancel_order(order_id)
                QMessageBox.information(self, "成功", "订单已取消")
                self.refresh_orders()
            except Exception as e:
                QMessageBox.critical(self, "错误", f"取消失败: {str(e)}")
    
    def filter_orders(self):
        """筛选订单"""
        status = self.order_status_combo.currentText()
        if status == "all":
            orders = self.order_controller.get_all_orders()
        else:
            orders = self.order_controller.get_orders_by_status(status)
        
        self.order_table.setRowCount(len(orders))
        for row, order in enumerate(orders):
            self.order_table.setItem(row, 0, QTableWidgetItem(str(order.id)))
            self.order_table.setItem(row, 1, QTableWidgetItem(order.order_number))
            self.order_table.setItem(row, 2, QTableWidgetItem(order.requester))
            self.order_table.setItem(row, 3, QTableWidgetItem(order.department or ""))
            self.order_table.setItem(row, 4, QTableWidgetItem(order.status))
            self.order_table.setItem(row, 5, QTableWidgetItem(order.priority))
            created_at = order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else 'N/A'
            self.order_table.setItem(row, 6, QTableWidgetItem(created_at))
    
    def generate_report(self):
        """生成订单报告"""
        selected_ranges = self.report_table.selectedRanges()
        if not selected_ranges:
            QMessageBox.warning(self, "警告", "请选择要生成报告的订单")
            return
        
        order_ids = set()
        for range_item in selected_ranges:
            top_row = range_item.topRow()
            bottom_row = range_item.bottomRow()
            for row in range(top_row, bottom_row + 1):
                if self.report_table.item(row, 0):
                    order_id = int(self.report_table.item(row, 0).text())
                    order_ids.add(order_id)
        
        if not order_ids:
            QMessageBox.warning(self, "警告", "请选择要生成报告的订单")
            return
        
        try:
            html_content = self.report_controller.generate_order_report(list(order_ids))
            
            filename, _ = QFileDialog.getSaveFileName(
                self, "保存报告", "order_report.html", "HTML文件 (*.html)"
            )
            
            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write(html_content)
                QMessageBox.information(self, "成功", f"报告已保存到: {filename}")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"生成报告失败: {str(e)}")
    
    # ==================== ADC相关方法 ====================
    
    def refresh_adcs(self):
        """刷新ADC列表"""
        adcs = self.adc_controller.get_all_adcs()
        self.update_adc_cards(adcs)
    
    def update_adc_cards(self, adcs: List[ADC]):
        """更新ADC卡片"""
        # 清空现有卡片
        for card in self.adc_cards.values():
            card.deleteLater()
        self.adc_cards.clear()
        
        # 清空详情面板缓存
        for panel in self.adc_detail_panels.values():
            panel.deleteLater()
        self.adc_detail_panels.clear()
        
        self.selected_adc_id = None
        
        # 创建新卡片
        container = QWidget()
        layout = QVBoxLayout()
        container.setLayout(layout)
        
        for adc in adcs:
            card = ADCCard(adc)
            card.clicked.connect(self._on_adc_card_clicked)
            layout.addWidget(card)
            self.adc_cards[adc.id] = card
        
        layout.addStretch()
        
        self.adc_scroll.setWidget(container)
        
        # 显示placeholder
        self.adc_detail_placeholder.show()
    
    def _on_adc_card_clicked(self, adc_id: int):
        """ADC卡片点击事件"""
        # 取消之前选中的卡片
        if self.selected_adc_id:
            if self.selected_adc_id in self.adc_cards:
                self.adc_cards[self.selected_adc_id].set_selected(False)
        
        # 选中当前卡片
        if adc_id in self.adc_cards:
            self.adc_cards[adc_id].set_selected(True)
        self.selected_adc_id = adc_id
        
        # 显示详情
        self._show_adc_detail(adc_id)
    
    def _show_adc_detail(self, adc_id: int):
        """显示ADC详情"""
        # 隐藏placeholder
        self.adc_detail_placeholder.hide()
        
        # 如果已经有缓存的面板，直接显示
        if adc_id in self.adc_detail_panels:
            for aid, panel in self.adc_detail_panels.items():
                panel.hide()
            self.adc_detail_panels[adc_id].show()
            return
        
        # 从缓存获取ADC信息
        adc = self.adc_controller.get_adc(adc_id)
        if not adc:
            return
        
        # 创建新的详情面板并缓存
        panel = ADCDetailPanel(adc, self.adc_detail_widget)
        panel.edit_requested.connect(self.edit_adc_by_id)
        panel.delete_requested.connect(self.delete_adc_by_id)
        self.adc_detail_panels[adc_id] = panel
        self.adc_detail_layout.addWidget(panel)
    
    def add_adc(self):
        """添加ADC"""
        dialog = ADCDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            adc = dialog.result
            if adc:
                try:
                    self.adc_controller.create_adc(adc)
                    QMessageBox.information(self, "成功", "ADC添加成功")
                    self.refresh_adcs()
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"添加失败: {str(e)}")
    
    def edit_adc(self):
        """编辑ADC"""
        if not self.selected_adc_id:
            QMessageBox.warning(self, "警告", "请先选择一个ADC")
            return
        
        self.edit_adc_by_id(self.selected_adc_id)
    
    def edit_adc_by_id(self, adc_id: int):
        """根据ID编辑ADC"""
        adc = self.adc_controller.get_adc(adc_id)
        if not adc:
            QMessageBox.critical(self, "错误", "ADC不存在")
            return
        
        dialog = ADCDialog(self, adc)
        if dialog.exec_() == QDialog.Accepted:
            updated_adc = dialog.result
            if updated_adc:
                try:
                    success, message = self.adc_controller.update_adc(updated_adc)
                    if success:
                        QMessageBox.information(self, "成功", message)
                        self.refresh_adcs()
                    else:
                        QMessageBox.critical(self, "错误", message)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"更新失败: {str(e)}")
    
    def delete_adc(self):
        """删除ADC"""
        if not self.selected_adc_id:
            QMessageBox.warning(self, "警告", "请先选择一个ADC")
            return
        
        self.delete_adc_by_id(self.selected_adc_id)
    
    def delete_adc_by_id(self, adc_id: int):
        """根据ID删除ADC"""
        if QMessageBox.question(self, "确认", "确定要删除这个ADC吗？") == QMessageBox.Yes:
            try:
                self.adc_controller.delete_adc(adc_id)
                QMessageBox.information(self, "成功", "ADC删除成功")
                self.refresh_adcs()
            except Exception as e:
                QMessageBox.critical(self, "错误", f"删除失败: {str(e)}")
    
    def search_adcs(self):
        """搜索ADC（支持多条件组合搜索）"""
        lot_number = self.adc_lot_search_edit.text().strip()
        sample_id = self.adc_search_edit.text().strip()
        antibody = self.adc_antibody_search_edit.text().strip()
        linker_payload = self.adc_linker_search_edit.text().strip()
        
        # 获取所有ADC
        adcs = self.adc_controller.get_all_adcs()
        
        # 按条件过滤
        if lot_number:
            lot_number_lower = lot_number.lower()
            adcs = [adc for adc in adcs if lot_number_lower in adc.lot_number.lower()]
        
        if sample_id:
            sample_id_lower = sample_id.lower()
            adcs = [adc for adc in adcs if sample_id_lower in adc.sample_id.lower()]
        
        if antibody:
            antibody_lower = antibody.lower()
            adcs = [adc for adc in adcs if antibody_lower in adc.antibody.lower()]
        
        if linker_payload:
            linker_lower = linker_payload.lower()
            adcs = [adc for adc in adcs if linker_lower in adc.linker_payload.lower()]
        
        self.update_adc_cards(adcs)
    
    def export_adc_to_csv(self):
        """导出ADC库存到CSV文件"""
        import csv
        
        # 获取所有ADC数据
        adcs = self.adc_controller.get_all_adcs()
        
        if not adcs:
            QMessageBox.warning(self, "警告", "没有可导出的ADC数据")
            return
        
        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出ADC库存",
            "adc_inventory.csv",
            "CSV Files (*.csv)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                
                # 写入表头
                writer.writerow([
                    'Lot Number', 'Sample ID', 'Description', 'Concentration (mg/mL)',
                    'Owner', 'Storage Temp', 'Storage Position', '入库时间',
                    '规格 (mg)', '数量 (小管)', '小计 (mg)'
                ])
                
                # 写入数据
                for adc in adcs:
                    # 入库时间格式化
                    created_at_str = ""
                    if adc.created_at:
                        if isinstance(adc.created_at, datetime):
                            created_at_str = adc.created_at.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            created_at_str = str(adc.created_at)
                    
                    # 每个规格一行
                    if adc.specs:
                        for spec in adc.specs:
                            spec_mg = spec.spec_mg if isinstance(spec, ADCSpec) else spec.get('spec_mg', 0)
                            quantity = spec.quantity if isinstance(spec, ADCSpec) else spec.get('quantity', 0)
                            subtotal = spec_mg * quantity
                            
                            writer.writerow([
                                adc.lot_number,
                                adc.sample_id,
                                adc.description,
                                adc.concentration,
                                adc.owner,
                                adc.storage_temp,
                                adc.storage_position,
                                created_at_str,
                                spec_mg,
                                quantity,
                                f"{subtotal:.2f}"
                            ])
                    else:
                        # 没有规格的ADC也导出一行
                        writer.writerow([
                            adc.lot_number,
                            adc.sample_id,
                            adc.description,
                            adc.concentration,
                            adc.owner,
                            adc.storage_temp,
                            adc.storage_position,
                            created_at_str,
                            '',
                            '',
                            ''
                        ])
            
            QMessageBox.information(self, "成功", f"已成功导出到: {file_path}")
        
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    def export_adc_to_excel(self):
        """导出ADC库存到Excel文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "错误", "请先安装openpyxl库: pip install openpyxl")
            return
        
        # 获取所有ADC数据
        adcs = self.adc_controller.get_all_adcs()
        
        if not adcs:
            QMessageBox.warning(self, "警告", "没有可导出的ADC数据")
            return
        
        # 选择保存路径
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "导出ADC库存",
            "adc_inventory.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ADC库存"
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 交替行颜色
            color1 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            color2 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # 表头
            headers = [
                'Lot Number', 'Sample ID', 'Description', 'Concentration (mg/mL)',
                'Owner', 'Storage Temp', 'Storage Position', '入库时间',
                '规格 (mg)', '数量 (小管)', '小计 (mg)', 'ADC汇总 (mg)'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 设置列宽
            column_widths = [18, 15, 25, 18, 12, 14, 18, 20, 12, 14, 14, 16]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # 写入数据
            current_row = 2
            for adc_index, adc in enumerate(adcs):
                # 入库时间格式化
                created_at_str = ""
                if adc.created_at:
                    if isinstance(adc.created_at, datetime):
                        created_at_str = adc.created_at.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        created_at_str = str(adc.created_at)
                
                # 计算该ADC的总毫克数
                total_mg = adc.get_total_mg()
                
                # 选择背景色（交替）
                row_fill = color1 if adc_index % 2 == 0 else color2
                
                # 获取规格列表
                specs = adc.specs if adc.specs else []
                spec_count = max(len(specs), 1)  # 至少1行
                
                start_row = current_row
                
                if specs:
                    for spec_index, spec in enumerate(specs):
                        spec_mg = spec.spec_mg if isinstance(spec, ADCSpec) else spec.get('spec_mg', 0)
                        quantity = spec.quantity if isinstance(spec, ADCSpec) else spec.get('quantity', 0)
                        subtotal = spec_mg * quantity
                        
                        row = current_row + spec_index
                        
                        # 只在第一行写入ADC基本信息
                        if spec_index == 0:
                            ws.cell(row=row, column=1, value=adc.lot_number)
                            ws.cell(row=row, column=2, value=adc.sample_id)
                            ws.cell(row=row, column=3, value=adc.description)
                            ws.cell(row=row, column=4, value=adc.concentration)
                            ws.cell(row=row, column=5, value=adc.owner)
                            ws.cell(row=row, column=6, value=adc.storage_temp)
                            ws.cell(row=row, column=7, value=adc.storage_position)
                            ws.cell(row=row, column=8, value=created_at_str)
                            ws.cell(row=row, column=12, value=total_mg)
                        
                        # 规格数据
                        ws.cell(row=row, column=9, value=spec_mg)
                        ws.cell(row=row, column=10, value=quantity)
                        ws.cell(row=row, column=11, value=subtotal)
                        
                        # 应用样式
                        for col in range(1, 13):
                            cell = ws.cell(row=row, column=col)
                            cell.fill = row_fill
                            cell.border = thin_border
                            if col in [4, 9, 10, 11, 12]:
                                cell.alignment = Alignment(horizontal="right")
                else:
                    # 没有规格的ADC
                    row = current_row
                    ws.cell(row=row, column=1, value=adc.lot_number)
                    ws.cell(row=row, column=2, value=adc.sample_id)
                    ws.cell(row=row, column=3, value=adc.description)
                    ws.cell(row=row, column=4, value=adc.concentration)
                    ws.cell(row=row, column=5, value=adc.owner)
                    ws.cell(row=row, column=6, value=adc.storage_temp)
                    ws.cell(row=row, column=7, value=adc.storage_position)
                    ws.cell(row=row, column=8, value=created_at_str)
                    ws.cell(row=row, column=12, value=0)
                    
                    for col in range(1, 13):
                        cell = ws.cell(row=row, column=col)
                        cell.fill = row_fill
                        cell.border = thin_border
                
                # 合并单元格（如果有多个规格）
                if spec_count > 1:
                    end_row = start_row + spec_count - 1
                    for col in [1, 2, 3, 4, 5, 6, 7, 8, 12]:
                        ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                        # 设置合并后的对齐方式
                        ws.cell(row=start_row, column=col).alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center", wrap_text=True)
                
                current_row += spec_count
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存文件
            wb.save(file_path)
            
            QMessageBox.information(self, "成功", f"已成功导出到: {file_path}")
        
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出失败: {str(e)}")
    
    # ==================== ADC出入库相关方法 ====================
    
    def refresh_adc_movements(self):
        """刷新出入库记录列表"""
        movements = self.adc_controller.get_all_movements()
        self._update_movement_table(movements)
        # 更新日期筛选器的默认范围
        self._update_movement_date_range(movements)
    
    def _update_movement_date_range(self, movements: List[Dict]):
        """根据出入库记录更新日期筛选器的范围"""
        if not movements:
            # 如果没有记录，使用默认范围
            self.movement_date_from.setDate(QDate(2000, 1, 1))
            self.movement_date_to.setDate(QDate(2099, 12, 31))
            return
        
        min_date = None
        max_date = None
        
        for m in movements:
            if m['date']:
                m_date = None
                if isinstance(m['date'], datetime):
                    m_date = QDate(m['date'].year, m['date'].month, m['date'].day)
                elif isinstance(m['date'], str):
                    try:
                        dt = datetime.strptime(m['date'], '%Y-%m-%d %H:%M:%S.%f')
                        m_date = QDate(dt.year, dt.month, dt.day)
                    except ValueError:
                        try:
                            dt = datetime.strptime(m['date'], '%Y-%m-%d %H:%M:%S')
                            m_date = QDate(dt.year, dt.month, dt.day)
                        except ValueError:
                            continue
                
                if m_date:
                    if min_date is None or m_date < min_date:
                        min_date = m_date
                    if max_date is None or m_date > max_date:
                        max_date = m_date
        
        # 设置日期范围
        if min_date and max_date:
            self.movement_date_from.setDate(min_date)
            self.movement_date_to.setDate(max_date)
        else:
            self.movement_date_from.setDate(QDate(2000, 1, 1))
            self.movement_date_to.setDate(QDate(2099, 12, 31))
    
    def _update_movement_table(self, movements: List[Dict]):
        """更新出入库记录表格"""
        # 保存movements列表以便选中时获取详细信息
        self._current_movements = movements
        self.movement_table.setRowCount(len(movements))
        
        for row, movement in enumerate(movements):
            # 类型
            type_text = "入库" if movement['type'] == 'inbound' else "出库"
            type_item = QTableWidgetItem(type_text)
            if movement['type'] == 'inbound':
                type_item.setBackground(QColor("#d4edda"))
            else:
                type_item.setBackground(QColor("#f8d7da"))
            self.movement_table.setItem(row, 0, type_item)
            
            # Lot Number
            self.movement_table.setItem(row, 1, QTableWidgetItem(movement['lot_number']))
            
            # 操作人
            self.movement_table.setItem(row, 2, QTableWidgetItem(movement['operator']))
            
            # 日期（精确到秒）
            date_str = ""
            if movement['date']:
                if isinstance(movement['date'], datetime):
                    date_str = movement['date'].strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(movement['date'], str):
                    # 尝试解析字符串格式的日期
                    try:
                        dt = datetime.strptime(movement['date'], '%Y-%m-%d %H:%M:%S.%f')
                        date_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        date_str = str(movement['date'])
                else:
                    date_str = str(movement['date'])
            self.movement_table.setItem(row, 3, QTableWidgetItem(date_str))
            
            # 明细
            items = movement['items']
            items_str = ", ".join([
                f"{item.spec_mg}mg×{item.quantity}" if isinstance(item, ADCMovementItem) 
                else f"{item.get('spec_mg', 0)}mg×{item.get('quantity', 0)}"
                for item in items
            ])
            self.movement_table.setItem(row, 4, QTableWidgetItem(items_str))
            
            # 合计
            total_mg = sum([
                item.spec_mg * item.quantity if isinstance(item, ADCMovementItem)
                else item.get('spec_mg', 0) * item.get('quantity', 0)
                for item in items
            ])
            self.movement_table.setItem(row, 5, QTableWidgetItem(f"{total_mg:.2f}"))
            
            # 备注
            record = movement['record']
            notes = record.notes if hasattr(record, 'notes') else ""
            self.movement_table.setItem(row, 6, QTableWidgetItem(notes))
    
    def search_adc_movements(self):
        """搜索出入库记录（支持多条件筛选）"""
        # 获取所有记录
        lot_keyword = self.movement_search_edit.text().strip()
        if lot_keyword:
            movements = self.adc_controller.search_movements_by_lot_number(lot_keyword)
        else:
            movements = self.adc_controller.get_all_movements()
        
        # 按类型筛选
        type_filter = self.movement_type_combo.currentText()
        if type_filter == "入库":
            movements = [m for m in movements if m['type'] == 'inbound']
        elif type_filter == "出库":
            movements = [m for m in movements if m['type'] == 'outbound']
        
        # 按操作人筛选
        operator_keyword = self.movement_operator_edit.text().strip()
        if operator_keyword:
            movements = [m for m in movements if operator_keyword.lower() in m['operator'].lower()]
        
        # 按日期范围筛选
        date_from = self.movement_date_from.date()
        date_to = self.movement_date_to.date()
        
        filtered_movements = []
        for m in movements:
            if m['date']:
                # 解析日期
                m_date = None
                if isinstance(m['date'], datetime):
                    m_date = QDate(m['date'].year, m['date'].month, m['date'].day)
                elif isinstance(m['date'], str):
                    try:
                        dt = datetime.strptime(m['date'], '%Y-%m-%d %H:%M:%S.%f')
                        m_date = QDate(dt.year, dt.month, dt.day)
                    except ValueError:
                        try:
                            dt = datetime.strptime(m['date'], '%Y-%m-%d %H:%M:%S')
                            m_date = QDate(dt.year, dt.month, dt.day)
                        except ValueError:
                            pass
                
                if m_date:
                    # 检查日期范围
                    if m_date >= date_from and m_date <= date_to:
                        filtered_movements.append(m)
                else:
                    # 无法解析日期的记录保留
                    filtered_movements.append(m)
            else:
                # 没有日期的记录保留
                filtered_movements.append(m)
        
        self._update_movement_table(filtered_movements)
    
    def clear_movement_search(self):
        """清除出入库搜索条件"""
        # 暂时断开信号，避免多次触发搜索
        self.movement_type_combo.blockSignals(True)
        self.movement_search_edit.blockSignals(True)
        self.movement_operator_edit.blockSignals(True)
        self.movement_date_from.blockSignals(True)
        self.movement_date_to.blockSignals(True)
        
        self.movement_type_combo.setCurrentIndex(0)  # 全部
        self.movement_search_edit.clear()
        self.movement_operator_edit.clear()
        
        # 重新连接信号
        self.movement_type_combo.blockSignals(False)
        self.movement_search_edit.blockSignals(False)
        self.movement_operator_edit.blockSignals(False)
        self.movement_date_from.blockSignals(False)
        self.movement_date_to.blockSignals(False)
        
        # 刷新数据（会自动更新日期范围为数据范围）
        self.refresh_adc_movements()
    
    def adc_inbound(self):
        """ADC入库"""
        dialog = ADCInboundDialog(self, self.adc_controller)
        if dialog.exec_() == QDialog.Accepted:
            inbound = dialog.result
            if inbound:
                try:
                    success, result = self.adc_controller.create_inbound(inbound)
                    if success:
                        QMessageBox.information(self, "成功", "入库成功")
                        self.refresh_adc_movements()
                        self.refresh_adcs()  # 刷新ADC库存
                    else:
                        QMessageBox.critical(self, "错误", result)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"入库失败: {str(e)}")
    
    def adc_outbound(self):
        """ADC出库"""
        dialog = ADCOutboundDialog(self, self.adc_controller)
        if dialog.exec_() == QDialog.Accepted:
            outbound = dialog.result
            if outbound:
                try:
                    success, result = self.adc_controller.create_outbound(outbound)
                    if success:
                        QMessageBox.information(self, "成功", "出库成功")
                        self.refresh_adc_movements()
                        self.refresh_adcs()  # 刷新ADC库存
                    else:
                        QMessageBox.critical(self, "错误", result)
                except Exception as e:
                    QMessageBox.critical(self, "错误", f"出库失败: {str(e)}")


def main():
    """主函数"""
    app = QApplication(sys.argv)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
